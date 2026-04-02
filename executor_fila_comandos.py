import json
import logging
import os
import subprocess
import sys
import time
from datetime import datetime
from pathlib import Path

import gspread
from google.oauth2.service_account import Credentials
import openpyxl

ROOT = Path(__file__).resolve().parent
CONFIG_PATH = ROOT / "config.json"
CMD_FILE = ROOT / "comando_imprimir.txt"
ROBO_SCRIPT = ROOT / "robo.py"
LOG_DIR = ROOT / "logs"
LOG_FILE = LOG_DIR / "executor_fila_comandos.log"

SHEET_NAME = os.getenv("EXECUTOR_SHEET_NAME", "COMANDOS")
POLL_SECONDS = max(2, int(os.getenv("EXECUTOR_POLL_SECONDS", "8")))
BATCH_SIZE = max(1, int(os.getenv("EXECUTOR_BATCH_SIZE", "25")))
ALLOW_LOCAL = str(os.getenv("EXECUTOR_ALLOW_LOCAL", "false")).strip().lower() in {"1", "true", "yes", "sim"}
AUTO_SYNC_LOCAL_EXCEL = str(os.getenv("EXECUTOR_AUTO_SYNC_LOCAL_EXCEL", "true")).strip().lower() in {"1", "true", "yes", "sim"}
AUTO_SYNC_MIN_INTERVAL_SECONDS = max(15, int(os.getenv("EXECUTOR_AUTO_SYNC_MIN_INTERVAL_SECONDS", "45")))

STATUS_SKIP = {
    "OK",
    "PROCESSANDO",
    "PROCESSANDO_EXTERNO",
    "IGNORADO",
    "IGNORADO_EXTERNO",
}

STATUS_ALLOW = {
    "PENDENTE",
    "ENCAMINHADO_EXTERNO",
    "ERRO",
}

DIRECT_CMD_PREFIXES = (
    "IMPRIMIR_FECHAMENTO:",
    "IMPRIMIR:",
    "IMPRIMIR_COMANDA:",
    "IMPRIMIR_GARANTIA:",
    "IMPRIMIR_CANCELADAS",
    "IMPRIMIR_RETIRADAS",
    "ENVIAR_WHATSAPP:",
)

DIRECT_CMDS = {
    "RECARREGAR_CONFIG",
    "VERIFICAR_HISTORICO",
    "ATUALIZAR_ESTOQUE",
    "FECHAMENTO_MANUAL",
}

SAFE_EXACT_CMDS = {
    "INICIAR_SISTEMA",
    "PARAR_SISTEMA",
    "IMPRIMIR_TICKETS",
    "SINCRONIZAR_EXCEL_LOCAL",
    "RECARREGAR_CONFIG",
    "VERIFICAR_HISTORICO",
    "ATUALIZAR_ESTOQUE",
    "FECHAMENTO_MANUAL",
    "GERAR_RELATORIO",
    "PROCESSAR_ALERTAS_ATRASO",
    "PROCESSAR_FILA_COMANDOS",
}

SAFE_PREFIXES = (
    "IMPRIMIR_FECHAMENTO:",
    "IMPRIMIR:",
    "IMPRIMIR_COMANDA:",
    "IMPRIMIR_GARANTIA:",
    "IMPRIMIR_CANCELADAS",
    "IMPRIMIR_RETIRADAS",
    "ENVIAR_WHATSAPP:",
    "TELEGRAM:",
)

CONFIG_SYNC_KEYS = (
    "pix_motoboys",
    "motoboys",
    "bairros",
)


def setup_logging() -> None:
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(levelname)s | %(message)s",
        handlers=[
            logging.FileHandler(LOG_FILE, encoding="utf-8"),
            logging.StreamHandler(sys.stdout),
        ],
    )


def load_config() -> dict:
    if not CONFIG_PATH.exists():
        raise FileNotFoundError(f"config.json não encontrado em {CONFIG_PATH}")
    with CONFIG_PATH.open("r", encoding="utf-8") as f:
        return json.load(f)


def resolve_sheets_credentials(config: dict) -> tuple[str, Path]:
    gs = config.get("google_sheets", {}) if isinstance(config, dict) else {}
    sheet_id = str(gs.get("sheets_id") or "").strip()
    if not sheet_id:
        raise RuntimeError("google_sheets.sheets_id não configurado no config.json")

    candidates = []
    env_json = os.getenv("GOOGLE_APPLICATION_CREDENTIALS", "").strip()
    if env_json:
        candidates.append(Path(env_json))

    cfg_json = str(gs.get("service_account_json") or "").strip()
    if cfg_json:
        candidates.append(Path(cfg_json))

    for p in ROOT.glob("*.json"):
        if p.name.startswith("gen-lang-client"):
            candidates.append(p)

    for path in candidates:
        if path.exists() and path.is_file():
            return sheet_id, path

    raise RuntimeError(
        "Credencial de service account não encontrada. Defina google_sheets.service_account_json "
        "ou GOOGLE_APPLICATION_CREDENTIALS."
    )


def open_sheet(sheet_id: str, credentials_path: Path):
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]
    creds = Credentials.from_service_account_file(str(credentials_path), scopes=scopes)
    gc = gspread.authorize(creds)
    sh = gc.open_by_key(sheet_id)
    return sh.worksheet(SHEET_NAME)


def open_spreadsheet(sheet_id: str, credentials_path: Path):
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]
    creds = Credentials.from_service_account_file(str(credentials_path), scopes=scopes)
    gc = gspread.authorize(creds)
    return gc.open_by_key(sheet_id)


def _safe_obj(value) -> dict:
    return value if isinstance(value, dict) else {}


def sync_local_config_maps_to_sheets(local_cfg: dict | None = None) -> str:
    cfg = local_cfg if isinstance(local_cfg, dict) else load_config()
    sheet_id, cred_path = resolve_sheets_credentials(cfg)

    sh = open_spreadsheet(sheet_id, cred_path)
    try:
        ws_cfg = sh.worksheet("CONFIG")
    except gspread.WorksheetNotFound:
        ws_cfg = sh.add_worksheet(title="CONFIG", rows=200, cols=2)

    values = ws_cfg.get_all_values()
    if not values:
        ws_cfg.append_row(["Chave", "Valor"])
        values = [["Chave", "Valor"]]

    row_by_key = {}
    for idx, row in enumerate(values[1:], start=2):
        key = str(row[0]).strip() if row else ""
        if key:
            row_by_key[key] = idx

    synced_parts = []
    for key in CONFIG_SYNC_KEYS:
        payload = _safe_obj(cfg.get(key))
        payload_text = json.dumps(payload, ensure_ascii=False)

        row = row_by_key.get(key)
        if row:
            ws_cfg.update(f"A{row}:B{row}", [[key, payload_text]])
        else:
            ws_cfg.append_row([key, payload_text])

        synced_parts.append(f"{key}={len(payload)}")

    return "Config sincronizada em CONFIG: " + ", ".join(synced_parts)


def read_rows(ws):
    records = ws.get_all_values()
    if not records:
        return [], {}

    header = records[0]
    idx = {name: i for i, name in enumerate(header)}
    required = ["Comando", "Status", "Tentativas", "Resultado", "AtualizadoEm"]
    missing = [c for c in required if c not in idx]
    if missing:
        raise RuntimeError(f"Colunas ausentes em {SHEET_NAME}: {', '.join(missing)}")

    rows = []
    for r, values in enumerate(records[1:], start=2):
        row = {"_sheet_row": r}
        for k, i in idx.items():
            row[k] = values[i] if i < len(values) else ""
        rows.append(row)
    return rows, idx


def update_status(ws, idx, sheet_row: int, status: str, tentativas: int, resultado: str) -> None:
    data = [
        [status, tentativas, resultado[:500], time.strftime("%Y-%m-%d %H:%M:%S")],
    ]
    c_status = idx["Status"] + 1
    ws.update(f"{col_letter(c_status)}{sheet_row}:{col_letter(c_status+3)}{sheet_row}", data)


def col_letter(col_number: int) -> str:
    result = ""
    n = col_number
    while n:
        n, rem = divmod(n - 1, 26)
        result = chr(65 + rem) + result
    return result


def write_command_file(command: str) -> str:
    CMD_FILE.write_text(command, encoding="utf-8")
    return "Comando enviado ao robo.py via arquivo"


def resolve_python_executable() -> str:
    venv1 = ROOT / ".venv-1" / "Scripts" / "python.exe"
    if venv1.exists():
        return str(venv1)

    venv = ROOT / ".venv" / "Scripts" / "python.exe"
    if venv.exists():
        return str(venv)

    return sys.executable


def is_robo_running() -> bool:
    try:
        output = subprocess.check_output(
            [
                "powershell",
                "-NoProfile",
                "-Command",
                "Get-CimInstance Win32_Process | Where-Object { ($_.Name -match 'python') -and ($_.CommandLine -match 'robo.py') } | Select-Object -First 1 -ExpandProperty ProcessId",
            ],
            text=True,
            stderr=subprocess.STDOUT,
        )
        return bool(str(output or "").strip())
    except Exception:
        return False


def start_robo_process() -> str:
    if not ROBO_SCRIPT.exists():
        raise FileNotFoundError("robo.py não encontrado")

    if is_robo_running():
        return "robo.py já está em execução"

    pyexe = resolve_python_executable()
    subprocess.Popen(["cmd", "/c", "start", "", pyexe, str(ROBO_SCRIPT)], cwd=str(ROOT), shell=False)
    return f"robo.py iniciado com {pyexe}"


def start_system() -> str:
    robo_msg = start_robo_process()

    bot_bat = ROOT / "start_telegram_bot.bat"
    if bot_bat.exists():
        subprocess.Popen(["cmd", "/c", "start", "", str(bot_bat)], cwd=str(ROOT), shell=False)
        return "Início do sistema disparado. " + robo_msg + " + telegram bot"

    return "Início do sistema disparado. " + robo_msg


def stop_system() -> str:
    bat = ROOT / "parar_tudo.bat"
    if not bat.exists():
        raise FileNotFoundError("parar_tudo.bat não encontrado")
    subprocess.Popen(["cmd", "/c", str(bat)], cwd=str(ROOT), shell=False)
    return "Parada do sistema disparada"


def extract_payload(raw_payload: str) -> dict:
    txt = (raw_payload or "").strip()
    if not txt:
        return {}
    try:
        obj = json.loads(txt)
        return obj if isinstance(obj, dict) else {"value": obj}
    except Exception:
        return {"raw": txt}


def is_command_allowed(command: str) -> bool:
    cmd = (command or "").strip()
    if not cmd:
        return False

    up = cmd.upper()

    if up in SAFE_EXACT_CMDS:
        return True

    if up.startswith(SAFE_PREFIXES):
        return True

    if up.startswith("LOCAL:"):
        return ALLOW_LOCAL

    return False


def resolve_external_command(command: str, payload: dict) -> tuple[str, str]:
    cmd = (command or "").strip()
    up = cmd.upper()

    if not is_command_allowed(cmd):
        return "IGNORED", "Comando bloqueado pela whitelist do executor"

    if up == "INICIAR_SISTEMA":
        return "LOCAL_START", start_system()

    if up == "PARAR_SISTEMA":
        return "LOCAL_STOP", stop_system()

    if up == "IMPRIMIR_TICKETS":
        forwarded = payload.get("command") or payload.get("comando") or "IMPRIMIR:RETIRADA"
        return "FORWARDED", write_command_file(str(forwarded))

    if up == "SINCRONIZAR_EXCEL_LOCAL":
        arquivo = str(payload.get("arquivo") or payload.get("file") or payload.get("path") or "").strip()
        data = str(payload.get("data") or "").strip()
        cfg_local = load_config()
        msg_extrato = sync_local_excel_to_sheets(
            ws_name="EXTRATO DETALHADO",
            arquivo=arquivo,
            data=data,
            source_sheet_name="EXTRATO DETALHADO",
        )
        msg_vales = sync_local_excel_to_sheets(
            ws_name="VALES",
            arquivo=arquivo,
            data=data,
            source_sheet_name="VALES",
            optional_if_missing=True,
        )
        msg_cfg = sync_local_config_maps_to_sheets(cfg_local)
        return "SYNC_LOCAL", f"{msg_extrato} | {msg_vales} | {msg_cfg}"

    if up in DIRECT_CMDS or up.startswith(DIRECT_CMD_PREFIXES):
        return "FORWARDED", write_command_file(cmd)

    if up.startswith("LOCAL:"):
        if not ALLOW_LOCAL:
            return "IGNORED", "Comando LOCAL bloqueado (EXECUTOR_ALLOW_LOCAL=false)"
        raw = cmd.split(":", 1)[1].strip()
        if not raw:
            return "IGNORED", "Comando LOCAL vazio"
        subprocess.Popen(raw, cwd=str(ROOT), shell=True)
        return "LOCAL_SHELL", "Comando LOCAL executado"

    return "IGNORED", "Comando não mapeado para executor externo"


def _resolver_arquivo_excel_local(arquivo: str = "", data: str = "") -> Path:
    if arquivo:
        p = Path(arquivo)
        if not p.is_absolute():
            p = ROOT / p
        if p.exists() and p.is_file():
            return p
        raise FileNotFoundError(f"Arquivo informado não encontrado: {p}")

    if data:
        nome = f"Controle_Financeiro_{data}.xlsx"
        p = ROOT / nome
        if p.exists() and p.is_file():
            return p
        raise FileNotFoundError(f"Arquivo da data {data} não encontrado: {p}")

    candidatos = sorted(ROOT.glob("Controle_Financeiro_*.xlsx"), key=lambda x: x.stat().st_mtime, reverse=True)
    if not candidatos:
        raise FileNotFoundError("Nenhum arquivo Controle_Financeiro_*.xlsx encontrado no diretório do projeto")
    return candidatos[0]


def sync_local_excel_to_sheets(
    ws_name: str = "EXTRATO DETALHADO",
    arquivo: str = "",
    data: str = "",
    source_sheet_name: str = "",
    optional_if_missing: bool = False,
) -> str:
    cfg = load_config()
    sheet_id, cred_path = resolve_sheets_credentials(cfg)

    excel_path = _resolver_arquivo_excel_local(arquivo=arquivo, data=data)
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    source_sheet = str(source_sheet_name or ws_name).strip() or "EXTRATO DETALHADO"
    if source_sheet not in wb.sheetnames:
        if optional_if_missing:
            return f"Aba '{source_sheet}' ausente em {excel_path.name}; sincronização ignorada"
        raise RuntimeError(f"Aba '{source_sheet}' não encontrada no Excel local")

    ws_xlsx = wb[source_sheet]
    valores = []
    for row in ws_xlsx.iter_rows(values_only=True):
        linha = list(row)
        if not any(v not in (None, "") for v in linha):
            continue
        valores.append(["" if v is None else v for v in linha])

    if not valores:
        raise RuntimeError("Excel local sem dados para sincronizar")

    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]
    creds = Credentials.from_service_account_file(str(cred_path), scopes=scopes)
    gc = gspread.authorize(creds)
    sh = gc.open_by_key(sheet_id)
    try:
        ws_dest = sh.worksheet(ws_name)
    except gspread.WorksheetNotFound:
        ws_dest = sh.add_worksheet(title=ws_name, rows=max(1000, len(valores) + 50), cols=max(20, len(valores[0]) + 5))

    ws_dest.clear()
    ws_dest.update(valores, "A1")

    return f"Sincronizado {len(valores)-1} registro(s) de {excel_path.name} ({source_sheet} -> {ws_name})"


def auto_sync_local_excel_if_needed(
    last_synced_path: str,
    last_synced_mtime: float,
    last_sync_ts: float,
) -> tuple[str, float, float, bool]:
    if not AUTO_SYNC_LOCAL_EXCEL:
        return last_synced_path, last_synced_mtime, last_sync_ts, False

    agora = time.time()
    if last_sync_ts and (agora - last_sync_ts) < AUTO_SYNC_MIN_INTERVAL_SECONDS:
        return last_synced_path, last_synced_mtime, last_sync_ts, False

    try:
        excel_path = _resolver_arquivo_excel_local()
    except Exception:
        return last_synced_path, last_synced_mtime, last_sync_ts, False

    try:
        mtime = excel_path.stat().st_mtime
    except Exception:
        return last_synced_path, last_synced_mtime, last_sync_ts, False

    mudou = (str(excel_path) != str(last_synced_path)) or (float(mtime) != float(last_synced_mtime))
    if not mudou:
        return last_synced_path, last_synced_mtime, last_sync_ts, False

    cfg_local = load_config()

    msg_extrato = sync_local_excel_to_sheets(
        ws_name="EXTRATO DETALHADO",
        arquivo=str(excel_path),
        data="",
        source_sheet_name="EXTRATO DETALHADO",
    )
    msg_vales = sync_local_excel_to_sheets(
        ws_name="VALES",
        arquivo=str(excel_path),
        data="",
        source_sheet_name="VALES",
        optional_if_missing=True,
    )
    msg_cfg = sync_local_config_maps_to_sheets(cfg_local)

    logging.info("Auto-sync Excel local: %s | %s | %s", msg_extrato, msg_vales, msg_cfg)
    return str(excel_path), float(mtime), agora, True


def process_batch(ws) -> tuple[int, int, int]:
    rows, idx = read_rows(ws)
    processed = 0
    errors = 0
    ignored = 0

    for row in rows:
        if processed >= BATCH_SIZE:
            break

        status = str(row.get("Status", "") or "PENDENTE").strip().upper()
        if status in STATUS_SKIP or status not in STATUS_ALLOW:
            continue

        cmd = str(row.get("Comando", "") or "").strip()
        if not cmd:
            ignored += 1
            continue

        tentativas = int(float(str(row.get("Tentativas", "0") or "0"))) + 1
        sheet_row = int(row["_sheet_row"])
        update_status(ws, idx, sheet_row, "PROCESSANDO_EXTERNO", tentativas, "")

        try:
            payload = extract_payload(str(row.get("Payload", "") or ""))
            action, result = resolve_external_command(cmd, payload)

            if action == "IGNORED":
                update_status(ws, idx, sheet_row, "IGNORADO_EXTERNO", tentativas, result)
                ignored += 1
                continue

            update_status(ws, idx, sheet_row, "OK", tentativas, f"{action}: {result}")
            processed += 1
        except Exception as exc:
            errors += 1
            update_status(ws, idx, sheet_row, "ERRO", tentativas, str(exc))
            logging.exception("Erro processando comando externo: %s", cmd)

    return processed, errors, ignored


def main() -> None:
    setup_logging()
    logging.info(
        "Executor externo da fila iniciado. Poll=%ss Batch=%s AllowLocal=%s",
        POLL_SECONDS,
        BATCH_SIZE,
        ALLOW_LOCAL,
    )
    logging.info(
        "AutoSyncLocalExcel=%s IntervaloMin=%ss",
        AUTO_SYNC_LOCAL_EXCEL,
        AUTO_SYNC_MIN_INTERVAL_SECONDS,
    )

    cfg = load_config()
    sheet_id, cred_path = resolve_sheets_credentials(cfg)
    logging.info("Conectando na planilha %s usando credencial %s", sheet_id, cred_path)

    last_synced_path = ""
    last_synced_mtime = 0.0
    last_sync_ts = 0.0

    while True:
        try:
            ws = open_sheet(sheet_id, cred_path)
            p, e, i = process_batch(ws)
            if p or e or i:
                logging.info("Ciclo fila: processados=%s erros=%s ignorados=%s", p, e, i)

            last_synced_path, last_synced_mtime, last_sync_ts, synced = auto_sync_local_excel_if_needed(
                last_synced_path,
                last_synced_mtime,
                last_sync_ts,
            )
            if synced:
                logging.info("Auto-sync aplicado para arquivo: %s", last_synced_path)
        except Exception:
            logging.exception("Falha no ciclo do executor externo")
        time.sleep(POLL_SECONDS)


if __name__ == "__main__":
    main()
