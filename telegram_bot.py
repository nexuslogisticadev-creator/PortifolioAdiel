# ==================================================================================
# INTEGRAÇÃO E SINCRONIA ENTRE SISTEMAS
# ----------------------------------------------------------------------------------
# Este arquivo implementa o BOT TELEGRAM (telegram_bot.py), responsável por receber
# comandos remotos e repassá-los ao robô principal (robo.py) via 'telegram_command.txt'.
# O painel (painel.py) também pode enviar comandos ao robô via 'comando_imprimir.txt'.
# Todos os comandos disponíveis devem ser mantidos sincronizados entre os três sistemas.
# O bot responde ao usuário no Telegram e registra comandos para execução pelo robô.
# Consulte os comentários de integração ao longo do código para pontos de comunicação.
# ==================================================================================
# -*- coding: utf-8 -*-
import os
import json
import subprocess
import psutil
import time
import sys
import asyncio
import threading
# Configurar encoding UTF-8 para stdout (suportar emojis no Windows)
if sys.platform == "win32":
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import atexit
import re
import requests
import logging
import openpyxl
import pandas as pd
import numpy as np
from datetime import datetime, timedelta, time as datetime_time
from telegram import Update, Bot, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    Application,
    CommandHandler,
    CallbackContext,
    MessageHandler,
    filters,
    ConversationHandler,
    CallbackQueryHandler,
)

# --- BIBLIOTECAS DE MANIPULAÇÃO DE JANELAS ---
try:
    import win32gui
    import win32con
    TEM_WIN32 = True
except ImportError:
    TEM_WIN32 = False

# --- BIBLIOTECAS DE IMPRESSÃO (WINDOWS) ---
try:
    import win32print
    import win32api
    TEM_IMPRESSORA = True
except ImportError:
    TEM_IMPRESSORA = False

# --- BIBLIOTECAS GOOGLE SHEETS ---
try:
    import gspread
    from google.oauth2.service_account import Credentials
    from gspread.exceptions import WorksheetNotFound
    TEM_GSHEETS = True
except ImportError:
    pass
    TEM_GSHEETS = False
    print("❌ ERRO: gspread nao instalado. Impossível ler token do Sheets!")

# ==================================================================================
# IMPORTAR UTILITÁRIOS COMPARTILHADOS
# ==================================================================================
from utils import (
    get_caminho_base, get_caminho_excel, carregar_config, salvar_config,
    get_data_operacional, normalizar_texto, arquivo_existe, ler_arquivo,
    escrever_arquivo, ARQUIVO_CONFIG, ARQUIVO_ROBO, ARQUIVO_COMANDO,
    ARQUIVO_COMANDO_TELEGRAM
)

# ==================================================================================
# VARIÁVEIS GLOBAIS DO BOT
# ==================================================================================
TELEGRAM_TOKEN = None
ADMIN_CHAT_ID = None

# ==================================================================================
# CONSTANTES DE IMPRESSÃO TÉRMICA
# ==================================================================================
CMD_INIT = b"\x1b\x40"
CMD_CENTER = b"\x1b\x61\x01"
CMD_LEFT = b"\x1b\x61\x00"
CMD_BOLD_ON = b"\x1b\x45\x01"
CMD_BOLD_OFF = b"\x1b\x45\x00"
CMD_DOUBLE_H = b"\x1b\x21\x10"
CMD_NORMAL = b"\x1b\x21\x00"
CMD_CUT = b"\x1d\x56\x00"
CMD_TIGHT_SPACING = b"\x1b\x33\x14"
CMD_NORMAL_SPACING = b"\x1b\x32"

# ==================================================================================
# STATUS CANCELADOS (para filtrar pedidos)
# ==================================================================================
STATUS_CANCELADOS_LISTA = ["CANCELADA", "CANCELADO", "DEVOLVIDO", "DEVOLUÇÃO", "RETORNADO"]

# ==================================================================================
# LOGGING
# ==================================================================================
logging.basicConfig(format='%(asctime)s - %(name)s - %(levelname)s - %(message)s', level=logging.INFO)

# Suprimir aviso informativo do python-telegram-bot sobre per_message
logging.getLogger('telegram.ext._conversationhandler').setLevel(logging.ERROR)

# ==================================================================================
# FUNÇÕES DE IMPRESSÃO INDEPENDENTE (sem depender do robo.py)
# ==================================================================================

def imprimir_lote_independente(lista_pedidos):
    """Imprime lote de pedidos diretamente via impressora térmica (sem robo.py)."""
    if not TEM_IMPRESSORA:
        return False
    try:
        print(f"🖨️ Imprimindo lote de {len(lista_pedidos)} tickets...")
        impressora_padrao = win32print.GetDefaultPrinter()
        hPrinter = win32print.OpenPrinter(impressora_padrao)
        try:
            hJob = win32print.StartDocPrinter(hPrinter, 1, ("Lote_Tickets", None, "RAW"))
            try:
                win32print.StartPagePrinter(hPrinter)
                buffer_impressao = CMD_INIT + CMD_LEFT + CMD_BOLD_ON
                for i, dados in enumerate(lista_pedidos):
                    buffer_impressao += f"PED: {dados['numero']}  |  {dados['hora']}\n".encode('cp850', errors='ignore')
                    buffer_impressao += f"CLI: {dados['cliente'][:28]}\n".encode('cp850', errors='ignore')
                    buffer_impressao += f"BAI: {dados['bairro']}\n".encode('cp850', errors='ignore')
                    if dados.get('itens'):
                        itens_fmt = dados['itens'].replace("\n", " ").replace(" | ", " ")
                        buffer_impressao += f"ITM: {itens_fmt[:40]}\n".encode('cp850', errors='ignore')
                    buffer_impressao += f"$$$: R$ {dados['valor']:.2f}\n".replace('.', ',').encode('cp850')
                    if i < len(lista_pedidos) - 1:
                        buffer_impressao += b"________________________________\n"
                    else:
                        buffer_impressao += b"\n\n\n"
                buffer_impressao += CMD_CUT
                win32print.WritePrinter(hPrinter, buffer_impressao)
                win32print.EndPagePrinter(hPrinter)
            finally:
                win32print.EndDocPrinter(hPrinter)
        finally:
            win32print.ClosePrinter(hPrinter)
        print("✅ Lote impresso.")
        return True
    except Exception as e:
        print(f"❌ Erro ao imprimir lote: {e}")
        return False

def imprimir_resumo_independente(nome_motoboy, lista_pedidos, qtd_8, qtd_11, total_valor, vale_total=0.0):
    """Imprime resumo/extrato de fechamento com vales (sem robo.py)."""
    if not TEM_IMPRESSORA:
        return False
    try:
        print(f"🖨️ Imprimindo EXTRATO para {nome_motoboy}...")
        impressora_padrao = win32print.GetDefaultPrinter()
        hPrinter = win32print.OpenPrinter(impressora_padrao)
        data_print = datetime.now().strftime('%d/%m/%Y %H:%M')
        try:
            hJob = win32print.StartDocPrinter(hPrinter, 1, (f"Extrato_{nome_motoboy}", None, "RAW"))
            try:
                win32print.StartPagePrinter(hPrinter)
                cupom = CMD_INIT + CMD_CENTER
                cupom += CMD_BOLD_ON + CMD_DOUBLE_H + b"FECHAMENTO\n" + CMD_NORMAL + CMD_BOLD_OFF
                cupom += CMD_BOLD_ON + f"{nome_motoboy.upper()}\n".encode('cp850', errors='ignore')
                cupom += f"{data_print}\n".encode('cp850') + b"================================\n"
                cupom += CMD_LEFT + b"PEDIDO          VALOR\n" + CMD_BOLD_OFF
                for p in lista_pedidos:
                    id_ped = str(p['numero']).ljust(15)
                    valor_ped = f"R$ {p['valor']:.2f}".replace('.', ',')
                    cupom += f"{id_ped} {valor_ped}\n".encode('cp850')
                cupom += b"--------------------------------\n" + CMD_BOLD_ON
                cupom += f"QTD R$ 8,00:  {qtd_8}\n".encode('cp850')
                cupom += f"QTD R$ 11,00: {qtd_11}\n".encode('cp850')
                if vale_total and vale_total > 0:
                    cupom += b"--------------------------------\n"
                    cupom += f"(-) DESCONTO VALE: R$ {vale_total:.2f}\n".replace('.', ',').encode('cp850')
                cupom += b"--------------------------------\n" + CMD_CENTER + CMD_DOUBLE_H
                cupom += f"TOTAL: R$ {total_valor:.2f}\n".replace('.', ',').encode('cp850')
                cupom += CMD_NORMAL + CMD_BOLD_ON
                cupom += f"ENTREGAS: {len(lista_pedidos)}\n".encode('cp850') + CMD_BOLD_OFF
                cupom += b"\n\n\n" + CMD_CUT
                win32print.WritePrinter(hPrinter, cupom)
                win32print.EndPagePrinter(hPrinter)
            finally:
                win32print.EndDocPrinter(hPrinter)
        finally:
            win32print.ClosePrinter(hPrinter)
        print("✅ Extrato impresso.")
        return True
    except Exception as e:
        print(f"❌ Erro ao imprimir extrato: {e}")
        return False

def buscar_e_imprimir_extrato(nome_motoboy, data_personalizada=None):
    """Busca e imprime o extrato de um motoboy diretamente do Excel (sem robo.py)."""
    try:
        if not data_personalizada:
            data_personalizada = get_data_operacional()
        
        excel_path = get_caminho_excel()
        if not os.path.exists(excel_path):
            print(f"❌ Arquivo de controle não encontrado: {excel_path}")
            return False

        wb = openpyxl.load_workbook(excel_path, data_only=True)
        if "EXTRATO DETALHADO" not in wb.sheetnames:
            print("❌ Aba 'EXTRATO DETALHADO' não encontrada no Excel.")
            return False

        ws = wb["EXTRATO DETALHADO"]
        pedidos_completos = []
        qtd_8 = 0
        qtd_11 = 0
        total = 0.0

        print(f"🔎 Buscando pedidos de '{nome_motoboy}' no Excel...")
        nome_buscado_norm = normalizar_texto(nome_motoboy)

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 9 or not row[6]:
                continue
            
            motoboy_excel_original = str(row[6])
            motoboy_excel_norm = normalizar_texto(motoboy_excel_original)
            status = str(row[5]).upper()
            val = float(row[8]) if row[8] else 0.0

            eh_retirada_busca = "retirada" in nome_buscado_norm
            match = False

            if eh_retirada_busca:
                if (val == 0.0 or "RETIRADA" in motoboy_excel_original.upper()) and not any(x in status for x in STATUS_CANCELADOS_LISTA):
                    match = True
            elif nome_buscado_norm in motoboy_excel_norm:
                if not any(x in status for x in STATUS_CANCELADOS_LISTA):
                    match = True

            if match:
                total += val
                if abs(val - 8.0) < 0.1:
                    qtd_8 += 1
                elif abs(val - 11.0) < 0.1:
                    qtd_11 += 1

                data_ped = row[0]
                if isinstance(data_ped, datetime):
                    data_ped = data_ped.strftime('%d/%m')

                pedidos_completos.append({
                    'numero': row[2],
                    'data': str(data_ped),
                    'hora': str(row[1]),
                    'cliente': str(row[3]),
                    'bairro': str(row[4]),
                    'motoboy': str(row[6]),
                    'valor': val,
                    'itens': str(row[9]) if len(row) > 9 and row[9] else ""
                })

        # Buscar vales
        vale_total = 0.0
        if "retirada" not in nome_buscado_norm and "VALES" in wb.sheetnames:
            ws_vales = wb["VALES"]
            for row in ws_vales.iter_rows(min_row=2, values_only=True):
                if not row or len(row) < 3 or not row[1]:
                    continue
                if normalizar_texto(str(row[1])) == nome_buscado_norm:
                    try:
                        vale_total += float(row[2]) if row[2] else 0.0
                    except Exception:
                        pass

        print(f"🏁 Encontrados {len(pedidos_completos)} pedidos")
        if pedidos_completos:
            print("🖨️ Imprimindo detalhes (Lote)...")
            imprimir_lote_independente(pedidos_completos)
            time.sleep(2)
            
            print("🖨️ Imprimindo resumo final...")
            nome_final = "RETIRADAS" if "retirada" in nome_buscado_norm else nome_motoboy
            total_liquido = total - vale_total
            if total_liquido < 0:
                total_liquido = 0.0
            
            imprimir_resumo_independente(
                nome_final, pedidos_completos, qtd_8, qtd_11, total_liquido,
                vale_total=vale_total
            )
            print("✅ Impressão completa!")
            return True
        else:
            print(f"⚠️ Nenhum pedido encontrado para '{nome_motoboy}'.")
            return False

    except Exception as e:
        print(f"❌ Erro ao buscar/imprimir extrato: {e}")
        import traceback
        traceback.print_exc()
        return False

# ==================================================================================
# GERENCIAMENTO DE PROCESSOS DO ROBÔ
# ==================================================================================

def find_robo_process():
    """Encontra o processo do robo.py no sistema."""
    target_script = ARQUIVO_ROBO.lower()
    for proc in psutil.process_iter(['pid', 'name', 'cmdline']):
        try:
            cmdline = proc.info.get('cmdline')
            if not cmdline: continue
            
            # Verifica se é um processo Python
            is_python = 'python' in proc.info.get('name', '').lower() or any('python' in arg.lower() for arg in cmdline)
            if is_python:
                # Verifica se o script alvo está nos argumentos (independente do caminho absoluto/relativo)
                if any(arg.lower().endswith(target_script) for arg in cmdline):
                    return proc
        except (psutil.NoSuchProcess, psutil.AccessDenied):
            continue
    return None

def start_robo():
    """Inicia o processo do robô se não estiver rodando."""
    if find_robo_process():
        return "⚠️ O robô já está em execução."
    
    caminho_python = sys.executable
    caminho_robo = os.path.abspath(os.path.join(get_caminho_base(), ARQUIVO_ROBO))
    comando = [caminho_python, caminho_robo, "--painel"]
    debug_info = (
        f"[DEBUG] Comando para iniciar robô: {comando}\n"
        f"[DEBUG] cwd: {get_caminho_base()}\n"
        f"[DEBUG] ARQUIVO_ROBO existe? {os.path.exists(caminho_robo)}\n"
    )
    print(debug_info)
    try:
        with open(os.path.join(get_caminho_base(), "debug_robo_start.txt"), "a", encoding="utf-8") as dbg:
            dbg.write(f"{datetime.datetime.now()}\n{debug_info}\n")
    except Exception:
        pass
    try:
        processo = subprocess.Popen(comando, cwd=get_caminho_base())
        
        # Verifica se o processo continua vivo após o início
        for i in range(20):
            time.sleep(1)
            if processo.poll() is not None:
                print("❌ O robô fechou inesperadamente logo após iniciar.")
                return "❌ O robô fechou sozinho. Verifique o 'robo.log' para ver o erro do Chrome."
            
            proc_detectado = find_robo_process()
            if proc_detectado:
                # Se o processo existe e o Chrome costuma demorar, esperamos ele estabilizar
                if i > 5: 
                    print(f"✅ Robô estabilizado (tentativa {i+1}).")
                    return "✅ Robô iniciado e Chrome carregado!"
        
        print("❌ Falha ao verificar o processo do robô após a inicialização.")
        return "❌ O robô iniciou, mas fechou em seguida. Verifique o arquivo 'robo.log' para ver o erro."
    except Exception as e:
        print(f"❌ Erro ao iniciar o robô: {e}")
        return f"❌ Erro ao iniciar o robô: {e}"

def stop_robo():
    """Para o processo do robô se estiver rodando."""
    processo = find_robo_process()
    if not processo:
        return "⚠️ O robô já está offline."
        
    print(f"⏹️ Parando o processo do robô (PID: {processo.pid})...")
    try:
        processo.terminate()
        processo.wait(timeout=5)
        print("✅ Robô parado com sucesso.")
        return "✅ Robô parado com sucesso!"
    except psutil.TimeoutExpired:
        print(f"⚠️ Processo {processo.pid} não terminou a tempo. Forçando parada...")
        processo.kill()
        print("✅ Robô forçadamente parado.")
        return "✅ Robô forçadamente parado."
    except Exception as e:
        print(f"❌ Erro ao parar o robô: {e}")
        return f"❌ Erro ao parar o robô: {e}"

def control_chrome_window(action: str):
    """Localiza e altera a visibilidade das janelas do Chrome (Modo Stealth)."""
    if not TEM_WIN32:
        return "❌ Erro: Biblioteca 'pywin32' não instalada no servidor."

    alvos = []
    try:
        def enum_handler(hwnd, results):
            title = win32gui.GetWindowText(hwnd)
            # Lógica idêntica ao painel.py
            if ("Google Chrome" in title or "Chrome" in title) and "DELIVERY" not in title:
                if action == "show" or win32gui.IsWindowVisible(hwnd):
                    results.append(hwnd)

        win32gui.EnumWindows(enum_handler, alvos)
        if not alvos:
            return "⚠️ Nenhuma janela do Chrome encontrada."

        for hwnd in alvos:
            if action == "hide":
                win32gui.ShowWindow(hwnd, win32con.SW_HIDE)
            else:
                win32gui.ShowWindow(hwnd, win32con.SW_SHOW)
                win32gui.ShowWindow(hwnd, win32con.SW_RESTORE)
        return f"✅ Chrome {'visível' if action == 'show' else 'ocultado'} ({len(alvos)} janelas)."
    except Exception as e:
        return f"❌ Erro ao controlar janelas: {e}"

def get_robo_status():
    """Verifica se o robô está online ou offline."""
    if find_robo_process():
        return "🟢 *Online*"
    else:
        return "🔴 *Offline*"

def send_panel_command(command: str):
    """Envia um comando para o robo.py através de um arquivo."""
    # INTEGRAÇÃO: Esta função grava o comando no arquivo 'telegram_command.txt',
    # que será lido e processado pelo robo.py. Mantenha os comandos sincronizados
    # entre este arquivo, robo.py e painel.py para garantir operação remota consistente.
    try:
        with open(ARQUIVO_COMANDO_TELEGRAM, 'w', encoding='utf-8') as f:
            f.write(command)
        return f"✅ Comando '{command}' enviado ao robô!"
    except Exception as e:
        print(f"❌ Erro ao enviar comando via arquivo: {e}")
        return f"❌ Erro ao enviar comando: {e}"

# ==================================================================================
# FILTROS
# ==================================================================================

# ==================================================================================
# HANDLERS DE COMANDOS
# ==================================================================================

async def start_command(update: Update, context: CallbackContext):
    """Handler para o comando /start."""
    message = start_robo()
    await update.message.reply_text(message)

async def stop_command(update: Update, context: CallbackContext):
    """Handler para o comando /stop."""
    message = stop_robo()
    await update.message.reply_text(message)

async def restart_command(update: Update, context: CallbackContext):
    """Handler para o comando /restart."""
    await update.message.reply_text("🔄 Reiniciando o robô...")
    
    # Para o robô e aguarda confirmação
    stop_result = stop_robo()
    
    # Aguarda até o processo realmente encerrar (máx 10 segundos)
    for _ in range(10):
        await asyncio.sleep(1)
        if not find_robo_process():
            break
    
    # Verifica se realmente parou
    if find_robo_process():
        await update.message.reply_text("❌ Não foi possível parar o robô para reiniciar.")
        return
    
    # Inicia novamente
    message = start_robo()
    await update.message.reply_text(message)

async def status_command(update: Update, context: CallbackContext):
    """Handler para o comando /status."""
    status = get_robo_status()
    await update.message.reply_text(f"Status do Robô: {status}", parse_mode='Markdown')

async def hide_chrome_command(update: Update, context: CallbackContext):
    """Handler para ocultar o Chrome."""
    message = control_chrome_window("hide")
    await update.message.reply_text(message)

async def show_chrome_command(update: Update, context: CallbackContext):
    """Handler para mostrar o Chrome."""
    message = control_chrome_window("show")
    await update.message.reply_text(message)

async def logs_command(update: Update, context: CallbackContext):
    """Handler para o comando /logs - Envia as últimas linhas do robo.log."""
    log_path = os.path.join(get_caminho_base(), "robo.log")
    if not os.path.exists(log_path):
        await update.message.reply_text("❌ Arquivo 'robo.log' ainda não foi criado.")
        return

    try:
        with open(log_path, "rb") as f:
            f.seek(0, os.SEEK_END)
            size = f.tell()
            # Lê os últimos 4KB para garantir que pegamos o final do log
            f.seek(max(0, size - 4000))
            content = f.read().decode('utf-8', errors='ignore')
            
        lines = content.splitlines()[-25:] # Pega as últimas 25 linhas
        texto_log = "\n".join(lines)
        
        if not texto_log.strip():
            texto_log = "O arquivo de log está vazio."
            
        await update.message.reply_text(f"📝 *Últimas linhas do robo.log:*\n\n```\n{texto_log}\n```", parse_mode='Markdown')
    except Exception as e:
        await update.message.reply_text(f"❌ Erro ao ler log: {e}")

async def help_command(update: Update, context: CallbackContext):
    """Handler para o comando /help ou /menu."""
    # Lista de comandos registrados
    comandos = [
        ("/start", "Inicia o robô"),
        ("/stop", "Para o robô"),
        ("/restart", "Reinicia o robô"),
        ("/status", "Mostra status do robô"),
        ("/logs", "Mostra últimas linhas do log"),
        ("/recarregar_config", "Recarrega configurações do bot"),
        ("/ocultar", "Oculta janela do Chrome"),
        ("/mostrar", "Mostra janela do Chrome"),
        ("/imprimir <Nome>", "Imprime relatório de motoboy ou cliente"),
        ("/garantia <Nome> <Início> <Fim>", "Gera recibo de garantia"),
        ("/enviar <Mensagem>", "Envia mensagem no grupo WhatsApp"),
        ("/alerta_auto", "Ativa/desativa alertas automáticos"),
        ("/mencao", "Ativa/desativa menção no WhatsApp"),
        ("/consultar_vale [Motoboy]", "Lista vales do dia (opcional filtrar por motoboy)"),
        ("/excluir_vale <ID>", "Exclui um vale pelo ID"),
        ("/lancar_vale", "Lança um novo vale para motoboy"),
        ("/historico", "Gera relatório de fechamento no Google Sheets"),
        ("/excel", "Gera relatório de fechamento (alternativo)"),
        ("/resumo", "Total de taxas e total do dia"),
        ("/canceladas", "Gera relatório de perdas"),
        ("/fechamento_manual", "Força geração do relatório de fechamento"),
        ("/atualizar_estoque", "Atualiza estoque pelo histórico"),
        ("/motos", "Vê entregadores na rua"),
        ("/pendentes", "Lista de pedidos na fila"),
        ("/estoque", "Vê itens com estoque baixo"),
        ("/help", "Mostra este menu de comandos"),
        ("/menu", "Mostra este menu de comandos"),
        ("/iniciar", "Alias para /start"),
        ("/parar", "Alias para /stop"),
        ("/reiniciar", "Alias para /restart"),
        ("/cancelar", "Cancela operações conversacionais"),
    ]

    # Função para escapar caracteres especiais do MarkdownV2
    def escape_md(text):
        # Escapa todos os caracteres especiais do MarkdownV2
        return re.sub(r'([_\*\[\]\(\)~`>#+\-=|{}.!\\])', r'\\\1', text)

    header = escape_md("🤖 ZÉ-BOT: MENU DE COMANDOS")
    help_text = f"*{header}*\n\n"
    for cmd, desc in comandos:
        help_text += f"🔹 `{escape_md(cmd)}` \\- {escape_md(desc)}\n"

    footer = escape_md("Todos os comandos acima estão disponíveis para qualquer usuário.\nSe precisar de detalhes, envie /help <comando> (não implementado).")
    help_text += f"\n*{footer}*"
    await update.message.reply_text(help_text, parse_mode='MarkdownV2')

async def recarregar_config_command(update: Update, context: CallbackContext):
    """Handler para o comando /recarregar_config."""
    await update.message.reply_text("🔄 Recarregando configurações para o bot do Telegram...")
    
    # Recarrega a config para o processo do telegram_bot
    config_ok = load_config()
    
    if config_ok:
        await update.message.reply_text("✅ Configurações do bot recarregadas.")
    else:
        await update.message.reply_text("❌ Falha ao recarregar configurações do bot. Verifique os logs.")

    # Envia o comando para o robo.py também recarregar
    await update.message.reply_text("🖥️ Enviando comando para o robô principal recarregar as configurações...")
    if "Online" not in get_robo_status():
        await update.message.reply_text("⚠️ O robô está offline. O comando para recarregar a configuração será ignorado por ele.")
    else:
        response = send_panel_command("RECARREGAR_CONFIG")
        await update.message.reply_text(f"Resposta do robô: {response}")

async def imprimir_command(update: Update, context: CallbackContext):
    """Handler para o comando /imprimir - INDEPENDENTE, sem depender do robo.py."""
    if not context.args:
        await update.message.reply_text("Uso: /imprimir <Nome do Motoboy>\n\nExemplo: /imprimir gledson")
        return
    
    nome_motoboy = " ".join(context.args)
    await update.message.reply_text(f"⏳ Buscando e imprimindo extrato de '{nome_motoboy}'...")
    
    # Executar em thread para não bloquear o bot
    import threading
    def executar_impressao():
        resultado = buscar_e_imprimir_extrato(nome_motoboy)
        if resultado:
            print(f"✅ Impressão executada para '{nome_motoboy}'")
        else:
            print(f"❌ Falha na impressão de '{nome_motoboy}'")
    
    thread = threading.Thread(target=executar_impressao, daemon=True)
    thread.start()
    
    await update.message.reply_text(f"✅ Comando de impressão enviado para '{nome_motoboy}'!")

async def garantia_command(update: Update, context: CallbackContext):
    """Handler para o comando /garantia."""
    if len(context.args) < 3:
        await update.message.reply_text("Uso: /garantia <Nome> <Início> <Fim>\nExemplo: /garantia jean dutra 18:00 02:00")
        return  # <-- Esta linha DEVE estar alinhada com o 'await' acima.
    fim = context.args[-1]
    inicio = context.args[-2]
    nome = " ".join(context.args[:-2])
    
    response = send_panel_command(f"GERAR_GARANTIA:{nome}|{inicio}|{fim}")
    await update.message.reply_text(response)
    return
    nome, inicio, fim = context.args[0], context.args[1], context.args[2]
    response = send_panel_command(f"GERAR_GARANTIA:{nome}|{inicio}|{fim}")
    await update.message.reply_text(response)
    
async def enviar_command(update: Update, context: CallbackContext):
    """Handler para o comando /enviar."""
    if not context.args:
        await update.message.reply_text("Uso: /enviar <Mensagem>")
    return
    mensagem = " ".join(context.args)
    response = send_panel_command(f"ENVIAR_MENSAGEM:{mensagem}")
    await update.message.reply_text(response)
    
async def alerta_auto_command(update: Update, context: CallbackContext):
    """Handler para o comando /alerta_auto."""
    response = send_panel_command("TOGGLE_ALERTA_AUTO")
    await update.message.reply_text(response)

async def mencao_command(update: Update, context: CallbackContext):
    """Handler para o comando /mencao."""
    response = send_panel_command("TOGGLE_MENCAO")
    await update.message.reply_text(response)

async def panel_commands(update: Update, context: CallbackContext):
    """Handler para comandos que interagem com o robô em execução."""
    command_map = {
        # Comandos que já existiam
        '/resumo': 'GERAR_RESUMO',
        '/canceladas': 'GERAR_CANCELADAS',
        # Comandos restaurados/adicionados
        '/fechamento_manual': 'FECHAMENTO_MANUAL',
        '/atualizar_estoque': 'ATUALIZAR_ESTOQUE',
        '/motos': 'VER_MOTOS',
        '/pendentes': 'VER_PENDENTES',
        '/estoque': 'VER_ESTOQUE',
    }
    command = update.message.text.split(' ')[0]

    if command in command_map:
        if "Online" not in get_robo_status():
            await update.message.reply_text("❌ O robô precisa estar online para executar este comando. Use /start primeiro.")
            return
            
        robo_command = command_map[command]
        response = send_panel_command(robo_command)
        await update.message.reply_text(response)

def _obter_pix_motoboy(nome, config_data):
    pix_map = config_data.get("pix_motoboys", {})
    if not nome:
        return ""
    nome_limpo = str(nome).strip()
    # Tenta busca direta
    pix = pix_map.get(nome_limpo)
    if pix:
        return str(pix).strip()
    # Tenta busca normalizada
    for k, v in pix_map.items():
        if k.lower() == nome_limpo.lower():
            return str(v).strip()
    return ""

def _get_vales_motoboy(nome_motoboy, excel_path):
    """Carrega e calcula o total de vales para um motoboy específico."""
    if not os.path.exists(excel_path):
        return 0.0
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        if "VALES" not in wb.sheetnames:
            return 0.0
        
        ws = wb["VALES"]
        total_vales = 0.0
        nome_motoboy_lower = nome_motoboy.lower()

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and len(row) > 2 and row[1] and row[1].lower() == nome_motoboy_lower:
                try:
                    total_vales += float(row[2])
                except (ValueError, TypeError):
                    continue
        return total_vales
    except Exception:
        return 0.0

def _carregar_dados_fechamento(excel_path):
    """Carrega e processa os dados do Excel para o fechamento."""
    if not os.path.exists(excel_path):
        return {}
    
    try:
        # Usa pandas para ler as duas abas necessárias
        df_detalhado = pd.read_excel(excel_path, sheet_name="EXTRATO DETALHADO")
        df_pagamentos = pd.read_excel(excel_path, sheet_name="PAGAMENTO_MOTOBOYS")
    except Exception as e:
        logging.error(f"Erro ao ler planilhas do Excel: {e}")
        return {}

    dados_fechamento = {}

    # Processa pagamentos para inicializar os motoboys
    for _, row in df_pagamentos.iterrows():
        nome = str(row.get("MOTOBOY", "")).strip()
        if nome and nome.upper() != "RETIRADA":
            dados_fechamento[nome] = {
                "qtd8": int(row.get("QTD R$ 8,00", 0) or 0),
                "qtd11": int(row.get("QTD R$ 11,00", 0) or 0),
                "prod_total": float(row.get("TOTAL A PAGAR (R$)", 0.0) or 0.0),
                "entregas": []
            }

    # Processa extrato detalhado para obter as entregas
    for _, row in df_detalhado.iterrows():
        nome = str(row.get('Motoboy', '')).strip()
        status = str(row.get('Status', '')).upper()
        if nome and nome in dados_fechamento and "CANCEL" not in status and "ABANDONED" not in status:
            try:
                valor = float(row.get('Valor (R$)', 0.0) or 0.0)
                hora_str = row.get('Hora', '')
                hora = pd.to_datetime(hora_str).strftime('%H:%M') if pd.notna(hora_str) else ''
                dados_fechamento[nome]['entregas'].append({"hora": hora, "valor": valor})
            except Exception:
                continue

    return dict(sorted(dados_fechamento.items()))


def _obter_nome_aba_sheets():
    data_str = get_data_operacional()
    partes = data_str.split("-")
    return f"{partes[0]}/{partes[1]}" if len(partes) >= 2 else data_str


# --- CONVERSATION HANDLER FOR /gerar_excel ---
GET_FECHAMENTO_DATE = range(3, 4)

async def gerar_excel_start(update: Update, context: CallbackContext):
    """Starts the conversation to generate the closing report."""
    await update.message.reply_text(
        "Para qual data você quer gerar o relatório?\n"
        "Envie a data no formato `dd-mm-aaaa` ou deixe em branco para usar a data de hoje.",
        parse_mode='Markdown'
    )
    return GET_FECHAMENTO_DATE

async def get_fechamento_date_handler(update: Update, context: CallbackContext):
    """Handles the date input and triggers the report generation."""
    date_str = update.message.text.strip()
    if not date_str:
        date_str = get_data_operacional()
        await update.message.reply_text(f"Nenhuma data informada. Usando a data de hoje: {date_str}")
    else:
        # Validar formato da data
        try:
            datetime.strptime(date_str, '%d-%m-%Y')
        except ValueError:
            await update.message.reply_text(
                "❌ Formato de data inválido. Por favor, use `dd-mm-aaaa` (ex: 22-02-2026).\n"
                "Ou envie uma mensagem vazia para usar a data de hoje.",
                parse_mode='Markdown'
            )
            return GET_FECHAMENTO_DATE # Permanece no mesmo estado

    # Chama a tarefa principal com a data correta
    await _gerar_excel_task(update, date_str)
    
    return ConversationHandler.END


async def _gerar_excel_task(update: Update, date_str: str):
    """Core logic to generate the closing report for a specific date in Google Sheets."""
    await update.message.reply_text(f"Iniciando geração do relatório para a data: {date_str}...")

    try:
        # Carregar config
        with open(ARQUIVO_CONFIG, 'r', encoding='utf-8') as f:
            config_data = json.load(f)
        
        google_sheets_config = config_data.get("google_sheets", {})
        cred_path = google_sheets_config.get("service_account_json")
        sheet_id = google_sheets_config.get("sheets_id")

        if not cred_path or not os.path.exists(cred_path) or not sheet_id:
            await update.message.reply_text("❌ Configuração do Google Sheets (JSON ou ID) não encontrada em config.json.")
            return

        excel_path = get_caminho_excel(date_str)
        if not os.path.exists(excel_path):
            await update.message.reply_text(f"❌ Arquivo de controle para a data {date_str} não foi encontrado.")
            return
            
        dados_fechamento = _carregar_dados_fechamento(excel_path)

        if not dados_fechamento:
            await update.message.reply_text(f"ℹ️ Não foram encontrados dados de motoboys no arquivo de {date_str} para gerar o fechamento.")
            return
            
        await update.message.reply_text("Conectando ao Google Sheets...")
        scopes = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
        creds = Credentials.from_service_account_file(cred_path, scopes=scopes)
        client = gspread.authorize(creds)
        sh = client.open_by_key(sheet_id)
        
        # Use the provided date to name the sheet
        partes = date_str.split("-")
        sheet_title = f"{partes[0]}/{partes[1]}" if len(partes) >= 2 else date_str
        
        try:
            ws = sh.worksheet(sheet_title)
            ws.clear()
        except WorksheetNotFound:
            ws = sh.add_worksheet(title=sheet_title, rows=200, cols=15)
        
        cabecalhos = ["Motoboy", "8", "11", "Entregas", "Pago", "PIX", "TOTAL", "Val", "Garantido", "8", "11", "", "Inicio", "Fim"]
        linhas = [cabecalhos]
        
        for nome, info in dados_fechamento.items():
            vale = _get_vales_motoboy(nome, excel_path)
            pix = _obter_pix_motoboy(nome, config_data)
            
            total_entregas = info['qtd8'] + info['qtd11']

            linhas.append([
                nome, info['qtd8'], info['qtd11'], total_entregas, "", pix, None, vale, 0.0, 0, 0, "", "", ""
            ])
            
        end_row = len(linhas)
        await update.message.reply_text(f"Enviando {end_row-1} registros para a planilha '{sheet_title}'...")
        ws.update(values=linhas, range_name=f"A1:N{end_row}", value_input_option="USER_ENTERED")
        
        for idx in range(2, end_row + 1):
            ws.update_cell(idx, 7, f"=B{idx}*8+C{idx}*11-H{idx}+I{idx}")

        requests_batch = [
            {"repeatCell": {"range": {"sheetId": ws.id, "startRowIndex": 0, "endRowIndex": 1}, "cell": {"userEnteredFormat": {"backgroundColor": {"red": 0.2, "green": 0.2, "blue": 0.2}, "textFormat": {"foregroundColor": {"red": 1, "green": 1, "blue": 1}, "bold": True}}}, "fields": "userEnteredFormat(backgroundColor,textFormat)"}},
            {"updateSheetProperties": {"properties": {"sheetId": ws.id, "gridProperties": {"frozenRowCount": 1}}, "fields": "gridProperties.frozenRowCount"}},
            {"setDataValidation": {"range": {"sheetId": ws.id, "startRowIndex": 1, "endRowIndex": end_row, "startColumnIndex": 4, "endColumnIndex": 5}, "rule": {"condition": {"type": "BOOLEAN"}}}},
            {"addConditionalFormatRule": {"rule": {"ranges": [{"sheetId": ws.id, "startRowIndex": 1, "endRowIndex": end_row}], "booleanRule": {"condition": {"type": "CUSTOM_FORMULA", "values": [{"userEnteredValue": "=$E2=TRUE"}]}, "format": {"backgroundColor": {"red": 0.8, "green": 1, "blue": 0.8}}}}, "index": 0}}
        ]

        if requests_batch:
            sh.batch_update({"requests": requests_batch})

        sheet_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/edit#gid={ws.id}"
        await update.message.reply_text(f"✅ Relatório para {date_str} gerado com sucesso!\n\n[Abrir Planilha]({sheet_url})", parse_mode='Markdown', disable_web_page_preview=True)

    except Exception as e:
        logging.error(f"Erro ao gerar fechamento no Google Sheets: {e}")
        await update.message.reply_text(f"❌ Ocorreu um erro grave ao gerar o relatório: {e}")




# --- CONVERSATION HANDLER FOR /lancar_vale ---
CHOOSE_MOTOBOY, GET_VALOR, GET_MOTIVO = range(3)

def _get_motoboys_from_config():
    """Helper to get motoboy list from config.json."""
    try:
        with open(ARQUIVO_CONFIG, 'r', encoding='utf-8') as f:
            config = json.load(f)
        # The config stores motoboys as a dict {email: name}. We want the names.
        return list(config.get('motoboys', {}).values())
    except Exception as e:
        print(f"⚠️ Erro ao ler motoboys do config: {e}")
        return []

def _get_motoboys_from_excel():
    """Busca motoboys que estão com pedidos no dia (Excel)."""
    try:
        excel_path = get_caminho_excel()
        if not os.path.exists(excel_path):
            return []
        
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        if "EXTRATO DETALHADO" not in wb.sheetnames:
            return []
        
        ws = wb["EXTRATO DETALHADO"]
        motoboys_set = set()
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and len(row) > 6 and row[6]:  # Coluna 6 é o motoboy
                motoboy = str(row[6]).strip()
                if motoboy and motoboy.upper() not in ["MOTOBOY", "N/A"]:
                    motoboys_set.add(motoboy)
        
        return list(motoboys_set)
    except Exception as e:
        print(f"⚠️ Erro ao ler motoboys do Excel: {e}")
        return []

def _get_all_motoboys():
    """Combina motoboys do config + motoboys com pedidos no dia."""
    motoboys_config = _get_motoboys_from_config()
    motoboys_excel = _get_motoboys_from_excel()
    
    # Combinar e remover duplicatas (case-insensitive)
    combined = {}
    for m in motoboys_config:
        combined[m.lower()] = m
    for m in motoboys_excel:
        m_lower = m.lower()
        if m_lower not in combined:
            combined[m_lower] = m
        else:
            # Se já existe, usar a versão do config (mais "oficial")
            combined[m_lower] = combined.get(m_lower, m)
    
    return sorted(list(set(combined.values())))

async def lancar_vale_start(update: Update, context: CallbackContext):
    """Starts the conversation to add a vale."""
    # Buscar motoboys: config + os com pedidos no dia
    motoboys = _get_all_motoboys()
    
    if not motoboys:
        await update.message.reply_text("❌ Nenhum motoboy cadastrado ou com pedidos no dia.")
        return ConversationHandler.END

    keyboard = [[InlineKeyboardButton(name, callback_data=name)] for name in motoboys]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await update.message.reply_text(f"📋 *{len(motoboys)} motoboy(s) disponível(is)*\n\nPara qual você quer lançar o vale?", reply_markup=reply_markup, parse_mode='Markdown')
    return CHOOSE_MOTOBOY

async def choose_motoboy_handler(update: Update, context: CallbackContext):
    """Handles the motoboy selection."""
    query = update.callback_query
    await query.answer()
    
    motoboy = query.data
    context.user_data['motoboy'] = motoboy
    
    await query.edit_message_text(text=f"Motoboy selecionado: *{motoboy}*.\n\nQual o *valor* do vale? (ex: 50.50)", parse_mode='Markdown')
    return GET_VALOR

async def get_valor_handler(update: Update, context: CallbackContext):
    """Handles getting the vale value."""
    valor_str = update.message.text
    try:
        valor_float = float(valor_str.replace(",", "."))
        context.user_data['valor'] = valor_float
        await update.message.reply_text(f"Valor definido: *R$ {valor_float:.2f}*.\n\nAgora, digite o *motivo* do vale.", parse_mode='Markdown')
        return GET_MOTIVO
    except ValueError:
        await update.message.reply_text("❌ Valor inválido. Por favor, envie apenas números (ex: 50 ou 50.50). Qual o valor correto?")
        return GET_VALOR # Stay in the same state

async def get_motivo_handler(update: Update, context: CallbackContext):
    """Handles getting the reason and saving the vale."""
    motivo = update.message.text
    motoboy = context.user_data['motoboy']
    valor = context.user_data['valor']

    await update.message.reply_text(f"Registrando vale de *R$ {valor:.2f}* para *{motoboy}*...", parse_mode='Markdown')

    try:
        excel_path = get_caminho_excel()
        # Se o arquivo não existir, cria um novo com o cabeçalho
        if not os.path.exists(excel_path):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "VALES"
            ws.append(["HORA", "MOTOBOY", "VALOR", "MOTIVO"])
        else:
            wb = openpyxl.load_workbook(excel_path)
            ws = wb["VALES"] if "VALES" in wb.sheetnames else wb.create_sheet("VALES")

        # Garante cabeçalho se a planilha estiver vazia
        if ws.max_row == 1 and ws.cell(row=1, column=1).value is None:
            ws.append(["HORA", "MOTOBOY", "VALOR", "MOTIVO"])

        hora = datetime.now().strftime('%H:%M')
        ws.append([hora, motoboy, valor, motivo])
        wb.save(excel_path)
        
        await update.message.reply_text(f"✅ Vale registrado com sucesso!")
    except Exception as e:
        await update.message.reply_text(f"❌ Ocorreu um erro ao salvar no Excel: {e}")
    
    # Clean up user_data
    context.user_data.clear()
    return ConversationHandler.END

async def cancel_handler(update: Update, context: CallbackContext):
    """Cancels the current conversation."""
    await update.message.reply_text("Operação cancelada.")
    context.user_data.clear()
    return ConversationHandler.END

async def consultar_vale_command(update: Update, context: CallbackContext):
    """Handler para o comando /consultar_vale diretamente no Excel."""
    try:
        excel_path = get_caminho_excel()
        if not os.path.exists(excel_path):
            await update.message.reply_text(f"❌ Arquivo de controle do dia ({os.path.basename(excel_path)}) não encontrado.")
            return

        wb = openpyxl.load_workbook(excel_path, data_only=True)
        
        if "VALES" not in wb.sheetnames:
            await update.message.reply_text("ℹ️ A planilha 'VALES' ainda não existe no arquivo de hoje.")
            return

        ws = wb["VALES"]
        vales = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if any(cell is not None for cell in row):
                vales.append({"id": row_idx, "data": row})
        
        if not vales:
            await update.message.reply_text("ℹ️ Nenhum vale encontrado na planilha de hoje.")
            return

        motoboy_filtro = " ".join(context.args).lower() if context.args else None
        if motoboy_filtro:
            vales_filtrados = [v for v in vales if v["data"] and len(v["data"]) > 1 and str(v["data"][1]).lower() == motoboy_filtro]
            if not vales_filtrados:
                await update.message.reply_text(f"ℹ️ Nenhum vale encontrado para '{context.args[0]}'.")
                return
            vales = vales_filtrados

        mensagem = "📖 *Vales Registrados Hoje*\n\n"
        total_vales = 0.0
        for vale in vales:
            row_data = vale["data"]
            vale_id = vale["id"]
            try:
                hora = row_data[0] or "--:--"
                if isinstance(hora, (datetime, datetime_time)):
                    hora = hora.strftime('%H:%M')
                moto = row_data[1] or "N/A"
                valor = float(row_data[2]) if row_data[2] is not None else 0.0
                motivo = row_data[3] or ""
                
                mensagem += f"`ID: {vale_id}` | *{moto}*\n"
                mensagem += f"  R$ {valor:.2f} às {hora} - {motivo}\n"
                mensagem += "--------------------------------------\n"
                total_vales += valor
            except (IndexError, ValueError) as e:
                logging.warning(f"Skipping malformed row {vale_id} in VALES sheet: {e}")

        mensagem += f"\n*Total em vales: R$ {total_vales:.2f}*"
        await update.message.reply_text(mensagem, parse_mode='Markdown')

    except Exception as e:
        logging.error(f"Erro em consultar_vale_command: {e}")
        await update.message.reply_text(f"❌ Ocorreu um erro ao consultar os vales: {e}")

async def excluir_vale_command(update: Update, context: CallbackContext):
    """Handler para o comando /excluir_vale diretamente no Excel."""
    try:
        if not context.args:
            await update.message.reply_text("Uso: /excluir_vale <ID do Vale>\n(Use /consultar_vale para ver os IDs)")
            return
        
        try:
            vale_id = int(context.args[0])
            if vale_id < 2:
                 await update.message.reply_text("❌ ID inválido. O ID deve ser um número maior que 1.")
                 return
        except ValueError:
            await update.message.reply_text("❌ ID inválido. Forneça o número do ID do vale.")
            return

        excel_path = get_caminho_excel()
        if not os.path.exists(excel_path):
            await update.message.reply_text(f"❌ Arquivo de controle do dia ({os.path.basename(excel_path)}) não encontrado.")
            return

        wb = openpyxl.load_workbook(excel_path)

        if "VALES" not in wb.sheetnames:
            await update.message.reply_text("ℹ️ A planilha 'VALES' não existe no arquivo de hoje.")
            return
            
        ws = wb["VALES"]
        
        if vale_id > ws.max_row:
            await update.message.reply_text(f"❌ ID {vale_id} não encontrado na planilha.")
            return
            
        dados_linha = [cell.value for cell in ws[vale_id]]
        motoboy = dados_linha[1]
        valor = dados_linha[2]

        ws.delete_rows(vale_id)
        wb.save(excel_path)
        
        await update.message.reply_text(f"✅ Vale ID {vale_id} (R$ {valor} - {motoboy}) foi excluído com sucesso.")

    except Exception as e:
        logging.error(f"Erro em excluir_vale_command: {e}")
        await update.message.reply_text(f"❌ Ocorreu um erro ao excluir o vale: {e}")

async def log_command(update: Update, context: CallbackContext):
    """Handler para o comando /log."""
    logging.info(f"User {update.effective_user.id} issued command: {update.message.text}")

# ==================================================================================
# FUNÇÃO PRINCIPAL E INICIALIZAÇÃO DO BOT
# ==================================================================================

def load_config():
    """Carrega as configurações do JSON."""
    global TELEGRAM_TOKEN, ADMIN_CHAT_ID
    try:
        # Adicionar depuração aqui
        base_path = get_caminho_base()
        config_file_path = os.path.join(base_path, ARQUIVO_CONFIG)
        print(f"[DEBUG] Caminho base do bot: {base_path}")
        print(f"[DEBUG] Procurando config.json em: {config_file_path}")
        print(f"[DEBUG] config.json existe? {os.path.exists(config_file_path)}")
        
        with open(ARQUIVO_CONFIG, 'r', encoding='utf-8') as f:
            config = json.load(f)
        TELEGRAM_TOKEN = config.get('telegram_token')
        ADMIN_CHAT_ID = config.get('telegram_chat_id')
        
        if not TELEGRAM_TOKEN or not ADMIN_CHAT_ID:
            print("❌ ERRO: 'telegram_token' e 'telegram_chat_id' devem ser definidos no config.json")
            return False
        
        ADMIN_CHAT_ID = str(ADMIN_CHAT_ID)
        return True
        
    except FileNotFoundError:
        print(f"❌ Arquivo de configuração '{ARQUIVO_CONFIG}' não encontrado. Tentando ler token do Google Sheets...");
        return load_token_from_sheets()

    except (json.JSONDecodeError, KeyError) as e:
        print(f"❌ ERRO: Falha ao ler o arquivo de configuração. Verifique o formato. Detalhe: {e}")
        return False

def load_token_from_sheets():
    """Tenta carregar o token do telegram do Google Sheets"""
    global TELEGRAM_TOKEN, ADMIN_CHAT_ID
    if not TEM_GSHEETS:
        print("❌ Sem suporte a Google Sheets (gspread).")
        return False

    try:
        if not os.path.exists(ARQUIVO_CONFIG):
            print("❌ Arquivo config.json não encontrado.")
            return False
        with open(ARQUIVO_CONFIG, 'r', encoding='utf-8') as f:
            config = json.load(f)

        sheets_config = config.get("google_sheets", {})
        cred_file = sheets_config.get("service_account_json")
        sheet_id = sheets_config.get("sheets_id")
        if not cred_file or not sheet_id or not os.path.exists(cred_file):
            print("❌ Sem credenciais ou ID da planilha no config.json")
            return False

        scopes = ["https://www.googleapis.com/auth/spreadsheets.readonly"]
        creds = Credentials.from_service_account_file(cred_file, scopes=scopes)
        client = gspread.authorize(creds)
        sh = client.open_by_key(sheet_id)
        ws = sh.sheet1
        TELEGRAM_TOKEN = ws.acell('A1').value
        ADMIN_CHAT_ID = ws.acell('B1').value
        print(f"✅ Token do Telegram lido da planilha: {TELEGRAM_TOKEN[:4]}... (ChatID: {ADMIN_CHAT_ID})")
        return True

    except Exception as e:
        print(f"❌ Erro ao ler token do Sheets: {e}")
        return False

def criar_aplicacao(token):
    """Cria e configura a aplicação do Telegram Bot com retry logic."""
    for tentativa in range(3):
        try:
            app = Application.builder().token(token).connect_timeout(30).read_timeout(30).build()
            return app
        except Exception as e:
            if tentativa < 2:
                print(f"⚠️ Erro ao criar aplicação (tentativa {tentativa+1}/3): {e}")
                time.sleep(5)
            else:
                raise

async def limpar_updates_antigos(bot):
    """Limpar updates antigos do Telegram API para evitar erro 409 (Conflict)."""
    try:
        print("🧹 Limpando updates antigos da fila do Telegram...")
        updates = await bot.get_updates(limit=100, timeout=1)
        if updates:
            offset = updates[-1].update_id + 1
            await bot.get_updates(offset=offset, limit=1, timeout=1)
            print(f"✅ Limpeza concluída: {len(updates)} updates antigos descartados")
        else:
            print("✅ Nenhum update antigo encontrado")
    except Exception as e:
        print(f"⚠️ Erro ao limpar updates antigos (não crítico): {e}")

async def error_handler(update, context):
    """Handler global para erros do bot."""
    try:
        from telegram.error import Conflict
        if isinstance(context.error, Conflict):
            print("⚠️ Erro 409 Conflito: Outra instância do bot está conectada")
            return
    except Exception:
        pass
    print(f"❌ Erro no bot: {context.error}")

def main():
    """Função principal que inicia o bot do Telegram."""
    if not load_config():
        sys.exit("Encerrando devido a erro na configuração.")

    lock_path = os.path.join(get_caminho_base(), "bot.lock")
    if os.path.exists(lock_path):
        try:
            with open(lock_path, 'r') as f:
                old_pid = int(f.read().strip())
            if psutil.pid_exists(old_pid):
                proc = psutil.Process(old_pid)
                # Adicionado para verificar se o processo é realmente o bot
                cmdline = proc.cmdline()
                if cmdline and any(os.path.basename(__file__).lower() in arg.lower() for arg in cmdline):
                    print(f"❌ ERRO: Bot já está rodando (PID: {old_pid}). Saindo.")
                    sys.exit(1)
        except (psutil.NoSuchProcess, psutil.AccessDenied, FileNotFoundError, ValueError):
            # Lock file antigo ou processo morto, pode sobrescrever
            pass
    
    with open(lock_path, 'w') as f:
        f.write(str(os.getpid()))
    atexit.register(lambda: os.remove(lock_path) if os.path.exists(lock_path) else None)

    print("🤖 Iniciando Bot de Controle do Telegram...")
    print("   Aguardando 5 segundos para garantir limpeza da conexão anterior...")
    time.sleep(5)

    application = criar_aplicacao(TELEGRAM_TOKEN)


    # --- CONVERSATION HANDLERS SETUP ---
    conv_handler_vale = ConversationHandler(
        entry_points=[CommandHandler("lancar_vale", lancar_vale_start)],
        states={
            CHOOSE_MOTOBOY: [CallbackQueryHandler(choose_motoboy_handler, pattern=".*")],
            GET_VALOR: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_valor_handler)],
            GET_MOTIVO: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_motivo_handler)],
        },
        fallbacks=[CommandHandler("cancelar", cancel_handler)],
    )

    conv_handler_excel = ConversationHandler(
        entry_points=[CommandHandler(["historico", "excel"], gerar_excel_start)],
        states={
            GET_FECHAMENTO_DATE: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_fechamento_date_handler)],
        },
        fallbacks=[CommandHandler("cancelar", cancel_handler)],
    )

    application.add_handler(conv_handler_vale)
    application.add_handler(conv_handler_excel)


    # Comandos básicos de controle
    application.add_handler(CommandHandler(["start", "iniciar"], start_command))
    application.add_handler(CommandHandler(["stop", "parar"], stop_command))
    application.add_handler(CommandHandler(["restart", "reiniciar"], restart_command))
    application.add_handler(CommandHandler("status", status_command))

    # Comandos de utilidade
    application.add_handler(CommandHandler(["help", "menu", "ajuda"], help_command))
    application.add_handler(CommandHandler("logs", logs_command))
    application.add_handler(CommandHandler("recarregar_config", recarregar_config_command))
    application.add_handler(CommandHandler("ocultar", hide_chrome_command))
    application.add_handler(CommandHandler("mostrar", show_chrome_command))

    # Comandos de interação com o robô
    application.add_handler(CommandHandler("imprimir", imprimir_command))
    application.add_handler(CommandHandler("garantia", garantia_command))
    application.add_handler(CommandHandler("enviar", enviar_command))
    application.add_handler(CommandHandler("alerta_auto", alerta_auto_command))
    application.add_handler(CommandHandler("mencao", mencao_command))

    # Comandos de Vales (não conversacionais)
    application.add_handler(CommandHandler("consultar_vale", consultar_vale_command))
    application.add_handler(CommandHandler("excluir_vale", excluir_vale_command))

    # Handler para comandos específicos do painel
    panel_command_list = [
        'resumo', 'canceladas', 'fechamento_manual', 'atualizar_estoque', 
        'motos', 'pendentes', 'estoque'
    ]
    application.add_handler(CommandHandler(panel_command_list, panel_commands))

    # Registrar error handler para erros de conflito (409)
    application.add_error_handler(error_handler)

    print("✅ Bot de controle pronto para receber comandos.")
    print("🚀 Iniciando polling do Telegram...")
    
# Aguardar limpeza de updates antigos
    import asyncio
    try:
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        loop.run_until_complete(limpar_updates_antigos(application.bot))
    except Exception as e:
        print(f"⚠️ Erro ao limpar updates antigos: {e}")
    
    # Iniciar polling (cria seu próprio event loop)
    application.run_polling()

if __name__ == "__main__":
    main()