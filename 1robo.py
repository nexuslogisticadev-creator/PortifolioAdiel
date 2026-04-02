import difflib
import requests
# Truque para ignorar o curl-cffi e usar o requests normal
import requests

class CFFIFake:
    @staticmethod
    def post(*args, **kwargs):
        if 'impersonate' in kwargs: del kwargs['impersonate']
        return requests.post(*args, **kwargs)
    @staticmethod
    def get(*args, **kwargs):
        if 'impersonate' in kwargs: del kwargs['impersonate']
        return requests.get(*args, **kwargs)
cffi_requests = CFFIFake()
import sys
import os
import openpyxl
from datetime import datetime
from tkinter import messagebox, simpledialog
LOG_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "robo.log")
import difflib
import requests
from curl_cffi import requests as cffi_requests
import sys
import io
import os
import openpyxl
from datetime import datetime
from tkinter import messagebox, simpledialog

# --- ADICIONE ESTAS LINHAS PARA CORRIGIR O ERRO DO EMOJI ---
LOG_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "robo.log")

class TeeStream(io.TextIOBase):
    def __init__(self, *streams):
        self.streams = streams

    def write(self, s):
        for stream in self.streams:
            try:
                stream.write(s)
                # Adicione estas duas linhas abaixo:
                if hasattr(stream, "flush"):
                    stream.flush()
            except Exception:
                pass
        return len(s)

    def flush(self):
        for stream in self.streams:
            try:
                stream.flush()
            except Exception:
                pass

# Corrige erro de recursão ao redefinir sys.stdout/sys.stderr para TeeStream
# Usa sys.__stdout__ e sys.__stderr__ para garantir acesso ao buffer original
try:
    _stdout = io.TextIOWrapper(sys.__stdout__.buffer, encoding='utf-8', line_buffering=True, write_through=True)
    _stderr = io.TextIOWrapper(sys.__stderr__.buffer, encoding='utf-8', line_buffering=True, write_through=True)
except AttributeError:
    # Fallback: usa sys.__stdout__/sys.__stderr__ diretamente
    _stdout = sys.__stdout__
    _stderr = sys.__stderr__
# Evita redefinir sys.stdout/sys.stderr se já forem TeeStream (previne recursão/erro)
_log_file = open(LOG_PATH, "a", encoding="utf-8", buffering=1)
if not isinstance(sys.stdout, TeeStream) and not hasattr(sys.stdout, 'buffer'):
    sys.stdout = TeeStream(_stdout, _log_file)
if not isinstance(sys.stderr, TeeStream) and not hasattr(sys.stderr, 'buffer'):
    sys.stderr = TeeStream(_stderr, _log_file)
# -----------------------------------------------------------
import openpyxl
from openpyxl.styles import Font, PatternFill
import time
import random
from urllib.parse import urlparse
# ... resto das importações ...from curl_cffi import requests as cffi_requests  # REDE INVISÍVEL
import time
import random 
import winsound
import pyperclip
import os
import sys
import math
import re
import unicodedata
import json
import subprocess
import psutil
from datetime import datetime, timedelta

# --- BIBLIOTECAS CHROME (ATUALIZADO) ---
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, WebDriverException
from webdriver_manager.chrome import ChromeDriverManager

# --- BIBLIOTECAS DE IMPRESSÃO (WINDOWS) ---
try:
    import win32print
    import win32api
    TEM_IMPRESSORA = True
except ImportError:
    TEM_IMPRESSORA = False
    print("⚠️ AVISO: Biblioteca de impressão não encontrada. Instale: pip install pywin32")

# --- 1. VERIFICAÇÃO DE BIBLIOTECAS ---
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ ERRO: FALTA 'openpyxl'. Instale com: pip install openpyxl")
    input("Enter para sair..."); exit()

try:
    import geocoder
    TEM_GPS = True
except ImportError:
    TEM_GPS = False
    print("⚠️ AVISO: Sem GPS (instale: pip install geocoder)")

# ================= CARREGAMENTO DE CONFIGURAÇÕES =================
def carregar_configuracoes():
    """Carrega todas as configurações do arquivo config.json"""
    try:
        with open('config.json', 'r', encoding='utf-8') as f:
            config = json.load(f)
        
        configuracoes = {
            'nome_grupo': config.get('grupo_whatsapp', 'Zé Número cliente'),
            'endereco_loja': config.get('endereco_loja', 'Rua Sete de Setembro 1178, Chapecó'),
            'email': config.get('email', ''),
            'senha': config.get('senha', ''),
            'telegram_token': config.get('telegram_token', ''),
            'telegram_chat_id': config.get('telegram_chat_id', ''),
            'path_backup': config.get('path_backup', ''),
            'motoboys': config.get('motoboys', {}),
            'bairros': config.get('bairros', {}),
            'pix_motoboys': config.get('pix_motoboys', {}),
            'google_sheets': config.get('google_sheets', {}),
            'debug_alerta_retirada_todos': config.get('debug_alerta_retirada_todos', False),
            'alerta_retirada_auto': config.get('alerta_retirada_auto', False),
            'whatsapp_mencao_ativa': config.get('whatsapp_mencao_ativa', False),
            'url_api': config.get('url_api', ''),
            'url_principal': config.get('url_principal', ''),
            'headers_api': config.get('headers_api', {}),
            'api_request': config.get('api_request', {}),
            'protecao': config.get('protecao', {}),
            'categorias_produtos': config.get('categorias_produtos', {}),
        }
        
        print("✅ Configurações carregadas do config.json")
        return configuracoes
    except FileNotFoundError:
        print("❌ ERRO: Arquivo config.json não encontrado!")
        return None
    except json.JSONDecodeError as e:
        print(f"❌ ERRO ao ler config.json: {e}")
        return None

def atualizar_config_flag(chave, valor):
    """Atualiza um flag booleano no config.json e no CONFIG em memoria."""
    global CONFIG
    try:
        with open('config.json', 'r', encoding='utf-8') as f:
            config = json.load(f)
        config[chave] = bool(valor)
        with open('config.json', 'w', encoding='utf-8') as f:
            json.dump(config, f, indent=4, ensure_ascii=False)
        if CONFIG is not None:
            CONFIG[chave] = bool(valor)
        return True
    except Exception as e:
        print(f"❌ Erro ao atualizar config {chave}: {e}")
        return False

# Carrega as configurações
CONFIG = carregar_configuracoes()
if CONFIG is None:
    print("❌ Não foi possível carregar as configurações. Encerrando...")
    input("Pressione Enter para sair...")
    exit()

#// ================= CONFIGURAÇÕES =================
URL_API = CONFIG['url_api']
NOME_GRUPO_FIXO = CONFIG['nome_grupo']

DISTANCIA_MAXIMA_ENTRE_CLIENTES = 2.0 
ANGULO_MAXIMO_DIFERENCA = 45 

# Carrega motoboys do config
MOTOBOYS_API = CONFIG['motoboys']

# Carrega bairros do config
BAIRROS_VALORES = CONFIG['bairros']
BAIRROS_NAO_CADASTRADOS_LOGADOS = set()

TELEGRAM_TOKEN = CONFIG['telegram_token']
TELEGRAM_CHAT_ID = CONFIG['telegram_chat_id']
DEBUG_ALERTA_RETIRADA_TODOS = CONFIG.get('debug_alerta_retirada_todos', False)
ALERTA_RETIRADA_AUTO = CONFIG.get('alerta_retirada_auto', False)

# Inicialização de variáveis globais de controle de tempo
LAST_MAIN_REFRESH = time.time()

STATUS_CANCELADOS_LISTA = [
    "ABANDONED", "CANCEL", "DEVOLVIDO", "POC_ABANDONED", 
    "CANCELLED", "POC_EXPIRED", "USER_CANCELLED", "SYS_CANCELLED",
    "POC_REJECTED", "DELIVERY_FAILED"
]

STATUS_FINALIZADOS = STATUS_CANCELADOS_LISTA + ["DELIVERED", "POC_DELIVERED", "FINISHED"]

# --- VARIÁVEIS GLOBAIS ---
TOKEN_ATUAL = ""
IDS_PROCESSADOS = set() 

class TeeStream(io.TextIOBase):
    def __init__(self, *streams):
        self.streams = streams

    def write(self, s):
        for stream in self.streams:
            try:
                stream.write(s)
            except Exception:
                pass
        return len(s)

    def flush(self):
        for stream in self.streams:
            try:
                stream.flush()
            except Exception:
                pass

_stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', line_buffering=True, write_through=True)
_stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', line_buffering=True, write_through=True)
_log_file = open(LOG_PATH, "a", encoding="utf-8", buffering=1)
sys.stdout = TeeStream(_stdout, _log_file)
sys.stderr = TeeStream(_stderr, _log_file)
# -----------------------------------------------------------
import openpyxl
from openpyxl.styles import Font, PatternFill
import time
import random
from urllib.parse import urlparse
# ... resto das importações ...from curl_cffi import requests as cffi_requests  # REDE INVISÍVEL
import time
import random 
import winsound
import pyperclip
import os
import sys
import math
import re
import unicodedata
import json
import subprocess
import psutil
from datetime import datetime, timedelta

# --- BIBLIOTECAS CHROME (ATUALIZADO) ---
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, WebDriverException
from webdriver_manager.chrome import ChromeDriverManager

# --- BIBLIOTECAS DE IMPRESSÃO (WINDOWS) ---
try:
    import win32print
    import win32api
    TEM_IMPRESSORA = True
except ImportError:
    TEM_IMPRESSORA = False
    print("⚠️ AVISO: Biblioteca de impressão não encontrada. Instale: pip install pywin32")

# --- 1. VERIFICAÇÃO DE BIBLIOTECAS ---
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ ERRO: FALTA 'openpyxl'. Instale com: pip install openpyxl")
    input("Enter para sair..."); exit()

try:
    import geocoder
    TEM_GPS = True
except ImportError:
    TEM_GPS = False
    print("⚠️ AVISO: Sem GPS (instale: pip install geocoder)")

# ================= CARREGAMENTO DE CONFIGURAÇÕES =================
def carregar_configuracoes():
    """Carrega todas as configurações do arquivo config.json"""
    try:
        with open('config.json', 'r', encoding='utf-8') as f:
            config = json.load(f)
        
        configuracoes = {
            'nome_grupo': config.get('grupo_whatsapp', 'Zé Número cliente'),
            'endereco_loja': config.get('endereco_loja', 'Rua Sete de Setembro 1178, Chapecó'),
            'email': config.get('email', ''),
            'senha': config.get('senha', ''),
            'telegram_token': config.get('telegram_token', ''),
            'telegram_chat_id': config.get('telegram_chat_id', ''),
            'path_backup': config.get('path_backup', ''),
            'motoboys': config.get('motoboys', {}),
            'bairros': config.get('bairros', {}),
            'pix_motoboys': config.get('pix_motoboys', {}),
            'google_sheets': config.get('google_sheets', {}),
            'debug_alerta_retirada_todos': config.get('debug_alerta_retirada_todos', False),
            'alerta_retirada_auto': config.get('alerta_retirada_auto', False),
            'whatsapp_mencao_ativa': config.get('whatsapp_mencao_ativa', False),
            'url_api': config.get('url_api', ''),
            'url_principal': config.get('url_principal', ''),
            'headers_api': config.get('headers_api', {}),
            'api_request': config.get('api_request', {}),
            'protecao': config.get('protecao', {}),
            'categorias_produtos': config.get('categorias_produtos', {}),
        }
        
        print("✅ Configurações carregadas do config.json")
        return configuracoes
    except FileNotFoundError:
        print("❌ ERRO: Arquivo config.json não encontrado!")
        return None
    except json.JSONDecodeError as e:
        print(f"❌ ERRO ao ler config.json: {e}")
        return None

def atualizar_config_flag(chave, valor):
    """Atualiza um flag booleano no config.json e no CONFIG em memoria."""
    global CONFIG
    try:
        with open('config.json', 'r', encoding='utf-8') as f:
            config = json.load(f)
        config[chave] = bool(valor)
        with open('config.json', 'w', encoding='utf-8') as f:
            json.dump(config, f, indent=4, ensure_ascii=False)
        if CONFIG is not None:
            CONFIG[chave] = bool(valor)
        return True
    except Exception as e:
        print(f"❌ Erro ao atualizar config {chave}: {e}")
        return False

# Carrega as configurações
CONFIG = carregar_configuracoes()
if CONFIG is None:
    print("❌ Não foi possível carregar as configurações. Encerrando...")
    input("Pressione Enter para sair...")
    exit()

#// ================= CONFIGURAÇÕES =================
URL_API = CONFIG['url_api']
ENDERECO_LOJA = CONFIG['endereco_loja']
NOME_GRUPO_FIXO = CONFIG['nome_grupo']

DISTANCIA_MAXIMA_ENTRE_CLIENTES = 2.0 
ANGULO_MAXIMO_DIFERENCA = 45 

# Carrega motoboys do config
MOTOBOYS_API = CONFIG['motoboys']

# Carrega bairros do config
BAIRROS_VALORES = CONFIG['bairros']
BAIRROS_NAO_CADASTRADOS_LOGADOS = set()

TELEGRAM_TOKEN = CONFIG['telegram_token']
TELEGRAM_CHAT_ID = CONFIG['telegram_chat_id']
DEBUG_ALERTA_RETIRADA_TODOS = CONFIG.get('debug_alerta_retirada_todos', False)
ALERTA_RETIRADA_AUTO = CONFIG.get('alerta_retirada_auto', False)

# Inicialização de variáveis globais de controle de tempo
LAST_MAIN_REFRESH = time.time()

STATUS_CANCELADOS_LISTA = [
    "ABANDONED", "CANCEL", "DEVOLVIDO", "POC_ABANDONED", 
    "CANCELLED", "POC_EXPIRED", "USER_CANCELLED", "SYS_CANCELLED",
    "POC_REJECTED", "DELIVERY_FAILED"
]

STATUS_FINALIZADOS = STATUS_CANCELADOS_LISTA + ["DELIVERED", "POC_DELIVERED", "FINISHED"]

# --- VARIÁVEIS GLOBAIS ---
TOKEN_ATUAL = ""
IDS_PROCESSADOS = set() 
pedidos_ja_enviados = set()
pedidos_em_espera = {} 
CACHE_NOMES_DO_DIA = {} 
CACHE_STATUS_PEDIDOS = {} 
TIMESTAMP_ACEITOS = {}  # Guarda quando cada pedido foi CRIADO (hora original do pedido)
RELATORIO_ENVIADO_HOJE = False # <--- ADICIONE ISSO

REQUISICOES_HOJE = 0
DATA_ULTIMO_RESET = datetime.now().date()
ERROS_CONSECUTIVOS = 0  # Para backoff exponencial

driver = None
LOJA_COORDS = None
LAST_WHATSAPP_REFRESH = 0
LAST_REFRESH_1 = 0
LAST_REFRESH_2 = 0
LAST_CHROME_RESTART = 0

# Carrega configurações de proteção do config
PROTECAO = CONFIG.get('protecao', {})
REFRESH_INTERVAL_1 = PROTECAO.get('refresh_interval_1', 7200)
REFRESH_INTERVAL_2 = PROTECAO.get('refresh_interval_2', 1200)
CHROME_RESTART_COOLDOWN = PROTECAO.get('chrome_restart_cooldown', 300)
# REMOVA A LINHA ABAIXO:
# LIMITE_REQUISICOES_DIA = PROTECAO.get('limite_requisicoes_dia', 3000)
# ================= TELEGRAM BOT (ADICIONADO) =================
TELEGRAM_TOKEN = CONFIG.get('telegram_token', '')
TELEGRAM_CHAT_ID = CONFIG.get('telegram_chat_id', '')
LAST_UPDATE_ID = 0
ARQUIVO_COMANDO_TELEGRAM = 'telegram_command.txt'

def enviar_telegram(mensagem):
    """Envia mensagem para o Telegram usando o token carregado."""
    if not TELEGRAM_TOKEN or not TELEGRAM_CHAT_ID:
        return
    try:
        url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage"
        payload = {
            "chat_id": TELEGRAM_CHAT_ID,
            "text": mensagem,
            "parse_mode": "Markdown"
        }
        cffi_requests.post(url, json=payload, timeout=20)
    except Exception as e:
        print(f"⚠️ Erro Telegram: {e}")

def enviar_foto_telegram(caminho_imagem, legenda=""):
    """Envia uma foto para o Telegram."""
    if not TELEGRAM_TOKEN or not TELEGRAM_CHAT_ID:
        return False
    try:
        url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendPhoto"
        with open(caminho_imagem, 'rb') as foto:
            files = {'photo': foto}
            data = {'chat_id': TELEGRAM_CHAT_ID, 'caption': legenda}
            import requests
            requests.post(url, files=files, data=data, timeout=30)
        return True
    except Exception as e:
        print(f"⚠️ Erro ao enviar foto Telegram: {e}")
        return False

import threading
def telegram_polling():
    global LAST_UPDATE_ID
    if not TELEGRAM_TOKEN:
        print("Telegram desativado (sem token)")
        return
    print("[Telegram] Polling iniciado e escutando comandos...")
    url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/getUpdates"
    while True:
        try:
            params = {"timeout": 30, "offset": LAST_UPDATE_ID + 1}
            resp = cffi_requests.get(url, params=params, timeout=35)
            data = resp.json()
            if not data.get("ok"): continue
            
            for update in data.get("result", []):
                LAST_UPDATE_ID = update["update_id"]
                msg = update.get("message", {})
                text = msg.get("text", "")
                chat_id = str(msg.get("chat", {}).get("id", ""))
                
                if chat_id != TELEGRAM_CHAT_ID: continue
                
                # Mantém o texto original para não estragar maiúsculas/minúsculas dos nomes
                cmd_original = text.strip()
                cmd = cmd_original.lower()
                
                if cmd in ["/status", "status"]:
                    enviar_telegram("🤖 Robô ativo! Use /ajuda para comandos.")
                elif cmd in ["/ajuda", "ajuda", "/help"]:
                    enviar_telegram("Comandos disponíveis:\n/status - status\n/parar - parar robô\n/reiniciar - reiniciar robô\n/gerar_excel - gerar planilha\n/gerar_resumo - resumo do dia\n🖨️ *Impressões:*\n`IMPRIMIR:nome_motoboy`\n`GERAR_GARANTIA:dados_com_barras_verticais`")
                elif cmd in ["/parar", "parar"]:
                    with open(ARQUIVO_COMANDO_TELEGRAM, 'w', encoding='utf-8') as f:
                        f.write('PARAR')
                    enviar_telegram("🛑 Parando robô...")
                elif cmd in ["/reiniciar", "reiniciar"]:
                    with open(ARQUIVO_COMANDO_TELEGRAM, 'w', encoding='utf-8') as f:
                        f.write('REINICIAR')
                    enviar_telegram("🔄 Reiniciando robô...")
                elif cmd in ["/gerar_excel", "gerar_excel"]:
                    with open(ARQUIVO_COMANDO_TELEGRAM, 'w', encoding='utf-8') as f:
                        f.write('GERAR_EXCEL')
                    enviar_telegram("⏳ Gerando Excel...")
                elif cmd in ["/gerar_resumo", "gerar_resumo"]:
                    with open(ARQUIVO_COMANDO_TELEGRAM, 'w', encoding='utf-8') as f:
                        f.write('GERAR_RESUMO')
                    enviar_telegram("⏳ Gerando resumo...")

                    
                # --- NOVA ROTA PARA AS IMPRESSÕES E GARANTIAS ---
                elif cmd_original.startswith("IMPRIMIR:") or cmd_original.startswith("IMPRIMIR_PEDIDO:") or cmd_original.startswith("GERAR_GARANTIA:") or cmd_original.startswith("IMPRIMIR_GARANTIA:"):
                    with open(ARQUIVO_COMANDO_TELEGRAM, 'w', encoding='utf-8') as f:
                        f.write(cmd_original) # Escreve o comando exatamente como o usuário mandou (com as letras maiúsculas/minúsculas corretas)
                    # Não precisa enviar msg aqui, pois a função final já vai avisar se deu certo ou errado

        except Exception as e:
            print(f"[Telegram] Erro polling: {e}")
            
        # O sleep fica alinhado com o try/except, dentro do while
        time.sleep(2)

# ==================================================================================
# Inicia polling do Telegram em thread separada (FORA DA FUNÇÃO, encostado na esquerda)
# Certifique-se de que essa parte NÃO tem o símbolo "#" na frente do if e do threading
# ==================================================================================
#if TELEGRAM_TOKEN:
 #threading.Thread(target=telegram_polling, daemon=True).start()
# ==================================================================================
#  SEÇÃO 3: TELEGRAM BOT - INICIALIZAÇÃO
# ==================================================================================
# Responsável por: Enviar mensagens para o Telegram usando o token do bot.
# Integra com a API do Telegram para comunicação bidirecional em tempo real.
# ==================================================================================

# ================= NOVAS FUNÇÕES DE SEGURANÇA (REQ + JITTER) =================

def esperar_humano(min_s=2, max_s=4):
    """Cria um atraso aleatório para simular comportamento humano (Jitter)."""
    tempo = random.uniform(min_s, max_s)
    time.sleep(tempo)

def _formatar_erro_requisicao(exc, url):
    """Gera uma mensagem curta e util para logs de falha de rede."""
    host = urlparse(url).hostname or "host_desconhecido"
    texto = str(exc)
    texto_lower = texto.lower()

    if "could not resolve host" in texto_lower or "name or service not known" in texto_lower:
        return f"DNS: nao foi possivel resolver o host {host}. Verifique internet/DNS/proxy/VPN."
    if "timed out" in texto_lower or "timeout" in texto_lower:
        return f"Timeout ao conectar em {host}. Verifique latencia, firewall ou indisponibilidade."
    if "ssl" in texto_lower or "certificate" in texto_lower:
        return f"Falha SSL ao conectar em {host}. Verifique certificado/rede interceptada."
    if "connection" in texto_lower and "refused" in texto_lower:
        return f"Conexao recusada por {host}. Servico pode estar fora do ar."

    return f"Falha de rede ao acessar {host}. Detalhes: {texto}"

def _resumir_payload(payload):
    """Extrai um resumo curto do payload GraphQL para logs."""
    if not isinstance(payload, dict):
        return "payload_desconhecido"

    operation = payload.get("operationName")
    if not operation:
        query = payload.get("query") or ""
        match = re.search(r"\b(query|mutation)\s+(\w+)", query)
        if match:
            operation = match.group(2)

    if not operation:
        operation = "operacao_desconhecida"

    variables = payload.get("variables")
    if isinstance(variables, dict) and variables:
        chaves = ", ".join(sorted(variables.keys()))
        return f"{operation} (vars: {chaves})"

    return operation

# ==================================================================================
# ==================================================================================
# Responsável por: Fazer requisições HTTP seguras à API do Zé Delivery com
# proteção contra detecção (User-Agent, delays aleatórios, tratamento de erros).
# ==================================================================================

def requisicao_segura(payload, ignorar_limite=False):
    global TOKEN_ATUAL, REQUISICOES_HOJE, DATA_ULTIMO_RESET, ERROS_CONSECUTIVOS, driver
    
    # 1. RESET CONTADOR DIÁRIO
    hoje = datetime.now().date()
    if hoje != DATA_ULTIMO_RESET:
        REQUISICOES_HOJE = 0
        DATA_ULTIMO_RESET = hoje
        print(f"\n🔄 Contador de requisições resetado ({hoje})")
    
    # 2. VERIFICAR LIMITE DIÁRIO (REINÍCIO COMPLETO)
    if REQUISICOES_HOJE >= 3000 and not ignorar_limite:
        print(f"\n⚠️ LIMITE DE 3000 REQUISIÇÕES ATINGIDO. REINICIANDO ROBÔ...")
        REQUISICOES_HOJE = 0 
        reiniciar_robo()
        return None  
    # 3. PAUSA LEVE EM HORÁRIO DE MADRUGADA (2h-6h)
    hora_atual = datetime.now().hour
    if 2 <= hora_atual < 6 and not ignorar_limite:
        esperar_humano(3, 6)  # Pausa leve em vez de 30s
    
    # 4. BACKOFF EXPONENCIAL EM CASO DE ERROS
    if ERROS_CONSECUTIVOS > 0 and not ignorar_limite:
        backoff = min(2 ** ERROS_CONSECUTIVOS, 60)  # Max 60s
        print(f"⏳ Backoff: {backoff}s (erros: {ERROS_CONSECUTIVOS})")
        time.sleep(backoff)
    
    # 5. DELAY ALEATÓRIO (comportamento humano - ignorado se for urgente do zap)
    if not ignorar_limite:
        esperar_humano(1, 3)
    
    # 6. HEADERS COMPLETOS E REALISTAS
    headers_completos = CONFIG.get('headers_api', {})

    try:
        # Parâmetros de requisição da API vindos do config
        API_REQUEST = CONFIG.get('api_request', {})
        impersonate_val = API_REQUEST.get('impersonate', None)
        timeout_min = API_REQUEST.get('timeout_min', 20)
        timeout_max = API_REQUEST.get('timeout_max', 30)
        cookie_token_key = API_REQUEST.get('cookie_token_key', 'token')
        r = cffi_requests.post(
            URL_API,
            json=payload,
            cookies={cookie_token_key: TOKEN_ATUAL},
            timeout=random.randint(timeout_min, timeout_max),
            impersonate=impersonate_val,
            headers=headers_completos
        )
        
        REQUISICOES_HOJE += 1
        ERROS_CONSECUTIVOS = 0  # Reset em sucesso

        if r.status_code == 429: # Too Many Requests
            ERROS_CONSECUTIVOS += 1
            resumo = _resumir_payload(payload)
            print("\n🛑 ALERTA VERMELHO: API retornou 429.")
            print(f"   Operacao: {resumo}")
            enviar_telegram(f"API 429 em {resumo}.")
            print("🔄 Reiniciando Chrome devido a status 429...")
            _reiniciar_chrome_se_preciso("API 429 - Reiniciando Chrome")
            return None
        
        if r.status_code == 403: # Forbidden
            ERROS_CONSECUTIVOS += 1
            resumo = _resumir_payload(payload)
            print("\n🛑 ALERTA: Erro 403 (Proibido).")
            print(f"   Operacao: {resumo}")
            enviar_telegram(f"API 403 em {resumo}.")
            print("⏳ Aguardando 2 minutos...")
            time.sleep(120)
            return None
        
        if r.status_code >= 500:  # Erro do servidor
            ERROS_CONSECUTIVOS += 1
            resumo = _resumir_payload(payload)
            print(f"\n⚠️ Erro do servidor: {r.status_code}")
            print(f"   Operacao: {resumo}")
            enviar_telegram(f"API {r.status_code} em {resumo}.")
            return None

        return r
    except Exception as e:
        ERROS_CONSECUTIVOS += 1
        detalhe = _formatar_erro_requisicao(e, URL_API)
        resumo = _resumir_payload(payload)
        print("⚠️ Erro na requisicao segura.")
        print(f"   {detalhe}")
        print(f"   Operacao: {resumo}")
        enviar_telegram(f"Erro de rede em {resumo}. {detalhe}")
        return None# ================= FUNÇÕES DE IMPRESSÃO TÉRMICA =================
CMD_INIT = b"\x1b\x40"
CMD_CENTER = b"\x1b\x61\x01"
CMD_LEFT = b"\x1b\x61\x00"
CMD_BOLD_ON = b"\x1b\x45\x01"
CMD_BOLD_OFF = b"\x1b\x45\x00"
CMD_DOUBLE_H = b"\x1b\x21\x10"
CMD_NORMAL = b"\x1b\x21\x00"
CMD_CUT = b"\x1d\x56\x00"
CMD_TIGHT_SPACING = b"\x1b\x33\x14"
CMD_NORMAL_SPACING = b"\x1b\x32"

# ==================================================================================
#  SEÇÃO 9: IMPRESSÃO TÉRMICA E RECIBOS
# ==================================================================================
# Responsável por: Gerar e imprimir recibos, relatórios e tickets na impressora
# térmica. Inclui geração de código de barras e formatação de documentos.
# ==================================================================================

def imprimir_lote_continuo(lista_pedidos):
    if not TEM_IMPRESSORA: return
    try:
        print(f"🖨️ Imprimindo lote de {len(lista_pedidos)} tickets contínuos...")
        impressora_padrao = win32print.GetDefaultPrinter()
        hPrinter = win32print.OpenPrinter(impressora_padrao)
        try:
            hJob = win32print.StartDocPrinter(hPrinter, 1, ("Lote_Tickets_Continuo", None, "RAW"))
            try:
                win32print.StartPagePrinter(hPrinter)
                buffer_impressao = CMD_INIT + CMD_LEFT + CMD_BOLD_ON 
                for i, dados in enumerate(lista_pedidos):
                    buffer_impressao += f"PED: {dados['numero']}  |  {dados['hora']}\n".encode('cp850', errors='ignore')
                    buffer_impressao += f"CLI: {dados['cliente'][:28]}\n".encode('cp850', errors='ignore')
                    buffer_impressao += f"BAI: {dados['bairro']}\n".encode('cp850', errors='ignore')
                    if dados.get('itens'):
                        itens_fmt = dados['itens'].replace("\n", " ").replace(" | ", " ")
                        buffer_impressao += f"ITM: {itens_fmt[:40]}\n".encode('cp850', errors='ignore')
                    buffer_impressao += f"$$$: R$ {dados['valor']:.2f}\n".replace('.', ',').encode('cp850')
                    if i < len(lista_pedidos) - 1:
                        buffer_impressao += b"________________________________\n"
                    else:
                        buffer_impressao += b"\n\n\n"
                buffer_impressao += CMD_CUT
                win32print.WritePrinter(hPrinter, buffer_impressao)
                win32print.EndPagePrinter(hPrinter)
            finally:
                win32print.EndDocPrinter(hPrinter)
        finally:
            win32print.ClosePrinter(hPrinter)
        print("✅ Lote enviado.")
    except Exception as e:
        print(f"❌ Erro ao imprimir lote: {e}")

def imprimir_resumo_extrato(nome_motoboy, lista_pedidos, qtd_8, qtd_11, total_valor, data_personalizada=None, vale_total=0.0, dados_fechamento=None):
    if not TEM_IMPRESSORA: return
    try:
        print(f"🖨️ Imprimindo EXTRATO DE FECHAMENTO para {nome_motoboy}...")
        impressora_padrao = win32print.GetDefaultPrinter()
        hPrinter = win32print.OpenPrinter(impressora_padrao)
        if data_personalizada: data_print = data_personalizada
        else: data_print = datetime.now().strftime('%d/%m/%Y')

        # Comandos ESC/POS extras
        CMD_FONT_B = b"\x1b\x4d\x01"        # Fonte menor (Font B)
        CMD_FONT_A = b"\x1b\x4d\x00"        # Fonte normal (Font A)
        CMD_UNDERLINE_ON = b"\x1b\x2d\x01"  # Sublinhado
        CMD_UNDERLINE_OFF = b"\x1b\x2d\x00"
        CMD_LINE_3 = b"\x1b\x33\x0A"        # Espaçamento compacto
        CMD_LINE_NORMAL = b"\x1b\x32"

        try:
            hJob = win32print.StartDocPrinter(hPrinter, 1, (f"Extrato_{nome_motoboy}", None, "RAW"))
            try:
                win32print.StartPagePrinter(hPrinter)
                c = CMD_INIT

                # ╔══════════════════════════════╗
                # ║        CABEÇALHO             ║
                # ╚══════════════════════════════╝
                c += CMD_CENTER + CMD_LINE_3
                c += b"\n"
                c += CMD_BOLD_ON + CMD_DOUBLE_H
                c += b"FECHAMENTO\n"
                c += CMD_NORMAL
                c += CMD_BOLD_ON + b"\x1b\x21\x30"
                c += f"{nome_motoboy.upper()}\n".encode('cp850', errors='ignore')
                c += CMD_NORMAL
                c += CMD_BOLD_ON + CMD_DOUBLE_H
                c += f"{data_print[:10]}\n".encode('cp850')
                c += CMD_NORMAL
                c += b"\n"

                # ════════════════════════════════
                #         ENTREGAS
                # ════════════════════════════════
                c += CMD_CENTER + CMD_BOLD_ON
                c += b"--- ENTREGAS ---\n"
                c += CMD_BOLD_OFF
                c += CMD_LEFT
                c += CMD_BOLD_ON + b"  #PEDIDO   HORA    VALOR\n" + CMD_BOLD_OFF
                c += b"  --------- -----  ---------\n"

                for p in lista_pedidos:
                    num = str(p['numero'])[:9].ljust(9)
                    hora_raw = str(p.get('hora', ''))
                    if len(hora_raw) > 5:
                        hora_raw = hora_raw[:5]
                    hora_ped = hora_raw.ljust(5)
                    valor_ped = f"R${p['valor']:.2f}".replace('.', ',').rjust(9)
                    c += f"  {num} {hora_ped}  {valor_ped}\n".encode('cp850')

                c += b"  ________________________________\n"

                # ════════════════════════════════
                #     DETALHAMENTO FECHAMENTO
                # ════════════════════════════════
                if dados_fechamento:
                    df = dados_fechamento
                    c += CMD_LEFT
                    c += f"  Periodo: {df['hora_ini']} as {df['hora_fim']}\n".encode('cp850')
                    c += b"  --------------------------------\n"
                    
                    # --- Produção ---
                    c += CMD_BOLD_ON
                    c += b"  PRODUCAO NO PERIODO\n"
                    prod_str = f"R${df['prod_dentro']:.2f}".replace('.', ',')
                    c += f"  Valor:            {prod_str.rjust(12)}\n".encode('cp850')
                    c += CMD_BOLD_OFF

                    # --- Garantia ---
                    c += CMD_BOLD_ON
                    c += b"  GARANTIA\n"
                    c += CMD_BOLD_OFF
                    gar_str = f"R${df['gar_valor']:.2f}".replace('.', ',')
                    horas = df.get('horas', 0)
                    if not horas and df.get('hora_ini') and df.get('hora_fim'):
                        try:
                            t1 = datetime.strptime(df['hora_ini'], "%H:%M")
                            t2 = datetime.strptime(df['hora_fim'], "%H:%M")
                            horas = (t2 - t1).total_seconds() / 3600
                        except Exception:
                            horas = 0
                    c += f"  {horas:.0f}h x R$15/h\n".encode('cp850')
                    c += CMD_BOLD_ON
                    c += f"  Subtotal:         {gar_str.rjust(12)}\n".encode('cp850')
                    c += CMD_BOLD_OFF

                    # --- Resultado ---
                    c += b"  ................................\n"
                    c += CMD_BOLD_ON
                    if df['usar_garantia']:
                        c += b"  * GARANTIA > PRODUCAO\n"
                        val_usado = f"R${df['gar_valor']:.2f}".replace('.', ',')
                    else:
                        c += b"  * PRODUCAO > GARANTIA\n"
                        val_usado = f"R${df['prod_dentro']:.2f}".replace('.', ',')
                    c += f"  VALOR BASE:       {val_usado.rjust(12)}\n".encode('cp850')
                    c += CMD_BOLD_OFF

                    # --- Extras ---
                    if df['prod_fora'] > 0:
                        c += b"\n"
                        c += CMD_BOLD_ON
                        c += f"  EXTRAS (apos {df['hora_fim']})\n".encode('cp850')
                        extras_str = f"R${df['prod_fora']:.2f}".replace('.', ',')
                        c += f"  Valor:            {extras_str.rjust(12)}\n".encode('cp850')
                        c += CMD_BOLD_OFF

                else:
                    # --- Resumo simples (sem garantia) ---
                    c += CMD_LEFT
                    c += CMD_BOLD_ON
                    c += b"  PRODUCAO\n"
                    prod_total = f"R${total_valor + vale_total:.2f}".replace('.', ',')
                    c += f"  Valor:            {prod_total.rjust(12)}\n".encode('cp850')
                    c += CMD_BOLD_OFF

                # ════════════════════════════════
                #           VALES
                # ════════════════════════════════
                if vale_total and vale_total > 0:
                    c += b"\n"
                    vale_str = f"R${vale_total:.2f}".replace('.', ',')
                    c += CMD_BOLD_ON
                    c += f"  (-) VALE:         {vale_str.rjust(12)}\n".encode('cp850')
                    c += CMD_BOLD_OFF

                # ╔══════════════════════════════╗
                # ║        TOTAL FINAL           ║
                # ╚══════════════════════════════╝
                c += CMD_LEFT
                c += f"  Entregas de 8:  {qtd_8}\n".encode('cp850')
                c += f"  Entregas de 11: {qtd_11}\n".encode('cp850')
                c += b"\n"
                c += CMD_CENTER
                c += b"================================\n"
                c += CMD_BOLD_ON + CMD_DOUBLE_H
                total_str = f"R$ {total_valor:.2f}".replace('.', ',')
                c += f"TOTAL {total_str}\n".encode('cp850')
                c += CMD_NORMAL + CMD_BOLD_OFF
                c += b"================================\n"

                c += CMD_LINE_NORMAL
                c += b"\n\n\n" + CMD_CUT
                win32print.WritePrinter(hPrinter, c)
                win32print.EndPagePrinter(hPrinter)
            finally:
                win32print.EndDocPrinter(hPrinter)
        finally:
            win32print.ClosePrinter(hPrinter)
    except Exception as e:
        print(f"❌ Erro ao imprimir extrato: {e}")

def imprimir_relatorio_canceladas(lista_canceladas, data_relatorio=None):
    if not TEM_IMPRESSORA: return
    try:
        print(f"🖨️ Imprimindo RELATORIO DE CANCELAMENTOS...")
        impressora_padrao = win32print.GetDefaultPrinter()
        hPrinter = win32print.OpenPrinter(impressora_padrao)
        data_print = data_relatorio if data_relatorio else datetime.now().strftime('%d-%m-%Y')
        try:
            hJob = win32print.StartDocPrinter(hPrinter, 1, ("Relatorio_Canceladas", None, "RAW"))
            try:
                win32print.StartPagePrinter(hPrinter)
                cupom = CMD_INIT + CMD_CENTER
                cupom += CMD_BOLD_ON + CMD_DOUBLE_H + b"CANCELADOS\n" + CMD_NORMAL + CMD_BOLD_OFF
                cupom += f"Data: {data_print}\n".encode('cp850', errors='ignore')
                cupom += b"================================\n"
                
                if not lista_canceladas:
                    cupom += CMD_LEFT + b"Nenhum pedido cancelado encontrado.\n"
                else:
                    for p in lista_canceladas:
                        cupom += CMD_LEFT
                        cupom += CMD_BOLD_ON + f"PEDIDO: {p.get('numero', 'N/A')} ({p.get('hora', 'N/A')})\n".encode('cp850', errors='ignore') + CMD_BOLD_OFF
                        cupom += f"Cliente: {p.get('cliente', 'N/A')[:25]}\n".encode('cp850', errors='ignore')
                        cupom += f"Bairro:  {p.get('bairro', 'N/A')[:25]}\n".encode('cp850', errors='ignore')
                        st = p.get('status', '').replace("POC_", "").replace("USER_", "").replace("SYS_", "")
                        cupom += f"Status:  {st}\n".encode('cp850', errors='ignore')
                        cupom += b"--------------------------------\n"

                cupom += CMD_CENTER + CMD_BOLD_ON
                cupom += f"TOTAL: {len(lista_canceladas)}\n".encode('cp850', errors='ignore')
                cupom += CMD_NORMAL + b"\n\n\n" + CMD_CUT
                win32print.WritePrinter(hPrinter, cupom)
                win32print.EndPagePrinter(hPrinter)
            finally:
                win32print.EndDocPrinter(hPrinter)
        finally:
            win32print.ClosePrinter(hPrinter)
    except Exception as e:
        print(f"❌ Erro ao imprimir canceladas: {e}")
def imprimir_recibo_garantia(dados_str):
    if not TEM_IMPRESSORA: return
    try:
        # Formato esperado: NOME|QTD8|QTD11|VALOR_PROD|HORAS|VALOR_GARANTIA|TOTAL|TIPO|DESCONTO|VALE
        partes = dados_str.split('|')
        
        if len(partes) < 8: 
            print("⚠️ Dados de garantia incompletos.")
            return
            
        # Extrai os dados básicos
        nome, qtd8, qtd11, v_prod, horas, v_garantia, v_total, tipo = partes[:8]
        
        # Captura o desconto (9º item) e o vale (10º item). Se não existir, assume 0.00
        desconto = partes[8] if len(partes) > 8 else "0,00"
        vale = partes[9] if len(partes) > 9 else "0,00"
        
        print(f"🖨️ Imprimindo Recibo com Desconto: {nome}")
        impressora_padrao = win32print.GetDefaultPrinter()
        hPrinter = win32print.OpenPrinter(impressora_padrao)
        
        try:
            hJob = win32print.StartDocPrinter(hPrinter, 1, (f"Fechamento_{nome}", None, "RAW"))
            try:
                win32print.StartPagePrinter(hPrinter)
                
                # Montagem do Cupom
                cupom = CMD_INIT + CMD_CENTER
                cupom += CMD_BOLD_ON + CMD_DOUBLE_H + b"RECIBO PAGAMENTO\n" + CMD_NORMAL + CMD_BOLD_OFF
                cupom += f"Data: {datetime.now().strftime('%d/%m/%Y %H:%M')}\n".encode('cp850')
                cupom += b"================================\n"
                cupom += CMD_BOLD_ON + CMD_DOUBLE_H + f"{nome}\n".encode('cp850', errors='ignore') + CMD_NORMAL + CMD_BOLD_OFF
                cupom += b"--------------------------------\n"
                
                # Produção
                cupom += CMD_LEFT + CMD_BOLD_ON + b"PRODUCAO:\n" + CMD_BOLD_OFF
                cupom += f"Entregas R$ 8,00:  {qtd8}\n".encode('cp850')
                cupom += f"Entregas R$ 11,00: {qtd11}\n".encode('cp850')
                cupom += f"Subtotal Prod: R$ {v_prod.replace('.', ',')}\n".encode('cp850')
                cupom += b"--------------------------------\n"
                
                # Garantia
                cupom += CMD_BOLD_ON + b"GARANTIA / HORAS:\n" + CMD_BOLD_OFF
                cupom += f"Horario: {horas}\n".encode('cp850')
                cupom += f"Subtotal Gar: R$ {v_garantia.replace('.', ',')}\n".encode('cp850')
                
                # --- NOVA SEÇÃO: IMPRESSÃO DO DESCONTO/VALE ---
                val_desc_float = float(desconto.replace(",", "."))
                val_vale_float = float(vale.replace(",", "."))
                if val_desc_float > 0 or val_vale_float > 0:
                    cupom += b"--------------------------------\n"
                    if val_desc_float > 0:
                        cupom += CMD_BOLD_ON + b"(-) DESCONTO:\n" + CMD_BOLD_OFF
                        cupom += f"Valor Retido: R$ {desconto.replace('.', ',')}\n".encode('cp850')
                    if val_vale_float > 0:
                        cupom += CMD_BOLD_ON + b"(-) DESCONTO VALE:\n" + CMD_BOLD_OFF
                        cupom += f"Valor Vale: R$ {vale.replace('.', ',')}\n".encode('cp850')
                
                cupom += b"================================\n"
                
                # Total Final
                cupom += CMD_CENTER + b"VALOR A PAGAR:\n"
                cupom += CMD_BOLD_ON + CMD_DOUBLE_H
                cupom += f"R$ {v_total.replace('.', ',')}\n".encode('cp850')
                cupom += CMD_NORMAL + f"({tipo})\n".encode('cp850')
                
                cupom += CMD_BOLD_OFF + b"\n\n\n________________________________\nAssinatura do Entregador\n\n\n" + CMD_CUT
                
                win32print.WritePrinter(hPrinter, cupom)
                win32print.EndPagePrinter(hPrinter)
            finally:
                win32print.EndDocPrinter(hPrinter)
        finally:
            win32print.ClosePrinter(hPrinter)
            
    except Exception as e:
        print(f"❌ Erro ao imprimir recibo: {e}")

# ==================================================================================
#  SEÇÃO 2: UTILITÁRIOS DE ARQUIVO E CAMINHO
# ==================================================================================
# Responsável por: Operações básicas com filesystem, carregamento de credenciais,
# e inicialização da estrutura de arquivos necessários para o robô funcionar.
# ==================================================================================

def get_caminho_base():
    if getattr(sys, 'frozen', False): return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

def get_caminho_excel():
    agora = datetime.now()
    if agora.hour < 10: agora -= timedelta(days=1)
    data_str = agora.strftime('%d-%m-%Y')
    return os.path.join(get_caminho_base(), f'Controle_Financeiro_{data_str}.xlsx')

def get_caminho_excel_por_data(data_str):
    """Retorna o caminho do Excel para uma data específica ('dd-mm-yyyy')."""
    if not data_str:
        return get_caminho_excel()
    return os.path.join(get_caminho_base(), f'Controle_Financeiro_{data_str}.xlsx')

def salvar_excel_seguro(wb, arquivo):
    """Salva o Excel em arquivo temporário primeiro, depois substitui o original.
    Isso evita corrupção se o processo for interrompido durante o save."""
    import tempfile, shutil
    pasta = os.path.dirname(arquivo)
    try:
        # Salva em arquivo temporário na mesma pasta
        fd, tmp_path = tempfile.mkstemp(suffix='.xlsx', dir=pasta)
        os.close(fd)
        wb.save(tmp_path)
        # Cria backup do original se existir
        if os.path.exists(arquivo):
            backup = arquivo + '.bak'
            try:
                shutil.copy2(arquivo, backup)
            except Exception:
                pass
        # Substitui o original pelo temporário (operação atômica no Windows com replace)
        shutil.move(tmp_path, arquivo)
    except Exception as e:
        # Se falhar, tenta salvar direto como fallback
        print(f"⚠️ Save seguro falhou ({e}), tentando save direto...")
        try:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
        except Exception:
            pass
        wb.save(arquivo)

def inicializar_excel_agora():
    global pedidos_ja_enviados, CACHE_STATUS_PEDIDOS
    arquivo = get_caminho_excel()
    def criar_novo_excel():
        print(f"📊 Criando NOVO arquivo Excel do dia: {os.path.basename(arquivo)}")
        pedidos_ja_enviados.clear()
        CACHE_STATUS_PEDIDOS.clear()
        wb = openpyxl.Workbook()
        
        # === ABA 1: EXTRATO DETALHADO ===
        ws1 = wb.active
        ws1.title = "EXTRATO DETALHADO"
        ws1.append(['Data', 'Hora', 'Numero', 'Cliente', 'Bairro', 'Status', 'Motoboy', 'Combo', 'Valor (R$)', 'Itens'])
        
        # Formatação Header Extrato
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, size=11, color="FFFFFF")
        for cell in ws1[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Largura das colunas
        ws1.column_dimensions['A'].width = 12  # Data
        ws1.column_dimensions['B'].width = 8   # Hora
        ws1.column_dimensions['C'].width = 12  # Numero
        ws1.column_dimensions['D'].width = 25  # Cliente
        ws1.column_dimensions['E'].width = 20  # Bairro
        ws1.column_dimensions['F'].width = 15  # Status
        ws1.column_dimensions['G'].width = 20  # Motoboy
        ws1.column_dimensions['H'].width = 10  # Combo
        ws1.column_dimensions['I'].width = 12  # Valor
        ws1.column_dimensions['J'].width = 40  # Itens
        
        ws1.freeze_panes = 'A2'  # Congela header
        
        # === ABA 2: PAGAMENTO MOTOBOYS ===
        ws2 = wb.create_sheet("PAGAMENTO_MOTOBOYS")
        ws2.append(["MOTOBOY", "QTD TOTAL", "QTD R$ 8,00", "QTD R$ 11,00", "TOTAL A PAGAR (R$)"])
        
        # Formatação Header Pagamentos
        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Largura das colunas
        ws2.column_dimensions['A'].width = 25
        ws2.column_dimensions['B'].width = 12
        ws2.column_dimensions['C'].width = 12
        ws2.column_dimensions['D'].width = 12
        ws2.column_dimensions['E'].width = 18
        
        ws2.freeze_panes = 'A2'
        
        try:
            wb.save(arquivo)
            print(f"✅ Arquivo salvo: {arquivo}")
        except Exception as e:
            print(f"❌ Erro ao criar Excel: {e}")
    if os.path.exists(arquivo):
        try:
            wb = openpyxl.load_workbook(arquivo)
            ws1 = wb["EXTRATO DETALHADO"]
            for row in ws1.iter_rows(min_row=2, values_only=True):
                if row and row[2]:
                    numero = str(row[2]).strip()
                    status_excel = str(row[5]).upper() if row[5] else ""
                    CACHE_STATUS_PEDIDOS[numero] = status_excel
                    if "POC_ACCEPTED" not in status_excel:
                        pedidos_ja_enviados.add(numero)
        except Exception as e:
            print(f"⚠️ ARQUIVO CORROMPIDO: {e}")
            try: os.remove(arquivo)
            except: pass
            criar_novo_excel()
    else:
        criar_novo_excel()
def registrar_vale(nome_moto, valor, motivo="Desconto/Vale"):
    """
    Grava um registro de vale na aba 'VALES' do Excel.
    Essa função é essencial para que o robô entenda o comando vindo do painel.
    """
    arquivo = get_caminho_excel()
    if not os.path.exists(arquivo): inicializar_excel_agora()
    
    try:
        wb = openpyxl.load_workbook(arquivo)
        
        # Cria a aba VALES se ela não existir na planilha do dia
        if "VALES" not in wb.sheetnames:
            ws = wb.create_sheet("VALES")
            ws.append(["Hora", "Motoboy", "Valor", "Motivo"])
            
            # Formatação profissional do header
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, size=11, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Largura das colunas
            ws.column_dimensions['A'].width = 10  # Hora
            ws.column_dimensions['B'].width = 25  # Motoboy
            ws.column_dimensions['C'].width = 12  # Valor
            ws.column_dimensions['D'].width = 30  # Motivo
            
            ws.freeze_panes = 'A2'
        else:
            ws = wb["VALES"]
            
        hora_atual = datetime.now().strftime('%H:%M')
        # Adiciona a nova linha de vale
        ws.append([hora_atual, nome_moto, float(valor), motivo])
        
        wb.save(arquivo)
        print(f"💾 Registro de Vale salvo: {nome_moto} - R$ {valor}")
        return True
    except Exception as e:
        print(f"❌ Erro crítico ao registrar vale no Excel: {e}")
        return False

# ==================================================================================
#  SEÇÃO 10: RELATÓRIOS E ANÁLISE
# ==================================================================================
# Responsável por: Geração de relatórios executivos, análise de dados
# e cálculo de métricas para fechamento diário.
# ==================================================================================

def gerar_relatorio_executivo():
    arquivo = get_caminho_excel()
    if not os.path.exists(arquivo):
        return "❌ Sem planilha hoje."

    try:
        wb = openpyxl.load_workbook(arquivo, data_only=True)
        ws = wb["EXTRATO DETALHADO"]

        total_venda = 0.0
        qtd_entregas = 0
        qtd_retiradas = 0
        pagamentos = {}
        contagem_produtos = {}

        # 1. Processa Entregas e Produtos do Extrato
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[5]:
                continue
            st = str(row[5]).upper()

            # Ignora cancelados
            if any(x in st for x in STATUS_CANCELADOS_LISTA):
                continue

            nome_moto = str(row[6]) if row[6] else "Desconhecido"
            val = float(row[8]) if row[8] else 0.0

            # Contagem de Produtos (Coluna 10)
            itens_str = str(row[9]) if len(row) > 9 and row[9] else ""
            if itens_str:
                partes = itens_str.split(',')
                for p in partes:
                    if "x " in p:
                        try:
                            qtd_item, nome_item = p.split("x ", 1)
                            nome_item = nome_item.strip().upper()
                            if nome_item not in contagem_produtos:
                                contagem_produtos[nome_item] = 0
                            contagem_produtos[nome_item] += int(qtd_item)
                        except:
                            pass

            if val > 0:
                total_venda += val
                qtd_entregas += 1
                if nome_moto not in pagamentos:
                    pagamentos[nome_moto] = {'qtd': 0, 'bruto': 0.0, 'vales': 0.0}
                pagamentos[nome_moto]['qtd'] += 1
                pagamentos[nome_moto]['bruto'] += val
            else:
                qtd_retiradas += 1

        # 2. Processa os Vales para desconto (Aba VALES)
        if "VALES" in wb.sheetnames:
            ws_vales = wb["VALES"]
            for row in ws_vales.iter_rows(min_row=2, values_only=True):
                if not row or len(row) < 3 or not row[1]:
                    continue
                nome_vale = normalizar_texto(str(row[1]))
                
                # Converte valor com validação robusta
                try:
                    valor_vale = float(row[2]) if row[2] else 0.0
                except (ValueError, TypeError):
                    continue  # Ignora linhas com valores inválidos

                encontrou = False
                for nome_pag in pagamentos:
                    if nome_vale in normalizar_texto(nome_pag):
                        pagamentos[nome_pag]['vales'] += valor_vale
                        encontrou = True
                        break

                if not encontrou:
                    nome_real = str(row[1]).capitalize()
                    if nome_real not in pagamentos:
                        pagamentos[nome_real] = {'qtd': 0, 'bruto': 0.0, 'vales': 0.0}
                    pagamentos[nome_real]['vales'] += valor_vale

        # 3. Montagem da Mensagem Final
        msg = "📊 *FECHAMENTO DO DIA*\n"
        msg += f"📅 Data: {datetime.now().strftime('%d/%m/%Y')}\n"
        msg += "━━━━━━━━━━━━━━━━━━\n"
        msg += f"💰 *Total Corridas: R$ {total_venda:.2f}*\n"
        msg += f"🛵 Entregas: {qtd_entregas} | 🛒 Retiradas: {qtd_retiradas}\n"
        msg += "━━━━━━━━━━━━━━━━━━\n"

        # Pagamentos por Motoboy
        total_liquido_geral = 0.0
        for nome, dados in pagamentos.items():
            if "RETIRADA" in nome.upper():
                continue
            bruto = dados['bruto']
            vale = dados['vales']
            liquido = bruto - vale
            total_liquido_geral += liquido

            msg += f"👤 *{nome}*\n"
            msg += f"   Produção: R$ {bruto:.2f}\n"
            if vale > 0:
                msg += f"   🔻 Vale: -R$ {vale:.2f}\n"
            msg += f"   💰 *Liquido: R$ {liquido:.2f}*\n"
            msg += "   ----------------\n"

        msg += f"\n🏆 *TOTAL A PAGAR: R$ {total_liquido_geral:.2f}*"
        return msg

    except Exception as e:
        return f"❌ Erro ao gerar relatório: {e}"
def salvar_no_excel(dados_pedido):
    global CACHE_STATUS_PEDIDOS
    
    # FILTRO: Não salva se motoboy for "Desconhecido" ou "Aguardando..."
    motoboy = dados_pedido.get('motoboy', '').strip()
    if motoboy in ["Desconhecido", "Aguardando..."]:
        return  # Ignora completamente esse pedido
    
    arquivo = get_caminho_excel()
    if not os.path.exists(arquivo): inicializar_excel_agora()
    
    # Definição de Cores e Estilos
    VERMELHO_CLARO = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    VERDE_CLARO = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    AMARELO_ALERTA = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    FONTE_VERMELHA = Font(color="990000", bold=True)
    FONTE_VERDE = Font(color="006600", bold=False)
    
    try:
        wb = openpyxl.load_workbook(arquivo)
        ws1 = wb["EXTRATO DETALHADO"]
        if "PAGAMENTO_MOTOBOYS" not in wb.sheetnames: wb.create_sheet("PAGAMENTO_MOTOBOYS")
        ws2 = wb["PAGAMENTO_MOTOBOYS"]
        
        # PROTEÇÃO: Garantir que PAGAMENTO_MOTOBOYS sempre tem headers
        if ws2.max_row == 0 or not ws2.cell(row=1, column=1).value:
            ws2.cell(row=1, column=1).value = "MOTOBOY"
            ws2.cell(row=1, column=2).value = "QTD TOTAL"
            ws2.cell(row=1, column=3).value = "QTD R$ 8,00"
            ws2.cell(row=1, column=4).value = "QTD R$ 11,00"
            ws2.cell(row=1, column=5).value = "TOTAL A PAGAR (R$)"
            for cell in ws2[1]: 
                cell.font = Font(bold=True, size=11)
        
        # --- PARTE 1: REGISTRO NO EXTRATO DETALHADO ---
        dt = parse_data_pedido(dados_pedido.get('data_pedido')) or datetime.now()
        numero = str(dados_pedido.get('numero', '')).strip()
        status_novo = str(dados_pedido.get('status', '')).upper()
        
        if numero in CACHE_STATUS_PEDIDOS and CACHE_STATUS_PEDIDOS[numero] == status_novo: return
        CACHE_STATUS_PEDIDOS[numero] = status_novo
        
        try: valor_float = float(dados_pedido.get('valor', 0.0))
        except: valor_float = 0.0
        
        eh_cancelado = any(termo in status_novo for termo in STATUS_CANCELADOS_LISTA)
        if eh_cancelado: valor_float = 0.0
        
        eh_valor_padrao = dados_pedido.get('valor_padrao_usado', False)
        combo_info = dados_pedido.get('combo', '')
        itens_str = dados_pedido.get('itens', '')
        motoboy_atual = dados_pedido.get('motoboy', '')

        linha_existente = None
        for r in range(2, ws1.max_row + 1):
            cell_num = ws1.cell(row=r, column=3).value
            if cell_num and str(cell_num).strip() == numero:
                linha_existente = r; break
                
        if linha_existente:
            ws1.cell(row=linha_existente, column=1).value = dt.strftime('%d/%m/%Y')
            ws1.cell(row=linha_existente, column=2).value = dt.strftime('%H:%M')
            ws1.cell(row=linha_existente, column=6).value = status_novo
            ws1.cell(row=linha_existente, column=7).value = motoboy_atual
            ws1.cell(row=linha_existente, column=8).value = combo_info
            ws1.cell(row=linha_existente, column=9).value = valor_float
            if itens_str: ws1.cell(row=linha_existente, column=10).value = itens_str
            
            for col in range(1, 11):
                cel = ws1.cell(row=linha_existente, column=col)
                if eh_cancelado: cel.fill = VERMELHO_CLARO; cel.font = FONTE_VERMELHA
                elif eh_valor_padrao: cel.fill = AMARELO_ALERTA
                else: cel.fill = VERDE_CLARO; cel.font = FONTE_VERDE
        else:
            ws1.append([dt.strftime('%d/%m/%Y'), dt.strftime('%H:%M'), numero, dados_pedido.get('cliente', ''), dados_pedido.get('bairro', ''), status_novo, motoboy_atual, combo_info, valor_float, itens_str])
            for col in range(1, 11):
                cel = ws1.cell(row=ws1.max_row, column=col)
                if eh_cancelado: cel.fill = VERMELHO_CLARO; cel.font = FONTE_VERMELHA
                elif eh_valor_padrao: cel.fill = AMARELO_ALERTA
                else: cel.fill = VERDE_CLARO; cel.font = FONTE_VERDE

        # --- PARTE 2: RECALCULAR PAGAMENTOS (COM DESCONTO DE VALES) ---
        ws2.delete_rows(2, ws2.max_row + 1)  # Deleta dados, mantém header na row 1

        resumo = {}

        # Soma produção do Extrato
        for row in ws1.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 7 or not row[6]: continue
            st = str(row[5]).upper()
            if any(termo in st for termo in STATUS_CANCELADOS_LISTA): continue
            
            nm = str(row[6])
            val = float(row[8]) if row[8] else 0.0
            
            if nm not in resumo: 
                resumo[nm] = {'qtd': 0, 'qtd_8': 0, 'qtd_11': 0, 'valor': 0.0, 'vales': 0.0}
            
            resumo[nm]['qtd'] += 1
            resumo[nm]['valor'] += val
            if abs(val - 8.0) < 0.1: resumo[nm]['qtd_8'] += 1
            elif abs(val - 11.0) < 0.1: resumo[nm]['qtd_11'] += 1

        # Subtrai Vales registrados na aba "VALES"
        if "VALES" in wb.sheetnames:
            ws_vales = wb["VALES"]
            for row in ws_vales.iter_rows(min_row=2, values_only=True):
                if not row or len(row) < 3 or not row[1]: continue
                moto_v = normalizar_texto(str(row[1]))
                
                # Converte valor com validação robusta
                try:
                    valor_v = float(row[2]) if row[2] else 0.0
                except (ValueError, TypeError):
                    continue  # Ignora linhas com valores inválidos
                
                for k in resumo:
                    if moto_v in normalizar_texto(k):
                        resumo[k]['vales'] += valor_v
                        break

        # Escreve os resultados finais na aba de pagamentos
        row_start = 2  # Começa após o header
        for nome, d in resumo.items():
            # Ignora "RETIRADA", "Desconhecido" e "Aguardando..."
            if "RETIRADA" in nome.upper() or nome.strip() in ["Desconhecido", "Aguardando..."]:
                continue
            liquido = d['valor'] - d['vales']
            ws2.append([nome, d['qtd'], d['qtd_8'], d['qtd_11'], liquido])
        
        # === FORMATAÇÃO PROFISSIONAL DA TABELA ===
        from openpyxl.styles import Border, Side, numbers
        
        # Bordas
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # Aplica bordas e formatação
        for row in ws2.iter_rows(min_row=1, max_row=ws2.max_row, min_col=1, max_col=5):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Formata a coluna de valores como moeda
        for row in range(2, ws2.max_row + 1):
            ws2.cell(row=row, column=5).number_format = 'R$ #,##0.00'
        
        # Ajusta largura das colunas
        ws2.column_dimensions['A'].width = 25  # Nome
        ws2.column_dimensions['B'].width = 12  # QTD Total
        ws2.column_dimensions['C'].width = 12  # QTD 8
        ws2.column_dimensions['D'].width = 12  # QTD 11
        ws2.column_dimensions['E'].width = 18  # Valor
        
        # Congela primeira linha (header)
        ws2.freeze_panes = 'A2'
        
        # Aplica cor no header
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, size=11, color="FFFFFF")
        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # --- SALVAMENTO E FECHAMENTO DO BLOCO ---
        wb.save(arquivo)
        
    except Exception as e:
        print(f"❌ Erro ao salvar dados no Excel: {e}")

# ==================================================================================
#  SEÇÃO 12: NORMALIZAÇÃO E PROCESSAMENTO DE TEXTO
# ==================================================================================
# Responsável por: Limpeza, padronização e formatação de strings de dados
# provenientes da API (nomes, bairros, datas, itens, etc).
# ==================================================================================

def normalizar_texto(texto):
    if not texto: return ""
    try:
        nfkd = unicodedata.normalize('NFKD', texto)
        # Converte para minúsculas e remove acentos
        t = "".join([c for c in nfkd if not unicodedata.combining(c)]).lower().strip()
        # Remove APENAS palavras completas (com word boundaries), não letras isoladas
        palavras_remover = ["bairro", "loteamento", "residencial", "condominio", "pq"]
        for p in palavras_remover:
            # Remove a palavra apenas se for uma palavra completa, não parte de outra
            t = re.sub(r'\b' + p + r'\b', '', t)
        # Remove números e caracteres especiais para padronizar
        t = re.sub(r'\d+', '', t)  # Remove números
        t = re.sub(r'[^\w\s]', '', t)  # Remove caracteres especiais
        return t.strip()
    except: return texto.lower().strip()

def parse_data_pedido(data_str):
    if not data_str:
        return None
    if isinstance(data_str, datetime):
        return data_str
    try:
        s = str(data_str).strip()
        if s.endswith("Z"):
            s = s.replace("Z", "+00:00")
        
        # Parse com timezone info se tiver
        dt = datetime.fromisoformat(s)
        
        # Se tem timezone info (aware), converter de UTC para São Paulo (UTC-3)
        if dt.tzinfo is not None:
            # Subtrair 3 horas para converter de UTC para São Paulo
            from datetime import timezone, timedelta
            utc_tz = timezone.utc
            sp_tz = timezone(timedelta(hours=-3))
            
            # Converter para UTC primeiro, depois para São Paulo
            dt_utc = dt.astimezone(utc_tz)
            dt_sp = dt_utc.astimezone(sp_tz)
            
            # Retorna naive datetime (sem tz info) para ser salvo no Excel
            return dt_sp.replace(tzinfo=None)
        else:
            # Se não tem timezone, assumir que é UTC e converter para São Paulo
            from datetime import timezone, timedelta
            utc_tz = timezone.utc
            dt_with_tz = dt.replace(tzinfo=utc_tz)
            sp_tz = timezone(timedelta(hours=-3))
            dt_sp = dt_with_tz.astimezone(sp_tz)
            return dt_sp.replace(tzinfo=None)
    except Exception as e:
        try:
            return datetime.strptime(str(data_str).strip(), "%Y-%m-%dT%H:%M:%S.%f")
        except Exception:
            return None

def normalizar_bairro(texto):
    t = normalizar_texto(texto)
    t = re.sub(r'[^a-z0-9\s]', ' ', t)
    t = re.sub(r'\s+', ' ', t).strip()
    return t

def limpar_texto_busca(texto):
    if not texto: return ""
    try:
        nfkd = unicodedata.normalize('NFKD', texto)
        # Converte para minúsculas e remove acentos
        t = "".join([c for c in nfkd if not unicodedata.combining(c)]).lower().strip()
        # Remove APENAS palavras completas (com word boundaries), não letras isoladas
        palavras_remover = ["imprimir", "relatorio", "telefone", "numero", "contato", "celular", "cliente", "canceladas"]
        for p in palavras_remover:
            # Remove a palavra apenas se for uma palavra completa, não parte de outra
            t = re.sub(r'\b' + p + r'\b', '', t)
        # Remove emails, números e caracteres especiais
        t = t.replace("@gmail.com", "").replace("@hotmail.com", "")
        t = re.sub(r'\d+', '', t)  # Remove números
        t = re.sub(r'[^\w\s]', '', t)  # Remove caracteres especiais
        return t.strip()
    except: return texto.strip()

def calcular_valor_entrega(bairro_api):
    if not bairro_api or bairro_api == "Não disp.": 
        return 8.00, True
    
    b = normalizar_bairro(bairro_api)
    
    # Busca o bairro no dicionário unificado
    for bairro_cadastrado, valor in BAIRROS_VALORES.items():
        if normalizar_bairro(bairro_cadastrado) in b:
            return valor, False
    
    # Se não encontrar, retorna valor padrão
    if b not in BAIRROS_NAO_CADASTRADOS_LOGADOS:
        BAIRROS_NAO_CADASTRADOS_LOGADOS.add(b)
        print(f"⚠️ BAIRRO NÃO CADASTRADO: '{bairro_api}' (normalizado: '{b}')")
    return 8.00, True

def identificar_motoboy(email):
    if not email: return "Desconhecido"
    return MOTOBOYS_API.get(email.lower().strip(), email.split('@')[0].capitalize())

def formatar_itens_para_string(lista_produtos):
    if not lista_produtos: return ""
    try:
        itens = []
        for prod in lista_produtos:
            nome = prod.get('name', 'Item')
            qtd = prod.get('amount', 1)
            itens.append(f"{qtd}x {nome}")
        return ", ".join(itens)
    except: return ""

# ==================================================================================
#  SEÇÃO 11: GEOLOCALIZAÇÃO E GEOPROCESSAMENTO
# ==================================================================================
# Responsável por: Cálculos de distância entre coordenadas GPS, identificação
# de bairro pela localização, e determinação de valor de entrega correto.
# ==================================================================================

def calcular_distancia_real_km(lat1, lon1, lat2, lon2):
    try:
        R = 6371; dLat = math.radians(lat2 - lat1); dLon = math.radians(lon2 - lon1)
        a = math.sin(dLat/2)**2 + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) * math.sin(dLon/2)**2
        return R * (2 * math.atan2(math.sqrt(a), math.sqrt(1-a)))
    except: return 99.0

def calcular_direcao_gps(lat_dest, lng_dest):
    if not LOJA_COORDS: return None
    lat_loja, lng_loja = LOJA_COORDS
    rlat_loja = math.radians(lat_loja); rlng_loja = math.radians(lng_loja)
    rlat_dest = math.radians(lat_dest); rlng_dest = math.radians(lng_dest)
    dLon = (rlng_dest - rlng_loja)
    y = math.sin(dLon) * math.cos(rlat_dest)
    x = math.cos(rlat_loja) * math.sin(rlat_dest) - math.sin(rlat_loja) * math.cos(rlat_dest) * math.cos(dLon)
    ang = (math.degrees(math.atan2(y, x)) + 360) % 360
    if 315 <= ang or ang < 45: zona = "NORTE"
    elif 45 <= ang < 135: zona = "LESTE"
    elif 135 <= ang < 225: zona = "SUL"
    else: zona = "OESTE"
    return {"erro": False, "zona": zona, "angulo": ang, "lat": lat_dest, "lng": lng_dest, "msg": f" (🧭 ZONA {zona})"}

# --- FUNÇÃO DE GPS (RESTAURADA!) ---
def preparar_gps_loja():
    global LOJA_COORDS
    if not TEM_GPS: return
    try:
        g = geocoder.arcgis(ENDERECO_LOJA); LOJA_COORDS = g.latlng if g.ok else [-27.1000, -52.6000]
        print(f"🌍 GPS Loja: {LOJA_COORDS}")
    except: pass

def fazer_barulho():
    try: winsound.Beep(1000, 300)
    except: pass

def reiniciar_robo():
    """Reinicia o processo do robô."""
    print("🤖 COMANDO DE REINICIO RECEBIDO. Reiniciando em 3 segundos...")
    enviar_telegram("🔄 *Reiniciando o robô...*\nO processo será reiniciado em 3 segundos.")
    time.sleep(3)
    try:
        if driver:
            try:
                print("Tentando fechar aba ativa do Chrome...")
                driver.close()
            except Exception as e:
                print(f"Aviso: erro ao fechar aba ativa: {e}")
            try:
                print("Tentando encerrar sessão do Chrome (driver.quit)...")
                driver.quit()
                print("driver.quit() executado com sucesso.")
            except Exception as e:
                print(f"Aviso: erro ao encerrar sessão do Chrome: {e}")
        driver = None
        
        # Fallback agressivo: se após quit ainda existem chrome na porta 9222, mata via taskkill
        try:
            import subprocess
            subprocess.run(
                ['taskkill', '/F', '/IM', 'chromedriver.exe'],
                capture_output=True, timeout=5
            )
        except Exception:
            pass
        
        iniciar_chrome_persistente()        
        # Verifica se o restarlmente funcionou
        if not _driver_ativo():
            print("⚠️ Chrome reiniciou mas driver não está ativo. Próximo ciclo tentará novamente.")
    except Exception as e:
        print(f"❌ Falha ao reiniciar Chrome: {e}")

def refresh_main_periodically():
    global LAST_MAIN_REFRESH, driver
    if not driver:
        return

    agora = time.time()
    if agora - LAST_MAIN_REFRESH < REFRESH_INTERVAL_2:
        return

    try:
        _recarregar_pagina_principal("manutencao")
    except Exception as e:
        print(f"⚠️ Erro ao atualizar página principal: {e}")
    finally:
        LAST_MAIN_REFRESH = agora

# ================= FUNÇÕES DE PROCESSAMENTO E RELATÓRIOS (ADICIONADAS) =================

def processar_relatorio_canceladas(data_filtro=None):
    lista = []
    arquivo = get_caminho_excel_por_data(data_filtro)
    if not os.path.exists(arquivo):
        return f"Planilha para data {data_filtro} nao encontrada."
    try:
        wb = openpyxl.load_workbook(arquivo, data_only=True)
        ws = wb["EXTRATO DETALHADO"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[5]:
                st = str(row[5]).upper()
                if any(x in st for x in STATUS_CANCELADOS_LISTA):
                    lista.append({
                        'hora': str(row[1]),
                        'numero': str(row[2]),
                        'cliente': str(row[3]),
                        'bairro': str(row[4]),
                        'status': st
                    })
        imprimir_relatorio_canceladas(lista, data_filtro)
        return f"Relatorio de Canceladas gerado: {len(lista)} pedidos."
    except Exception as e:
        return f"Erro ao gerar relatorio de canceladas: {e}"

def processar_impressao_individual(texto):
    nome = limpar_texto_busca(texto.replace("imprimir", ""))
    pedidos = buscar_todos_pedidos_excel_por_nome(nome)
    if pedidos:
        imprimir_lote_continuo(pedidos)
        return f"🖨️ Imprimindo {len(pedidos)} pedidos para '{nome}'."
    return f"⚠️ Nenhum pedido encontrado para '{nome}'."


# ==================================================================================
#  SEÇÃO 5: WHATSAPP - INTELIGÊNCIA E MONITORAMENTO
# ==================================================================================
# Responsável por: Garantir que o Chrome/WhatsApp permaneça em foco, monitora
# mensagens do grupo e responde automaticamente com informações dos pedidos.
# ==================================================================================

def garantir_foco_no_grupo():
    global driver, NOME_GRUPO_FIXO
    try:
        if "whatsapp" not in driver.current_url:
            for handle in driver.window_handles:
                driver.switch_to.window(handle)
                if "whatsapp" in driver.current_url: break

        nome_normalizado = normalizar_texto(NOME_GRUPO_FIXO)
        # 1. Tenta pelo header
        try:
            headers = driver.find_elements(By.XPATH, "//header//span")
            for h in headers:
                texto = h.text.strip().lower()
                if normalizar_texto(texto) == nome_normalizado and h.is_displayed():
                    return
        except Exception:
            pass

        # 2. Tenta na lista lateral de conversas
        try:
            conversas = driver.find_elements(By.XPATH, "//span[@title]")
            for c in conversas:
                texto = c.get_attribute("title").strip().lower()
                if normalizar_texto(texto) == nome_normalizado:
                    c.click()
                    time.sleep(1)
                    print("✅ Foco restaurado para o grupo (lista lateral).")
                    return
        except Exception:
            pass

        print(f"🔍 Grupo não detectado como ativo. Abrindo: {NOME_GRUPO_FIXO}")

        def _abrir_grupo():
            # Compatível com <div> e <input> (versões diferentes do WhatsApp Web)
            search_box = WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.XPATH, '//*[@data-tab="3"][@role="textbox" or @contenteditable="true"]'))
            )
            driver.execute_script("arguments[0].click();", search_box)
            driver.execute_script("arguments[0].focus();", search_box)
            time.sleep(0.5)
            search_box.send_keys(Keys.CONTROL + "a")
            search_box.send_keys(Keys.BACKSPACE)
            time.sleep(0.5)
            search_box.send_keys(NOME_GRUPO_FIXO)
            time.sleep(2.0)
            print("⌨️ Usando Teclado para entrar no grupo...")
            search_box.send_keys(Keys.ARROW_DOWN)
            time.sleep(0.5)
            search_box.send_keys(Keys.ENTER)
            print("✅ Foco restaurado para o grupo.")

        if not _executar_com_retentativas("abrir grupo WhatsApp", _abrir_grupo, tentativas=2):
            print("⚠️ Falha ao abrir grupo após retentativas.")

    except Exception as e:
        _tratar_timeout_webdriver("Erro geral ao validar grupo", e)

def refresh_whatsapp_periodically():
    global LAST_WHATSAPP_REFRESH, driver
    if not driver:
        return

    agora = time.time()
    if agora - LAST_WHATSAPP_REFRESH < REFRESH_INTERVAL_1:
        return

    try:
        handle_atual = driver.current_window_handle
        handle_whatsapp = None

        for handle in driver.window_handles:
            driver.switch_to.window(handle)
            if "web.whatsapp.com" in driver.current_url:
                handle_whatsapp = handle
                break

        if handle_whatsapp:
            print("🔄 Atualizando WhatsApp (limpeza automatica)...")

            def _refresh():
                driver.refresh()
                WebDriverWait(driver, 60).until(
                    EC.presence_of_element_located((By.XPATH, '//*[@data-tab="3" or @data-tab="10"]'))
                )
                garantir_foco_no_grupo()

            _executar_com_retentativas("refresh WhatsApp", _refresh, tentativas=2)

        if handle_atual in driver.window_handles:
            driver.switch_to.window(handle_atual)

        LAST_WHATSAPP_REFRESH = agora
    except Exception as e:
        _tratar_timeout_webdriver("Erro ao atualizar WhatsApp", e)
        LAST_WHATSAPP_REFRESH = agora

def _eh_timeout_webdriver(exc):
    texto = str(exc).lower()
    return "read timed out" in texto or "max retries exceeded with url: /session" in texto

def _eh_sessao_invalida(exc):
    texto = str(exc).lower()
    return "invalid session id" in texto or "disconnected" in texto

def _tratar_timeout_webdriver(contexto, exc):
    print(f"⚠️ {contexto}: {exc}")
    if _eh_timeout_webdriver(exc):
        print("🧯 WebDriver sem resposta. Tentando recuperar página principal...")
        _recarregar_pagina_principal("timeout webdriver")
    if _eh_sessao_invalida(exc):
        print("🔁 Sessao do navegador invalida. Tentando reiniciar...")
        _reiniciar_chrome_se_preciso("sessao invalida")

def _recarregar_pagina_principal(motivo):
    global driver
    if not driver:
        return

    try:
        handle_atual = driver.current_window_handle
        handle_principal = None
        for handle in driver.window_handles:
            driver.switch_to.window(handle)
            try:
                url = driver.current_url
            except WebDriverException:
                continue
            url_principal = CONFIG.get('url_principal', CONFIG.get('url_api', 'https://pagina-principal.com/'))
            if url_principal.split('//')[-1].split('/')[0] in url:
                handle_principal = handle
                break

        if handle_principal:
            print(f"🔄 Recarregando página principal ({motivo})...")
            driver.refresh()
            WebDriverWait(driver, 60).until(
                EC.presence_of_element_located((By.TAG_NAME, "body"))
            )
        else:
            print(f"🧭 Abrindo página principal em nova aba ({motivo})...")
            url_principal = CONFIG.get('url_principal', CONFIG.get('url_api', 'https://pagina-principal.com/'))
            driver.execute_script(f"window.open('{url_principal}', '_blank');")

        if handle_atual in driver.window_handles:
            driver.switch_to.window(handle_atual)
    except Exception as e:
        print(f"⚠️ Falha ao recuperar página principal: {e}")

def _executar_com_retentativas(contexto, func, tentativas=2, pausa=2):
    for tentativa in range(1, tentativas + 1):
        try:
            func()
            return True
        except (TimeoutException, WebDriverException) as e:
            print(f"⚠️ {contexto} falhou (tentativa {tentativa}/{tentativas}): {e}")
            time.sleep(pausa)
        except Exception as e:
            print(f"⚠️ {contexto} erro inesperado: {e}")
            time.sleep(pausa)
    return False

def _driver_ativo():
    global driver
    if not driver:
        return False
    if not getattr(driver, "session_id", None):
        return False
    try:
        driver.execute_script("return 1")
        return True
    except Exception:
        return False

def _reiniciar_chrome_se_preciso(motivo):
    global LAST_CHROME_RESTART, driver
    agora = time.time()
    if agora - LAST_CHROME_RESTART < CHROME_RESTART_COOLDOWN:
        print("⏳ Reinicio do Chrome em cooldown.")
        return

    print(f"🔁 Reiniciando Chrome ({motivo})...")
    try:
        if driver:
            try:
                print("Tentando fechar aba ativa do Chrome...")
                driver.close()
            except Exception as e:
                print(f"Aviso: erro ao fechar aba ativa: {e}")
            try:
                print("Tentando encerrar sessão do Chrome (driver.quit)...")
                driver.quit()
                print("driver.quit() executado com sucesso.")
            except Exception as e:
                print(f"Aviso: erro ao encerrar sessão do Chrome: {e}")
        driver = None
        
        # Fallback agressivo: se após quit ainda existem chrome na porta 9222, mata via taskkill
        try:
            import subprocess
            subprocess.run(
                ['taskkill', '/F', '/IM', 'chromedriver.exe'],
                capture_output=True, timeout=5
            )
        except Exception:
            pass
        
        iniciar_chrome_persistente()
        LAST_CHROME_RESTART = agora
        
        # Verifica se o restart realmente funcionou
        if not _driver_ativo():
            print("⚠️ Chrome reiniciou mas driver não está ativo. Próximo ciclo tentará novamente.")
    except Exception as e:
        print(f"❌ Falha ao reiniciar Chrome: {e}")

def refresh_main_periodically():
    global LAST_MAIN_REFRESH, driver
    if not driver:
        return

    agora = time.time()
    if agora - LAST_MAIN_REFRESH < REFRESH_INTERVAL_2:
        return

    try:
        _recarregar_pagina_principal("manutencao")
    except Exception as e:
        print(f"⚠️ Erro ao atualizar página principal: {e}")
    finally:
        LAST_MAIN_REFRESH = agora

# ================= CONFIGURAÇÃO E LOGIN AUTOMÁTICO =================
ARQUIVO_COMANDO = 'comando_imprimir.txt'
ARQUIVO_CONFIG = 'config.json'
ARQUIVO_FECHAMENTO_STATUS = 'fechamento_status.json'
ARQUIVO_ALERTAS = 'alertas_atraso.json'
def carregar_credenciais():
    """Retorna credenciais do Zé Delivery do CONFIG global"""
    global TELEGRAM_TOKEN, TELEGRAM_CHAT_ID 
    
    # Usa as configurações já carregadas no início
    TELEGRAM_TOKEN = CONFIG.get('telegram_token', '')
    TELEGRAM_CHAT_ID = CONFIG.get('telegram_chat_id', '')
    
    if TELEGRAM_TOKEN and TELEGRAM_CHAT_ID:
        print(f"📱 Telegram Configurado! (ID: {TELEGRAM_CHAT_ID})")
    
    return CONFIG.get('email'), CONFIG.get('senha')

def carregar_motoboys_do_painel():
    """Atualiza lista de motoboys do CONFIG global"""
    global MOTOBOYS_API
    
    try:
        # Recarrega o config.json para pegar atualizações do painel
        with open('config.json', 'r', encoding='utf-8') as f:
            dados = json.load(f)
        
        motos_painel = dados.get("motoboys", {})
        if motos_painel:
            MOTOBOYS_API = motos_painel
            print(f"📋 Lista de Motoboys atualizada: {len(motos_painel)} cadastrados.")
    except Exception as e:
        print(f"⚠️ Erro ao recarregar motoboys: {e}")

def obter_config_chrome_():
    """Retorna configuração de perfil do Chrome a partir do config.json.

    Campos aceitos:
    - chrome_user_data_dir (string)
    - chrome_profile_directory (string, ex: Default, Profile 1)
    - chrome: { user_data_dir, profile_directory }
    """
    cfg_chrome = CONFIG.get('chrome', {}) if isinstance(CONFIG.get('chrome', {}), dict) else {}

    user_data_dir = str(
        cfg_chrome.get('user_data_dir')
        or CONFIG.get('chrome_user_data_dir')
        or ''
    ).strip()

    profile_directory = str(
        cfg_chrome.get('profile_directory')
        or CONFIG.get('chrome_profile_directory')
        or ''
    ).strip()

    return {
        'user_data_dir': user_data_dir,
        'profile_directory': profile_directory
    }

def ler_codigo_2fa_de_arquivo_(caminho='telegram_code.txt'):
    """Lê código 2FA de arquivo auxiliar, se existir."""
    try:
        base = get_caminho_base()
        p = os.path.join(base, caminho)
        if not os.path.exists(p):
            return ''
        with open(p, 'r', encoding='utf-8') as f:
            codigo = str(f.read() or '').strip()
        if len(codigo) == 6 and codigo.isdigit():
            return codigo
    except Exception:
        pass
    return ''

# Variável compartilhada para código 2FA (thread-safe)
_codigo_2fa_recebido = {'valor': '', 'origem': ''}
_codigo_2fa_lock = threading.Lock()

def _monitor_telegram_2fa(timeout_segundos):
    """Thread que monitora Telegram aguardando código 2FA."""
    global LAST_UPDATE_ID, _codigo_2fa_recebido
    
    if not TELEGRAM_TOKEN or not TELEGRAM_CHAT_ID:
        return
    
    url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/getUpdates"
    inicio = time.time()
    
    while time.time() - inicio < timeout_segundos:
        with _codigo_2fa_lock:
            if _codigo_2fa_recebido['valor']:
                return  # Já recebeu de outra fonte
        
        try:
            params = {"timeout": 3, "offset": LAST_UPDATE_ID + 1}
            resp = cffi_requests.get(url, params=params, timeout=8)
            data = resp.json()
            
            if not data.get("ok"):
                time.sleep(1)
                continue
            
            for update in data.get("result", []):
                LAST_UPDATE_ID = update["update_id"]
                msg = update.get("message", {})
                text = str(msg.get("text", "")).strip()
                chat_id = str(msg.get("chat", {}).get("id", ""))
                
                # Verifica se é do chat correto
                if chat_id not in str(TELEGRAM_CHAT_ID).split(','):
                    continue
                
                # Verifica se é um código de 6 dígitos
                if len(text) == 6 and text.isdigit():
                    with _codigo_2fa_lock:
                        if not _codigo_2fa_recebido['valor']:
                            _codigo_2fa_recebido['valor'] = text
                            _codigo_2fa_recebido['origem'] = 'telegram'
                    enviar_telegram(f"✅ Código recebido via Telegram!")
                    return
                    
        except Exception:
            pass
        
        time.sleep(1)

def _monitor_terminal_2fa(timeout_segundos):
    """Thread que monitora input do terminal aguardando código 2FA."""
    global _codigo_2fa_recebido
    import select
    import msvcrt
    
    inicio = time.time()
    buffer_input = ""
    
    while time.time() - inicio < timeout_segundos:
        with _codigo_2fa_lock:
            if _codigo_2fa_recebido['valor']:
                return  # Já recebeu de outra fonte
        
        try:
            # Windows: usa msvcrt para input não-bloqueante
            if msvcrt.kbhit():
                char = msvcrt.getwch()
                if char == '\r' or char == '\n':  # Enter pressionado
                    codigo = buffer_input.strip()
                    if len(codigo) == 6 and codigo.isdigit():
                        with _codigo_2fa_lock:
                            if not _codigo_2fa_recebido['valor']:
                                _codigo_2fa_recebido['valor'] = codigo
                                _codigo_2fa_recebido['origem'] = 'terminal'
                        print(f"\n✅ Código recebido via Terminal!")
                        return
                    else:
                        print(f"\n⚠️ Código inválido: '{codigo}' (deve ter 6 dígitos)")
                        buffer_input = ""
                        print(">>> ", end="", flush=True)
                elif char == '\b':  # Backspace
                    if buffer_input:
                        buffer_input = buffer_input[:-1]
                        print('\b \b', end="", flush=True)
                else:
                    buffer_input += char
                    print(char, end="", flush=True)
        except Exception:
            pass
        
        time.sleep(0.1)

def obter_codigo_2fa_():
    """Obtém código 2FA - aceita de arquivo, env, Telegram OU terminal (quem vier primeiro)."""
    global _codigo_2fa_recebido
    
    # 1) variável de ambiente (prioridade)
    codigo_env = str(os.getenv('ZE_2FA_CODE', '') or '').strip()
    if len(codigo_env) == 6 and codigo_env.isdigit():
        return codigo_env

    # 2) arquivo auxiliar local
    codigo_file = ler_codigo_2fa_de_arquivo_('telegram_code.txt')
    if codigo_file:
        return codigo_file

    # Reset da variável compartilhada
    with _codigo_2fa_lock:
        _codigo_2fa_recebido['valor'] = ''
        _codigo_2fa_recebido['origem'] = ''
    
    # 3) Pede código simultaneamente via Telegram e Terminal
    timeout = 120  # 2 minutos
    
    winsound.Beep(1000, 500)
    print("\n" + "=" * 50)
    print("🚨 CÓDIGO 2FA NECESSÁRIO!")
    print("Responda pelo TELEGRAM ou digite aqui no TERMINAL")
    print("(6 dígitos recebidos por e-mail)")
    print("=" * 50)
    print(">>> ", end="", flush=True)
    
    # Envia mensagem no Telegram
    if TELEGRAM_TOKEN and TELEGRAM_CHAT_ID:
        enviar_telegram("🔐 *CÓDIGO 2FA NECESSÁRIO!*\n\nDigite o código de 6 dígitos aqui OU no terminal do painel.\nQuem responder primeiro ganha! ⏱️ 2 min")
    
    # Inicia threads para monitorar ambas as fontes
    thread_telegram = threading.Thread(target=_monitor_telegram_2fa, args=(timeout,), daemon=True)
    thread_terminal = threading.Thread(target=_monitor_terminal_2fa, args=(timeout,), daemon=True)
    
    thread_telegram.start()
    thread_terminal.start()
    
    # Aguarda até receber código ou timeout
    inicio = time.time()
    while time.time() - inicio < timeout:
        with _codigo_2fa_lock:
            if _codigo_2fa_recebido['valor']:
                origem = _codigo_2fa_recebido['origem']
                codigo = _codigo_2fa_recebido['valor']
                print(f"\n📱 Código capturado via: {origem.upper()}")
                return codigo
        
        # Também checa arquivo periodicamente (para quem escrever manualmente)
        codigo_file = ler_codigo_2fa_de_arquivo_('telegram_code.txt')
        if codigo_file:
            print(f"\n📄 Código capturado via: ARQUIVO")
            return codigo_file
        
        time.sleep(0.5)
    
    print("\n⏰ Tempo esgotado para inserir código 2FA!")
    if TELEGRAM_TOKEN and TELEGRAM_CHAT_ID:
        enviar_telegram("⏰ Tempo esgotado! Reinicie o robô e tente novamente.")
    
    return ''

# ==================================================================================
#  SEÇÃO 6: CHROME E NAVEGAÇÃO WEB
# ==================================================================================
# Responsável por: Inicializar e gerenciar instância persistente do Chrome
# Selenium para automação de navegação no WhatsApp 
# ==================================================================================

def aguardar_qrcode_whatsapp():
    """Detecta se o QR code está sendo exibido, tira print e envia no Telegram."""
    global driver
    
    qr_enviado = False
    tentativas_qr = 0
    max_espera = 300  # 5 minutos máximo de espera
    
    # Primeiro aguarda a página do WhatsApp carregar minimamente
    print("⏳ Aguardando página do WhatsApp carregar...")
    for i in range(60):  # até 60s para a página carregar
        try:
            ready = driver.execute_script("return document.readyState")
            # Verifica se já tem algum conteúdo WhatsApp na página
            tem_conteudo = driver.execute_script(
                "return document.querySelector('#app') !== null || "
                "document.querySelector('[data-tab]') !== null || "
                "document.querySelector('canvas') !== null || "
                "document.querySelector('[data-ref]') !== null || "
                "document.querySelector('[role=\"textbox\"]') !== null"
            )
            if ready == 'complete' and tem_conteudo:
                print("✅ Página do WhatsApp carregada.")
                break
            if i > 0 and i % 10 == 0:
                print(f"   ... ainda carregando ({i}s)")
        except Exception:
            pass
        time.sleep(1)
    
    while True:
        try:
            # Verifica se já está logado - múltiplos seletores para compatibilidade
            logado = driver.execute_script("""
                // Campo de busca (versões diferentes)
                if (document.querySelector('[data-tab="3"][role="textbox"]') || document.querySelector('div[contenteditable="true"][data-tab="3"]')) return true;
                if (document.querySelector('div[contenteditable="true"][title*="Search"]')) return true;
                if (document.querySelector('div[contenteditable="true"][title*="Pesquisar"]')) return true;
                // Sidebar de conversas existente = logado
                if (document.querySelector('[aria-label="Chat list"]') || document.querySelector('[aria-label="Lista de conversas"]')) return true;
                // Painel de chat aberto
                if (document.querySelector('div[data-tab="10"]')) return true;
                // Header com foto de perfil
                if (document.querySelector('header span[data-icon="default-user"]') || document.querySelector('header img[draggable="false"]')) return true;
                return false;
            """)
            if logado:
                if qr_enviado:
                    print("✅ QR Code escaneado com sucesso!")
                    enviar_telegram("✅ WhatsApp conectado com sucesso!")
                else:
                    print("✅ WhatsApp já está logado!")
                return
            
            # Verifica se QR code está visível - múltiplos seletores
            tem_qr = driver.execute_script("""
                if (document.querySelector('canvas[aria-label="Scan this QR code to link a device!"]')) return true;
                if (document.querySelector('canvas[aria-label]')) {
                    var c = document.querySelector('canvas[aria-label]');
                    if (c.width > 200) return true;
                }
                if (document.querySelector('div[data-ref]')) return true;
                // Texto "Link a device" ou "Conectar um aparelho"
                var spans = document.querySelectorAll('span');
                for (var s of spans) {
                    var t = s.textContent.toLowerCase();
                    if (t.includes('link a device') || t.includes('conectar um aparelho') || t.includes('scan') || t.includes('escanei')) return true;
                }
                return false;
            """)
            
            if tem_qr and not qr_enviado:
                print("📸 QR Code detectado! Tirando screenshot...")
                
                caminho_qr = os.path.join(get_caminho_base(), "qrcode_whatsapp.png")
                driver.save_screenshot(caminho_qr)
                
                enviar_telegram("📱 *QR Code do WhatsApp*\nEscaneie com seu celular para conectar:")
                time.sleep(0.5)
                if enviar_foto_telegram(caminho_qr, "Escaneie este QR Code com o WhatsApp do seu celular"):
                    print("✅ QR Code enviado para o Telegram!")
                    qr_enviado = True
                else:
                    print("⚠️ Falha ao enviar QR Code no Telegram")
            
            tentativas_qr += 1
            if tentativas_qr > max_espera:
                print("⏰ Tempo esgotado aguardando escaneamento do QR Code!")
                enviar_telegram("⏰ Tempo esgotado aguardando QR Code! Reinicie o robô.")
                raise RuntimeError("Timeout aguardando QR Code")
            
            # Log periódico para saber que não travou
            if tentativas_qr % 15 == 0:
                print(f"   ⏳ Aguardando WhatsApp... ({tentativas_qr}s)")
            
            # A cada 30 segundos, envia novo QR (pode ter mudado)
            if qr_enviado and tentativas_qr % 30 == 0:
                qr_enviado = False
                print("🔄 Reenviando QR Code atualizado...")
            
            time.sleep(1)
            
        except Exception as e:
            if "Timeout" in str(e):
                raise
            print(f"⚠️ Erro ao verificar QR: {e}")
            time.sleep(2)

def iniciar_chrome_persistente():
    global driver, TOKEN_ATUAL
    print("🤖 Iniciando Chrome (MODO VISÍVEL - STEALTH)...")
    
    chrome_cfg = obter_config_chrome_()
    perfil_path = chrome_cfg.get('user_data_dir') or os.path.join(get_caminho_base(), "perfil_chrome")
    if not os.path.exists(perfil_path): os.makedirs(perfil_path)

    def matar_chrome_do_robo():
        """Finaliza de forma agressiva processos do Chrome e Driver."""
        perfil_lower = perfil_path.lower().replace('/', '\\')
        pids_mortos = []
        # Passo 1: mata por cmdline (match no perfil) + todos chromedriver
        for proc in psutil.process_iter(['pid', 'name', 'cmdline']):
            try:
                info = proc.info
                nome = str(info.get('name', '')).lower()
                cmdline_list = info.get('cmdline') or []
                cmdline = ' '.join(cmdline_list).lower().replace('/', '\\')
                
                if (nome == 'chrome.exe' and perfil_lower in cmdline) or (nome == 'chromedriver.exe'):
                    try:
                        proc.kill()
                        pids_mortos.append(info['pid'])
                        print(f"🛑 Processo {nome} (PID {info['pid']}) encerrado.")
                    except:
                        pass
            except Exception:
                pass
        # Passo 2: mata chrome órfão cujo cmdline está inacessível (Access Denied)
        # Esses processos travam o perfil mas não aparecem no filtro acima
        for proc in psutil.process_iter(['pid', 'name']):
            try:
                info = proc.info
                if str(info.get('name', '')).lower() != 'chrome.exe':
                    continue
                if info['pid'] in pids_mortos:
                    continue
                # Tenta ler cmdline - se der AccessDenied, é órfão
                try:
                    cmdline_list = proc.cmdline()
                    cmdline = ' '.join(cmdline_list).lower().replace('/', '\\')
                    if perfil_lower in cmdline:
                        proc.kill()
                        print(f"🛑 Chrome órfão (PID {info['pid']}) encerrado via cmdline().")
                except (psutil.AccessDenied, psutil.ZombieProcess):
                    # Chrome sem cmdline acessível: verificar se tem porta 9222 (nosso debug port)
                    try:
                        conns = proc.net_connections()
                        for c in conns:
                            if c.laddr and c.laddr.port == 9222:
                                proc.kill()
                                print(f"🛑 Chrome órfão (PID {info['pid']}) encerrado via porta 9222.")
                                break
                    except (psutil.AccessDenied, psutil.ZombieProcess):
                        pass
                except psutil.NoSuchProcess:
                    pass
            except Exception:
                pass

    matar_chrome_do_robo()
    time.sleep(3) # Tempo extra para SO liberar handles de arquivo

    # Limpeza de travas e verificação do arquivo de preferências
    travas = ["SingletonLock", "DevToolsActivePort", "SingletonCookie", "SingletonSocket", "lockfile", "Preferences.old"]
    for root, dirs, files in os.walk(perfil_path):
        for f in files:
            if f in travas:
                try: os.remove(os.path.join(root, f))
                except: pass
        if root == perfil_path: break # Foca apenas na raiz e subpastas imediatas

    # TRATAMENTO ESPECÍFICO: Se o Preferences estiver travado/corrompido, resetamos ele
    prefs_path = os.path.join(perfil_path, "Default", "Preferences")
    if os.path.exists(prefs_path):
        try:
            with open(prefs_path, 'a'): pass # Tenta abrir para escrita
        except Exception:
            print("⚠️ Arquivo de preferências corrompido ou travado. Resetando...")
            try:
                backup_prefs = prefs_path + ".old"
                if os.path.exists(backup_prefs): os.remove(backup_prefs)
                os.rename(prefs_path, backup_prefs)
            except: pass

    opts = Options()
    opts.add_argument(f"--user-data-dir={perfil_path}") 
    if chrome_cfg.get('profile_directory'):
        opts.add_argument(f"--profile-directory={chrome_cfg.get('profile_directory')}")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)
    opts.add_argument("--log-level=3") 
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--remote-debugging-port=9222")
    # Flags para evitar "tab crashed"
    opts.add_argument("--disable-gpu")
    opts.add_argument("--disable-software-rasterizer")
    opts.add_argument("--disable-extensions")
    opts.add_argument("--disable-background-networking")
    opts.add_argument("--disable-default-apps")
    opts.add_argument("--disable-sync")
    opts.add_argument("--no-first-run")
    opts.add_argument("--disable-infobars")
    opts.add_argument("--start-maximized")

    try:
        driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=opts)
        driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
            "source": """
            Object.defineProperty(navigator, 'webdriver', {
              get: () => undefined
            })
            """
        })
        
        url_principal = CONFIG.get('url_principal', CONFIG.get('url_api', 'https://pagina-principal.com/'))
        driver.get(url_principal)
        
        email_cfg, senha_cfg = carregar_credenciais()
        if email_cfg and senha_cfg:
            try:
                print("🔑 Tentando login automático com dados do config.json...")
                wait_login = WebDriverWait(driver, 20)
                def preencher_campo_hexa_blindado(nome_atributo, valor, eh_senha=False):
                    host_element = wait_login.until(EC.presence_of_element_located((By.XPATH, f"//hexa-v2-input-text[@name='{nome_atributo}']")))
                    input_real = driver.execute_script("return arguments[0].shadowRoot.querySelector('input')", host_element)
                    time.sleep(0.5)
                    driver.execute_script(f"arguments[0].value = '{valor}';", input_real)
                    driver.execute_script("arguments[0].dispatchEvent(new Event('input', { bubbles: true }));", input_real)
                    driver.execute_script("arguments[0].dispatchEvent(new Event('change', { bubbles: true }));", input_real)
                    driver.execute_script("arguments[0].dispatchEvent(new Event('blur', { bubbles: true }));", input_real)
                    time.sleep(1)
                    if eh_senha: pass
                preencher_campo_hexa_blindado("email", email_cfg)
                preencher_campo_hexa_blindado("password", senha_cfg, eh_senha=True)
                try:
                    time.sleep(2)
                    host_btn = driver.find_element(By.XPATH, "//hexa-v2-button")
                    driver.execute_script("var host = arguments[0]; var root = host.shadowRoot; var btn = root.querySelector('button[type=\"submit\"]'); if (btn) btn.click();", host_btn)
                except:
                    driver.execute_script("var all = document.querySelectorAll('*'); for (var i=0; i<all.length; i++) { if (all[i].shadowRoot) { var btn = all[i].shadowRoot.querySelector('button[type=\"submit\"]'); if (btn && btn.innerText.includes('Entrar')) { btn.click(); break; } } }")
                time.sleep(3)
                try:
                    time.sleep(3)
                    btn_enviar_email = driver.find_elements(By.ID, "send-email-button")
                    if btn_enviar_email:
                        driver.execute_script("arguments[0].click();", btn_enviar_email[0])
                        time.sleep(3)
                except:
                    pass
                try:
                    time.sleep(2)
                    input_token_0 = driver.find_elements(By.ID, "verification-code-input-0")
                    if input_token_0:
                        codigo = obter_codigo_2fa_()
                        if codigo:
                            for i in range(6):
                                try:
                                    driver.find_element(By.ID, f"verification-code-input-{i}").send_keys(codigo[i])
                                except:
                                    pass
                                time.sleep(0.1)
                            time.sleep(1)
                            btn_confirmar = driver.find_elements(By.ID, "send-code-verification")
                            if btn_confirmar:
                                driver.execute_script("arguments[0].click();", btn_confirmar[0])
                                time.sleep(3)
                        else:
                            print("⚠️ Login 2FA detectado e sem código disponível (console/arquivo/env).")
                            print("⚠️ Em modo serviço, configure o perfil real do Chrome no config.json para evitar 2FA recorrente.")
                except:
                    pass
            except Exception as e:
                print(f"ℹ️ Pulei o login automático: {e}")

        print("⏳ Aguardando validação do Token...")
        inicio_token = time.time()
        timeout_token = 180
        while not TOKEN_ATUAL:
            try:
                for c in driver.get_cookies():
                    if c['name'] == "seu_ze_access_token": TOKEN_ATUAL = c['value']; break
            except: pass
            if time.time() - inicio_token > timeout_token:
                raise RuntimeError("Token não capturado em 180s. Verifique login/perfil do Chrome.")
            time.sleep(2)
        
        print("🔓 Token capturado com sucesso!")
        
        # --- VERIFICAÇÃO INTELIGENTE DE ABAS ---
        # Verifica se o Chrome já restaurou a aba do WhatsApp automaticamente
        aba_whatsapp_encontrada = False
        for handle in driver.window_handles:
            driver.switch_to.window(handle)
            if "whatsapp" in driver.current_url:
                aba_whatsapp_encontrada = True
                break
                
        # Se não achou nenhuma aba do WhatsApp, abre uma nova
        if not aba_whatsapp_encontrada:
            print("[DEBUG] Abrindo nova aba do WhatsApp Web...")
            driver.execute_script("window.open('https://web.whatsapp.com', '_blank');")
            print("[DEBUG] Aba do WhatsApp Web aberta, trocando para a nova aba...")
            driver.switch_to.window(driver.window_handles[-1])
            print(f"[DEBUG] URL atual após troca: {driver.current_url}")
        # ---------------------------------------
        
        print("📱 Aguardando carregamento do WhatsApp...")
        
        # Aguarda carregamento e verifica se precisa escanear QR code
        aguardar_qrcode_whatsapp()
        
        print(f"🔒 Bloqueando no grupo: {NOME_GRUPO_FIXO}")
        garantir_foco_no_grupo() 
        print("✅ Sistemas OK!")
    except Exception as e: print(f"❌ Erro: {e}"); sys.exit()

def enviar_mensagem_grupo(mensagem):
    global driver
    try:
        garantir_foco_no_grupo()
        
        # Localiza a caixa de texto (compatível com <div> e <input>)
        caixas = driver.find_elements(By.XPATH, '//*[@data-tab="10"][@role="textbox" or @contenteditable="true"]')
        if not caixas: 
            caixas = driver.find_elements(By.XPATH, '//div[@contenteditable="true"]')
        
        if caixas:
            box = caixas[-1]
            driver.execute_script("arguments[0].focus();", box)
            
            # VERIFICA SE A MENSAGEM COME A COM MENÇÃO
            mencao_prefixo = "@+55 49 9172-7951 "
            tem_mencao = mensagem.startswith(mencao_prefixo)
            
            if tem_mencao:
                # Remove o prefixo de menção da mensagem
                mensagem_sem_mencao = mensagem[len(mencao_prefixo):]
                
                try:
                    # 1. Digita @ para abrir dropdown de contatos
                    box.send_keys("@")
                    time.sleep(0.8)  # Aguarda dropdown aparecer
                    
                    # 2. Digita "amor" para buscar o contato
                    box.send_keys("amor")
                    time.sleep(0.6)  # Aguarda busca filtrar
                    
                    # 3. Pressiona ENTER para selecionar o primeiro resultado
                    box.send_keys(Keys.ENTER)
                    time.sleep(0.3)  # Aguarda menção ser inserida
                    
                    # 4. Adiciona espaço após a menção
                    box.send_keys(" ")
                    time.sleep(0.2)
                    
                    print("✅ Menção '@amor' criada com sucesso")
                    
                    # Agora continua com o resto da mensagem
                    mensagem = mensagem_sem_mencao
                    
                except Exception as e:
                    print(f"⚠️ Erro ao criar menção, enviando texto normal: {e}")
                    # Se falhar, envia a mensagem original completa
                    mensagem = mensagem  # Mantém com o @+55...
            
            # --- LÓGICA DE BLOCO ÚNICO ---
            # Divide o texto onde tem quebra de linha
            linhas = mensagem.split('\n')
            
            for i, linha in enumerate(linhas):
                texto_limpo = linha.strip()
                
                # Digita a linha atual (se não for vazia, ou se quiser manter espaço vazio)
                if texto_limpo:
                    # Usa JS para inserir o texto (mais rápido e seguro contra emojis)
                    driver.execute_script(f"document.execCommand('insertText', false, {json.dumps(texto_limpo)});")
                
                # Se NÃO for a última linha, aperta SHIFT + ENTER para pular linha
                if i < len(linhas) - 1:
                    box.send_keys(Keys.SHIFT, Keys.ENTER)
                    time.sleep(0.1) # Breve pausa para o WhatsApp processar a quebra

            # Só depois de digitar tudo, aperta ENTER para enviar o bloco
            time.sleep(0.5)
            box.send_keys(Keys.ENTER)
            msg_tipo = "(com menção)" if tem_mencao else "(Bloco Único Organizado)"
            print(f"📤 Mensagem enviada {msg_tipo}.")
            
        else:
            print("❌ Não encontrei a caixa de texto para responder.")
            
    except Exception as e:
        print(f"❌ Erro ao enviar zap: {e}")

# --- FUNÇÃO NOVA: Traduzir status para ficar bonito no Zap ---
def traduzir_status(status_raw):
    mapa = {
        "POC_ACCEPTED": "👨‍🍳 Aceito (Na Loja)", "DISPATCHED": "🚚 Despachado", 
        "IN_TRANSIT": "🛵 Em Rota", "DM_EN_ROUTE": "🛵 Em Rota", 
        "EM_ROUTE": "🛵 Em Rota", "DELIVERED": "✅ Entregue", 
        "POC_DELIVERED": "✅ Entregue", "FINISHED": "🏁 Finalizado", 
        "CANCELLED": "❌ Cancelado", "DM_PICKED_UP": "🛵 Retirado pelo entregador"
    }
    return mapa.get(str(status_raw).upper(), status_raw)        
def buscar_telefone(num):
    try:
        p = {"operationName": "sellerGetCustomerPhoneNumber", "query": "mutation sellerGetCustomerPhoneNumber($orderNumber: String!, $contactReason: OrderContactReasonInput!) { getCustomerPhoneNumber(orderNumber: $orderNumber, contactReason: $contactReason) { phoneNumber } }", "variables": {"orderNumber": str(num), "contactReason": {"category": "REASON_CATEGORY_CHANGE_ORDER", "description": "Validar", "item": "REASON_ITEM_PRODUCT_MISSING"}}}
        
        # --- VIP PASS AQUI ---
        r = requisicao_segura(p, ignorar_limite=True) 
        
        if r and r.status_code == 200:
            return r.json()["data"]["getCustomerPhoneNumber"]["phoneNumber"].replace("+", "")
    except: return "Não disp."
    return "Não disp."

# === NOVA FUNÇÃO: CONSULTA DIRETA NA API (ACTIVE ORDERS) ===
def consultar_api_direta():
    # Query ajustada para pegar exatamente os dados do seu JSON
    q = """query sellerActiveOrders { 
        activeOrders { 
            number 
            date
            status 
            customer { name } 
            items { name amount } 
            delivery { 
                courier { email }
                address { neighborhood }
            } 
        } 
    }"""
    try:
        # --- VIP PASS AQUI TAMBÉM ---
        r = requisicao_segura({"query": q}, ignorar_limite=True)
        
        if r and r.status_code == 200:
            d = r.json()
            if "data" in d and "activeOrders" in d["data"]:
                return d["data"]["activeOrders"]
    except Exception as e:
        print(f"Erro ao consultar API direta: {e}")
    return []
# === NOVA FUNÇÃO: BUSCAR TODOS NO EXCEL (MANTIDA) ===
def buscar_todos_pedidos_excel_por_nome(nome_buscado):
    encontrados = []
    arquivo = get_caminho_excel()
    if not os.path.exists(arquivo): return encontrados
    
    try:
        wb = openpyxl.load_workbook(arquivo, data_only=True)
        ws = wb["EXTRATO DETALHADO"]
        nome_norm = normalizar_texto(nome_buscado)
        
        # Itera sobre as linhas de baixo para cima (mais recentes primeiro)
        for row in reversed(list(ws.iter_rows(min_row=2, values_only=True))):
            if not row or len(row) < 7: continue 
            
            # Índices do Excel: 3=Cliente, 5=Status, 6=Motoboy
            motoboy_excel = normalizar_texto(str(row[6])) if row[6] else ""
            status = str(row[5]).upper() if row[5] else ""

            # Não busca em pedidos cancelados
            if any(x in status for x in STATUS_CANCELADOS_LISTA): continue

            match = False
            
            # Se a busca for "retirada", mantém a lógica antiga
            if "retirada" in nome_norm:
                valor = float(row[8]) if row[8] else 0.0
                if "retirada" in motoboy_excel or valor == 0:
                    match = True
            
            # AGORA A BUSCA É ESTRITAMENTE NO NOME DO MOTOBOY
            elif nome_norm in motoboy_excel:
                match = True
            
            if match:
                encontrados.append({
                    'numero': str(row[2]),
                    'hora': str(row[1]),
                    'cliente': row[3], # Mantém a chave 'cliente' pro print funcionar
                    'bairro': row[4],
                    'status': status,
                    'motoboy': row[6],
                    'valor': float(row[8]) if row[8] else 0.0,
                    'itens': row[9] if len(row) > 9 else ""
                })
                
    except Exception as e:
        print(f"❌ Erro ao ler Excel para busca: {e}")
        
    return encontrados
# === NOVA FUNÇÃO: CONSULTA DIRETA NA API (ACTIVE ORDERS) ===
def consultar_api_direta():
    # Query ajustada para pegar exatamente os dados do seu JSON
    q = """query sellerActiveOrders { 
        activeOrders { 
            number 
            date
            status 
            customer { name } 
            items { name amount } 
            delivery { 
                courier { email }
                address { neighborhood }
            } 
        } 
    }"""
    try:
        # Usa a sua função de requisição segura existente
        r = requisicao_segura({"query": q})
        if r and r.status_code == 200:
            d = r.json()
            if "data" in d and "activeOrders" in d["data"]:
                return d["data"]["activeOrders"]
    except Exception as e:
        print(f"Erro ao consultar API direta: {e}")
    return []
# === WHATSAPP READ & REPLY (MANTIDA) ===
# --- VARIÁVEL DE CONTROLE PARA INICIALIZAÇÃO ---
PRIMEIRA_LEITURA_FEITA = False


# === WHATSAPP READ & REPLY ===
def verificar_solicitacoes_whatsapp():
    global driver, IDS_PROCESSADOS, PRIMEIRA_LEITURA_FEITA
    EMOJIS_ROBO = ["🚀", "📦", "👤", "📞", "🛒", "📍", "📊", "✅", "📝", "💰", "🖨️"]
    
    try:
        garantir_foco_no_grupo()
        rows = driver.find_elements(By.XPATH, '//div[@role="row"]')
        if not rows: return
        msgs = rows[-15:]

        # Pula histórico ao ligar
        if not PRIMEIRA_LEITURA_FEITA:
            for r in msgs:
                try: 
                    c = r.find_elements(By.XPATH, ".//div[@data-id]")
                    if c: IDS_PROCESSADOS.add(c[0].get_attribute("data-id"))
                except: pass
            PRIMEIRA_LEITURA_FEITA = True; return

        # === PÓ ÚLTIMA MENSAGEM VÁLIDA (não processada, não do robô) ===
        ultima_msg_valida = None
        for row in reversed(msgs):  # Itera de trás pra frente (mais recente primeiro)
            try:
                container = row.find_elements(By.XPATH, ".//div[@data-id]")
                if not container: continue
                msg_el = container[0]
                mid = msg_el.get_attribute("data-id")
                
                # Ignora mensagens já processadas, outgoing, etc
                if mid in IDS_PROCESSADOS: continue
                if str(mid).startswith("true_"): IDS_PROCESSADOS.add(mid); continue
                if "message-out" in msg_el.get_attribute("class"): IDS_PROCESSADOS.add(mid); continue
                
                # Extrai texto
                txt = ""
                try:
                    el_txt = msg_el.find_elements(By.XPATH, ".//span[contains(@class, 'copyable-text')]")
                    if el_txt:
                        txt = el_txt[-1].text.strip()
                    else:
                        linhas = msg_el.text.split('\n')
                        if len(linhas) >= 3: txt = linhas[1] 
                        elif len(linhas) == 2: txt = linhas[0] 
                        elif len(linhas) == 1: txt = linhas[0]
                except: continue
                
                if not txt or len(txt) < 2: continue
                if any(txt.startswith(e) for e in EMOJIS_ROBO): IDS_PROCESSADOS.add(mid); continue
                
                # Encontrou a última mensagem válida!
                ultima_msg_valida = {'mid': mid, 'txt': txt}
                break
            except:
                continue
        
        # === PROCESSA A ÚLTIMA MENSAGEM ===
        if ultima_msg_valida:
            mid = ultima_msg_valida['mid']
            txt = ultima_msg_valida['txt']
            
            # --- MARCA COMO PROCESSADA ---
            IDS_PROCESSADOS.add(mid)
            print(f"📩 Cliente diz: '{txt}'")
            t_low = txt.lower()

            # 1. É NÚMERO DE PEDIDO?
            match_num = re.search(r"(\d{5,})", txt)
            if match_num:
                pid = match_num.group(1); tel = buscar_telefone(pid); api_d = consultar_api_direta() or []
                pedidos_encontrados = []
                for p in api_d:
                    if str(p['number']) == pid:
                        pedidos_encontrados.append(p)
                
                # Envia TODOS os pedidos encontrados com esse número
                if pedidos_encontrados:
                    for p in pedidos_encontrados:
                        msg = f"🚀 *ATIVO*\n📊 {traduzir_status(p['status'])}\n📦 {pid}\n👤 {p['customer']['name']}\n📞 {tel}"
                        enviar_mensagem_grupo(msg)
                else:
                    msg = f"🔎 *PEDIDO {pid}*\n📞 Contato: {tel}"
                    enviar_mensagem_grupo(msg)
            
            elif "imprimir" in t_low: enviar_mensagem_grupo(processar_impressao_individual(txt))
            elif any(x in t_low for x in ["cancelada", "cancelado"]): enviar_mensagem_grupo(processar_relatorio_canceladas())
            
            # 2. BUSCA POR NOME
            else:
                termo = limpar_texto_busca(txt)
                if len(termo) >= 2:  # Reduzido de 3 para 2 caracteres (ex: "PA" para paula)
                    api_d = consultar_api_direta() or []
                    pedidos_encontrados = []  # Acumula TODOS os pedidos do cliente
                    
                    for p in api_d:
                        nome_cliente = normalizar_texto(p['customer']['name'])
                        status_p = str(p['status']).upper()
                        
                        # Tenta busca por substring ou similaridade
                        encontra_nome = (termo in nome_cliente) or (termo and nome_cliente and difflib.SequenceMatcher(None, termo, nome_cliente).ratio() > 0.75)
                        
                        # BUSCA EM TODOS OS STATUS ATIVOS (não só em rota)
                        if encontra_nome:
                            pid = str(p['number'])
                            tel = buscar_telefone(pid)
                            email_moto = "Desconhecido"
                            try: email_moto = p['delivery']['courier']['email']
                            except: pass
                            nome_moto = identificar_motoboy(email_moto)
                            
                            # Extrai informações do pedido
                            try:
                                bairro = p['delivery']['address']['neighborhood']
                            except:
                                bairro = "Não disp."
                            
                            try:
                                valor = p['charging']['total']
                            except:
                                valor = 0.0
                            
                            # Monta lista de itens
                            itens_list = []
                            try:
                                for item in p.get('items', []):
                                    nome_item = item.get('name', 'Item')
                                    qtd = item.get('amount', 1)
                                    itens_list.append(f"• {qtd}x {nome_item}")
                            except:
                                pass
                            
                            itens_str = "\n".join(itens_list) if itens_list else "Sem itens"
                            
                            # Monta mensagem completa (emoji varia por status)
                            status_trad = traduzir_status(status_p)
                            status_rua = ['DISPATCHED', 'IN_TRANSIT', 'EN_ROUTE', 'EM_ROUTE', 'DM_EN_ROUTE', 'DM_PICKED_UP', 'PICKED_UP', 'OUT_FOR_DELIVERY', 'ON_THE_WAY', 'IN_DELIVERY', 'DM_ROUTING']
                            emoji = "🛵" if status_p in status_rua else "📦"
                            msg = (f"{emoji} *{status_trad}*\n"
                                   f"📦 {pid}\n"
                                   f"👤 *{p['customer']['name'].upper()}*\n"
                                   f"📍 {bairro}\n"
                                   f"📞 {tel}\n"
                                   f"🏍️ {nome_moto}\n"
                                   f"🛒 **ITENS:**\n{itens_str}")
                            
                            pedidos_encontrados.append(msg)  # ACUMULA em vez de enviar direto
                    
                    # Envia TODOS os pedidos encontrados de uma vez
                    if pedidos_encontrados:
                        if len(pedidos_encontrados) > 1:
                            enviar_mensagem_grupo(f"📋 *Encontrei {len(pedidos_encontrados)} pedidos para '{txt.upper()}'*\n")
                        for msg in pedidos_encontrados:
                            enviar_mensagem_grupo(msg)
                    else:
                        print(f"⚠️ Nome '{termo}' não encontrado.")
    except Exception as e:
        print(f"⚠️ Erro Geral Zap: {e}")

# === HISTÓRICO RÁPIDO ===
def buscar_historico_do_dia(limite_paginas=None):
    global TOKEN_ATUAL, CACHE_NOMES_DO_DIA, CACHE_STATUS_PEDIDOS
    agora = datetime.now()
    inicio_turno = agora.replace(hour=8, minute=0, second=0) 
    if agora.hour < 8: inicio_turno -= timedelta(days=1)
    fim_turno = (inicio_turno + timedelta(days=1)).replace(hour=4, minute=0, second=0)
    s_utc = (inicio_turno + timedelta(hours=3)).strftime('%Y-%m-%dT%H:%M:%S.000Z')
    e_utc = (fim_turno + timedelta(hours=3)).strftime('%Y-%m-%dT%H:%M:%S.000Z')
    pagina = 1
    erros = 0
    total_lidos = 0
    print(f"📜 Buscando histórico de {inicio_turno.strftime('%H:%M')} até agora...")

    while True:
        if erros > 5: 
            print("❌ Muitos erros no histórico. Parando busca.")
            break
        if limite_paginas and pagina > limite_paginas: break 
        
        q = "query sellerOrderHistoryService($period: PeriodInput!, $pagination: PaginationInput!, $migration: Boolean) { orderHistory(pagination: $pagination, period: $period, migration: $migration) { number, date, status, customer { name }, items { name amount }, delivery { type, courier { email }, address { neighborhood } } } }"
        p = {"operationName": "sellerOrderHistoryService", "query": q, "variables": {"period": {"start": s_utc, "end": e_utc}, "pagination": {"page": pagina, "perPage": 100}, "migration": True}}
        
        try:
            print(f"   📄 Lendo página {pagina}...", end="", flush=True)
            r = requisicao_segura(p)
            if r is None: 
                erros += 1
                continue

            if r.status_code == 200:
                d = r.json()
                if "data" in d and d["data"] and d["data"]["orderHistory"]:
                    lista_historico = d["data"]["orderHistory"]
                    qtd_na_pagina = len(lista_historico)
                    print(f" ✅ Encontrei {qtd_na_pagina} pedidos.")
                    if qtd_na_pagina == 0: break 
                    for item in lista_historico:
                        num = item['number']; st = item.get('status', '')
                        dt_pedido = parse_data_pedido(item.get('date')) or datetime.now()
                        CACHE_NOMES_DO_DIA[num] = {
                            'nome': item['customer']['name'], 
                            'hora': dt_pedido,
                            'bairro': item['delivery']['address'].get('neighborhood', 'Não disp.'),
                            'itens': formatar_itens_para_string(item.get('items', []))
                        }
                        if num in CACHE_STATUS_PEDIDOS and CACHE_STATUS_PEDIDOS[num] == str(st).upper(): continue
                        try: email = item['delivery']['courier']['email']
                        except: email = None
                        bairro = item['delivery']['address'].get('neighborhood', 'Não disp.')
                        v, padrao = calcular_valor_entrega(bairro)
                        tipo_entrega = str(item['delivery'].get('type', '')).upper()
                        if "PICKUP" in tipo_entrega or "RETIRADA" in tipo_entrega:
                            v = 0.0
                            motoboy_final = "RETIRADA"
                        else:
                            motoboy_final = identificar_motoboy(email)
                        itens_str = formatar_itens_para_string(item.get('items', []))
                        salvar_no_excel({
                            'numero': num, 
                            'cliente': item['customer']['name'], 
                            'bairro': bairro, 
                            'status': st, 
                            'motoboy': motoboy_final, 
                            'combo': 'NAO', 
                            'valor': v, 
                            'valor_padrao_usado': padrao, 
                            'itens': itens_str,
                            'data_pedido': item.get('date')
                        })
                        CACHE_STATUS_PEDIDOS[num] = str(st).upper()
                        total_lidos += 1
                    pagina += 1
                    erros = 0
                    esperar_humano(3, 6)  # Delay aumentado entre páginas
                else:
                    print(" (Vazia/Fim)")
                    break
            else:
                print(f" ⚠️ Erro API ({r.status_code})")
                erros += 1
                time.sleep(2)
        except Exception as e:
            print(f" ❌ Erro de conexão: {e}")
            erros += 1
            time.sleep(2)
    print(f"✅ Histórico finalizado! Total processado: {total_lidos}")

# ==================================================================================
#  SEÇÃO 8: MONITORAMENTO E SINCRONIZAÇÃO
# ==================================================================================
# Responsável por: Monitor em tempo real de pedidos, sincronização com Excel,
# tratamento de mudanças de status e envio de notificações.
# ==================================================================================

def monitorar():
    global TOKEN_ATUAL, pedidos_em_espera, CACHE_NOMES_DO_DIA
    q = """query sellerActiveOrders { activeOrders { number date status customer { name } delivery { type address { neighborhood location { coordinates } } courier { email } } items { name amount } } }"""
    try:
        r = requisicao_segura({"query": q})
        if r and r.status_code == 200:
            d = r.json()
            if d.get("data") and d["data"].get("activeOrders"): 
                lista = d["data"]["activeOrders"]
                status_rota = {
                    'DISPATCHED', 'IN_TRANSIT', 'EN_ROUTE', 'EM_ROUTE', 'DM_EN_ROUTE',
                    'DM_PICKED_UP', 'PICKED_UP', 'OUT_FOR_DELIVERY', 'ON_THE_WAY',
                    'IN_DELIVERY', 'DM_ROUTING'
                }
                ativos_status = status_rota | {'POC_ACCEPTED'}
                ativos = [p['number'] for p in lista if str(p.get('status', '')).upper() in ativos_status]
                for k in list(pedidos_em_espera):
                    if k not in ativos:
                        del pedidos_em_espera[k]
                        # Remove timestamp quando pedido sai dos ativos
                        if k in TIMESTAMP_ACEITOS:
                            del TIMESTAMP_ACEITOS[k]

                status_cancel = set(STATUS_CANCELADOS_LISTA + ["CANCELLED", "ABANDONED", "CANCELED_BY_DRIVER"])

                for p in lista:
                    num = p['number']
                    st_raw = p.get('status', '')
                    st = str(st_raw).upper()
                    nome = p['customer']['name']
                    dt_pedido = parse_data_pedido(p.get('date')) or datetime.now()
                    CACHE_NOMES_DO_DIA[num] = {
                        'nome': nome, 
                        'hora': dt_pedido,
                        'bairro': p['delivery']['address'].get('neighborhood', 'Não disp.'),
                        'itens': formatar_itens_para_string(p.get('items', []))
                    }
                    retirada = any(t in str(p['delivery'].get('type', '')).upper() for t in ["PICKUP", "RETIRADA"])
                    try: email = p['delivery']['courier']['email']
                    except: email = None
                    bairro = p['delivery']['address'].get('neighborhood', 'Não disp.')
                    v, padrao = calcular_valor_entrega(bairro)
                    itens_str = formatar_itens_para_string(p.get('items', []))
                    status_antigo = CACHE_STATUS_PEDIDOS.get(num)
                    


                    # ALERTA DE ATRASO NA RETIRADA - Sempre verifica quando sai de POC_ACCEPTED
                    if status_antigo == 'POC_ACCEPTED' and st != 'POC_ACCEPTED' and num in TIMESTAMP_ACEITOS and not retirada:
                            tempo_pedido = TIMESTAMP_ACEITOS[num]
                            tempo_decorrido = (datetime.now() - tempo_pedido).total_seconds()
                            minutos_debug = int(tempo_decorrido // 60)
                            print(f"🔍 DEBUG ALERTA: Pedido {num} saiu de POC_ACCEPTED após {minutos_debug}min para status {st}")
                            
                            # Se passou mais de 9min30s (570 segundos) ou modo teste ativo
                            if tempo_decorrido > 570 or DEBUG_ALERTA_RETIRADA_TODOS:
                                # Conta quantos motoboys cadastrados estão na rua (info adicional)
                                emails_na_rua = set()
                                for k, v in CACHE_STATUS_PEDIDOS.items():
                                    if v in status_rota:
                                        try:
                                            # Busca o email do motoboy na lista original
                                            for ped_orig in lista:
                                                if ped_orig['number'] == k:
                                                    try:
                                                        email_moto = ped_orig['delivery']['courier']['email']
                                                        if email_moto and email_moto in MOTOBOYS_API:
                                                            emails_na_rua.add(email_moto)
                                                    except:
                                                        pass
                                        except:
                                            pass
                                
                                # Conta quantos motoboys CADASTRADOS existem
                                total_motoboys = len(MOTOBOYS_API)
                                motoboys_livres = total_motoboys - len(emails_na_rua)
                                
                                # SEMPRE SALVA O ALERTA (decisão manual no painel)
                                minutos = int(tempo_decorrido // 60)
                                segundos = int(tempo_decorrido % 60)
                                tipo_alerta = "atraso" if tempo_decorrido > 570 else "teste"
                                
                                # Salva alerta na fila para confirmação no painel
                                alerta = {
                                    "numero": num,
                                    "cliente": nome.upper(),
                                    "motoboy": identificar_motoboy(email),
                                    "tempo_minutos": minutos,
                                    "tempo_segundos": segundos,
                                    "motoboys_livres": motoboys_livres,
                                    "motoboys_ocupados": len(emails_na_rua),
                                    "tipo": tipo_alerta,
                                    "status_novo": st,
                                    "timestamp": datetime.now().strftime('%H:%M:%S'),
                                    "hora_aceito": tempo_pedido.strftime('%H:%M:%S')
                                }

                                # ENVIO AUTOMÁTICO apenas se passou de 9:55 min e o modo automático estiver ativo
                                enviar_automatico = (tempo_decorrido >= 595) and ALERTA_RETIRADA_AUTO
                                
                                if enviar_automatico:
                                    # Verifica se a menção está ativa
                                    mencao_ativa = CONFIG.get("whatsapp_mencao_ativa", False)
                                    mencao_txt = "@+55 49 9172-7951 " if mencao_ativa else ""
                                    
                                    msg_alerta = (
                                        f"{mencao_txt}⚠️ RETIRADA ATRASADA: {num}\n"
                                        f"👤 *{nome.upper()}*\n"
                                        f"🏍️ Motoboy: {identificar_motoboy(email)}\n"
                                        f"⏱️ Retirou com: {minutos}min {segundos}s\n"
                                        f"🕐 Pedido: {tempo_pedido.strftime('%H:%M:%S')} | Saida: {datetime.now().strftime('%H:%M:%S')}"
                                    )
                                    try:
                                        enviar_mensagem_grupo(msg_alerta)
                                        fazer_barulho()
                                        print(f"📤 ALERTA AUTOMÁTICO ENVIADO: Pedido {num} ({minutos}min {segundos}s)")
                                    except Exception as e:
                                        print(f"❌ Erro ao enviar alerta automatico: {e}")
                                
                                try:
                                    alertas_pendentes = []
                                    if os.path.exists(ARQUIVO_ALERTAS):
                                        with open(ARQUIVO_ALERTAS, 'r', encoding='utf-8') as f:
                                            try:
                                                alertas_pendentes = json.load(f)
                                            except json.JSONDecodeError:
                                                print("⚠️ alertas_atraso.json corrompido, resetando...")
                                                alertas_pendentes = []
                                    
                                    alertas_pendentes.append(alerta)
                                    
                                    with open(ARQUIVO_ALERTAS, 'w', encoding='utf-8') as f:
                                        json.dump(alertas_pendentes, f, indent=2, ensure_ascii=False)
                                    
                                    print(f"🚨 ALERTA GERADO E SALVO: Pedido {num} - {nome.upper()} - {identificar_motoboy(email)} ({minutos}min {segundos}s)")
                                    print(f"⚠️ Alerta pendente: {num} ({minutos}min {segundos}s | {motoboys_livres} livres / {len(emails_na_rua)} ocupados)")
                                except Exception as e:
                                    print(f"❌ ERRO ao salvar alerta: {e}")
                                    import traceback
                                    traceback.print_exc()
                            
                            # Remove do tracking após sair da loja
                            del TIMESTAMP_ACEITOS[num]

                    # 3. VERIFICA PEDIDOS EM ROTA (Ativos)
                    if st in status_rota:
                        if num not in pedidos_em_espera:
                            pedidos_em_espera[num] = {'cliente': nome, 'eh_retirada': retirada, 'bairro': bairro}
                        
                        moto_final = "RETIRADA" if retirada else identificar_motoboy(email)
                        val_final = 0.0 if retirada else v
                        salvar_no_excel({'numero': num, 'cliente': nome, 'bairro': bairro, 'status': st_raw, 'motoboy': moto_final, 'combo': 'NAO', 'valor': val_final, 'valor_padrao_usado': padrao, 'itens': itens_str, 'data_pedido': p.get('date')})
                        continue # Pula para o próximo pedido do loop
                    
                    # 4. VERIFICA NOVOS PEDIDOS (Aceitos na Loja)
                    if st == 'POC_ACCEPTED':
                        # Guarda timestamp da hora ORIGINAL do pedido (não do aceite)
                        if num not in TIMESTAMP_ACEITOS:
                            TIMESTAMP_ACEITOS[num] = dt_pedido
                        if CACHE_STATUS_PEDIDOS.get(num) != 'POC_ACCEPTED':
                            CACHE_STATUS_PEDIDOS[num] = 'POC_ACCEPTED'
                        
                        if num not in pedidos_em_espera:
                            pedidos_em_espera[num] = {'notificado': False, 'entrega_alert_sent': False}
                            pedidos_em_espera[num].update({'cliente': nome, 'bairro': bairro, 'eh_retirada': retirada})
                        
                        if not pedidos_em_espera[num]['notificado']:
                            print(f"🔔 Novo Pedido Detectado: {num}")
                            
                            if retirada:
                                msg = (f"RETIRADA: {num}\n👤 *{nome.upper()}*\n📞 {buscar_telefone(num)}")
                                enviar_mensagem_grupo(msg); fazer_barulho()
                                salvar_no_excel({'numero': num, 'cliente': nome, 'bairro': bairro, 'status': st_raw, 'motoboy': 'RETIRADA', 'combo': 'NAO', 'valor': 0, 'valor_padrao_usado': False, 'itens': itens_str, 'data_pedido': p.get('date')})
                            else:
                                deadline = dt_pedido + timedelta(minutes=7, seconds=30)
                                cutoff = dt_pedido + timedelta(minutes=10) 
                                # Se o cutoff já passou (pedido antigo detectado após reinício), marca como já alertado
                                if datetime.now() > cutoff:
                                    pedidos_em_espera[num].update({'entrega_deadline': deadline, 'entrega_cutoff': cutoff, 'entrega_alert_sent': True})
                                else:
                                    pedidos_em_espera[num].update({'entrega_deadline': deadline, 'entrega_cutoff': cutoff})
                                msg = (f"NOVO PEDIDO: {num}\n👤 *{nome.upper()}*\n📍 {bairro}\n💰 R$ {v:.2f}".replace('.', ',') + "\n" + f"📞 {buscar_telefone(num)}")
                                
                                enviar_mensagem_grupo(msg)
                                fazer_barulho()
                                # Anunciar bairro na voz para entregas
                                if bairro:
                                    # NÃO salva no Excel em POC_ACCEPTED - só salva quando for atribuído motoboy (DISPATCHED)
                                    pass
                            
                            pedidos_em_espera[num]['notificado'] = True
                
                # Checagem de atraso
                agora_check = datetime.now()
                for pid, pdados in list(pedidos_em_espera.items()):
                    if pdados.get('eh_retirada') or pdados.get('entrega_alert_sent'):
                        continue
                    # Só envia alerta se o pedido AINDA está em POC_ACCEPTED (não foi retirado)
                    if pid not in TIMESTAMP_ACEITOS:
                        continue
                    tempo_pedido = TIMESTAMP_ACEITOS[pid]
                    tempo_decorrido = (agora_check - tempo_pedido).total_seconds()
                    # Alerta de voz aos 7,5 minutos (450s), mas fala "nove minutos"
                    if 450 < tempo_decorrido < 470 and not pdados.get('voz_alerta_7min', False):
                        nome = pdados.get('cliente', 'Cliente')
                        bairro = pdados.get('bairro', '')
                        pedidos_em_espera[pid]['voz_alerta_7min'] = True
                    cutoff = pdados.get('entrega_cutoff')
                    if cutoff and agora_check > cutoff:
                        msg_alert = (f"ENTREGA ATRASADA: {pid}\n👤 *{pdados.get('cliente', 'Cliente').upper()}*\n⏰ Ultrapassou 11 minutos sem retirada!")
                        enviar_mensagem_grupo(msg_alert)
                        fazer_barulho()
                        pedidos_em_espera[pid]['entrega_alert_sent'] = True
    except Exception as e: print(f"⚠️ Erro monitoramento: {e}")
    
# ================= INTEGRAÇÃO COM PAINEL (COM DEBUG & INTERCEPTOR) =================
ARQUIVO_COMANDO = 'comando_imprimir.txt'

# --- VARIÁVEL DE CONTROLE DO TELEGRAM (ADICIONE ISSO AQUI OU NO INÍCIO) ---
LAST_UPDATE_ID = 0

def imprimir_extrato_por_nome(nome_alvo, data_str, hora_ini=None, hora_fim=None, valor_gar=None):
    print(f"\n🖨️ COMANDO RECEBIDO: Buscar '{nome_alvo}' na data {data_str} | Garantia: ini={hora_ini} fim={hora_fim} gar={valor_gar}")
    caminho_base = get_caminho_base()
    arquivo_excel = os.path.join(caminho_base, f'Controle_Financeiro_{data_str}.xlsx')

    print(f"📂 Tentando abrir planilha: {arquivo_excel}")
    if not os.path.exists(arquivo_excel):
        print(f"❌ ARQUIVO NÃO ENCONTRADO! Verifique se a data está certa.")
        fazer_barulho()
        return False

    wb = openpyxl.load_workbook(arquivo_excel, data_only=True)
    ws = wb["EXTRATO DETALHADO"]
    pedidos_completos = []
    qtd_8 = 0
    qtd_11 = 0
    total = 0.0

    print("🔎 --- INICIANDO VARREDURA NO EXCEL ---")
    nome_buscado_norm = normalizar_texto(nome_alvo)

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[6]:
            continue
        motoboy_excel_original = str(row[6])
        motoboy_excel_norm = normalizar_texto(motoboy_excel_original)
        status = str(row[5]).upper()
        val = float(row[8]) if row[8] else 0.0

        eh_retirada_busca = "retirada" in nome_buscado_norm
        match = False

        if eh_retirada_busca:
            if (val == 0.0 or "RETIRADA" in motoboy_excel_original.upper()) and not any(x in status for x in STATUS_CANCELADOS_LISTA):
                match = True
        elif nome_buscado_norm in motoboy_excel_norm:
            if not any(x in status for x in STATUS_CANCELADOS_LISTA):
                match = True

        if match:
            total += val
            if abs(val - 8.0) < 0.1:
                qtd_8 += 1
            elif abs(val - 11.0) < 0.1:
                qtd_11 += 1

            data_ped = row[0]
            if isinstance(data_ped, datetime):
                data_ped = data_ped.strftime('%d/%m')

            pedidos_completos.append({
                'numero': row[2],
                'data': str(data_ped),
                'hora': str(row[1]),
                'cliente': str(row[3]),
                'bairro': str(row[4]),
                'motoboy': str(row[6]),
                'valor': val,
                'itens': str(row[9]) if len(row) > 9 and row[9] else ""
            })

    vale_total = 0.0
    if "retirada" not in nome_buscado_norm and "VALES" in wb.sheetnames:
        ws_vales = wb["VALES"]
        for row in ws_vales.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 3 or not row[1]:
                continue
            if normalizar_texto(str(row[1])) == nome_buscado_norm:
                try:
                    vale_total += float(row[2]) if row[2] else 0.0
                except Exception:
                    pass

    print(f"🏁 Fim da varredura. Total encontrados: {len(pedidos_completos)}")
    if pedidos_completos:
        print("🖨️ Imprimindo detalhes (Lote)...")
        imprimir_lote_continuo(pedidos_completos)
        time.sleep(2)
        print("🖨️ Imprimindo resumo final...")
        nome_final = "RETIRADAS" if "retirada" in nome_buscado_norm else nome_alvo

        # ===== LÓGICA DE FECHAMENTO: produção vs garantia =====
        # Usa os dados passados pelo painel (hora_ini, hora_fim, valor_gar)
        valor_garantido = None
        if valor_gar:
            try:
                valor_garantido = float(str(valor_gar).replace(",", "."))
            except Exception:
                valor_garantido = None

        def parse_hora(txt):
            if txt is None:
                return None
            txt = str(txt).strip()
            # Tenta HH:MM
            try:
                return datetime.strptime(txt, "%H:%M")
            except Exception:
                pass
            # Tenta HH:MM:SS
            try:
                return datetime.strptime(txt, "%H:%M:%S")
            except Exception:
                pass
            # Tenta datetime completo e extrai só a hora
            try:
                dt = datetime.strptime(txt, "%Y-%m-%d %H:%M:%S")
                return dt.replace(year=1900, month=1, day=1)
            except Exception:
                pass
            # Tenta extrair HH:MM de qualquer string
            import re as _re
            m = _re.search(r'(\d{1,2}):(\d{2})', txt)
            if m:
                try:
                    return datetime.strptime(f"{m.group(1)}:{m.group(2)}", "%H:%M")
                except Exception:
                    pass
            return None

        t_in = parse_hora(hora_ini) if hora_ini else None
        t_out = parse_hora(hora_fim) if hora_fim else None

        # Se não houver horários, imprime tudo como antes
        if not t_in or not t_out or t_out <= t_in:
            total_liquido = total - vale_total
            if total_liquido < 0:
                total_liquido = 0.0
            imprimir_resumo_extrato(
                nome_final, pedidos_completos, qtd_8, qtd_11, total_liquido,
                data_personalizada=data_str, vale_total=vale_total
            )
            print("✅ Impressão completa!")
            return True

        # Se houver horários, aplica a lógica do fechamento
        prod_dentro = 0.0
        prod_fora = 0.0  # Só entregas APÓS hora_fim
        qtd8_dentro = 0
        qtd11_dentro = 0
        qtd8_fora = 0
        qtd11_fora = 0
        for ped in pedidos_completos:
            h = ped.get('hora')
            t_ped = parse_hora(h)
            val = float(ped.get('valor', 0.0) or 0.0)
            if t_ped and t_in <= t_ped <= t_out:
                # Entrega dentro do período garantido
                prod_dentro += val
                if abs(val - 8.0) < 0.1:
                    qtd8_dentro += 1
                elif abs(val - 11.0) < 0.1:
                    qtd11_dentro += 1
                print(f"   ✅ DENTRO: Ped {ped.get('numero')} hora={h} parsed={t_ped.strftime('%H:%M')} val={val}")
            elif t_ped and t_ped > t_out:
                # Entrega APÓS o horário garantido = extra
                prod_fora += val
                if abs(val - 8.0) < 0.1:
                    qtd8_fora += 1
                elif abs(val - 11.0) < 0.1:
                    qtd11_fora += 1
                print(f"   ➕ EXTRA: Ped {ped.get('numero')} hora={h} parsed={t_ped.strftime('%H:%M')} val={val}")
            elif t_ped and t_ped < t_in:
                print(f"   ⏭️ ANTES: Ped {ped.get('numero')} hora={h} parsed={t_ped.strftime('%H:%M')} val={val} (ignorada)")
            else:
                print(f"   ❓ SEM HORA: Ped {ped.get('numero')} hora={h} parsed=None val={val} (ignorada)")
            # Entregas ANTES do horário inicial são ignoradas no cálculo

        # Calcula garantia (manual ou automático)
        if valor_garantido is not None:
            gar_valor = valor_garantido
        else:
            horas = (t_out - t_in).total_seconds() / 3600
            gar_valor = max(horas * 15.0, 0.0)

        usar_garantia = gar_valor > prod_dentro
        if usar_garantia:
            total_final = gar_valor + prod_fora
        else:
            total_final = prod_dentro + prod_fora

        total_final -= vale_total
        if total_final < 0:
            total_final = 0.0

        # Debug detalhado
        print(f"📊 DETALHAMENTO FECHAMENTO {nome_final}:")
        print(f"   Horário garantido: {hora_ini} - {hora_fim}")
        print(f"   Produção DENTRO ({hora_ini}-{hora_fim}): R$ {prod_dentro:.2f} ({qtd8_dentro}x R$8 + {qtd11_dentro}x R$11)")
        print(f"   Produção FORA (extras): R$ {prod_fora:.2f} ({qtd8_fora}x R$8 + {qtd11_fora}x R$11)")
        print(f"   Garantia: R$ {gar_valor:.2f} | Usar garantia: {usar_garantia}")
        print(f"   Vale: R$ {vale_total:.2f}")
        print(f"   TOTAL FINAL: R$ {total_final:.2f}")

        # Dados de fechamento para o recibo detalhado
        dados_fechamento = {
            "hora_ini": hora_ini,
            "hora_fim": hora_fim,
            "prod_dentro": prod_dentro,
            "prod_fora": prod_fora,
            "gar_valor": gar_valor,
            "usar_garantia": usar_garantia,
            "qtd8_dentro": qtd8_dentro,
            "qtd11_dentro": qtd11_dentro,
            "qtd8_fora": qtd8_fora,
            "qtd11_fora": qtd11_fora,
        }

        # Imprime o resumo com o valor correto e detalhamento
        imprimir_resumo_extrato(
            nome_final, pedidos_completos, qtd8_dentro + qtd8_fora, qtd11_dentro + qtd11_fora, total_final,
            data_personalizada=data_str, vale_total=vale_total, dados_fechamento=dados_fechamento
        )
        print("✅ Impressão completa!")
        return True

    print("⚠️ NADA ENCONTRADO.")
    return False

def processar_relatorio_retiradas(data_str):
    """Busca e imprime todas as retiradas de uma data específica."""
    print(f"\n🖨️ COMANDO RECEBIDO: Imprimir todas as retiradas da data {data_str}")
    arquivo_excel = get_caminho_excel_por_data(data_str)

    if not os.path.exists(arquivo_excel):
        print(f"❌ ARQUIVO NÃO ENCONTRADO! Verifique se a data {data_str} está certa.")
        return False

    wb = openpyxl.load_workbook(arquivo_excel, data_only=True)
    ws = wb["EXTRATO DETALHADO"]
    pedidos_retirada = []
    total_valor_retirada = 0.0

    print("🔎 --- INICIANDO VARREDURA DE RETIRADAS NO EXCEL ---")

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[6]:
            continue
        
        motoboy_excel_original = str(row[6])
        status = str(row[5]).upper()
        val = float(row[8]) if row[8] else 0.0

        # Condição: Motoboy é 'RETIRADA' ou valor é 0, e não está cancelado.
        if ("RETIRADA" in motoboy_excel_original.upper() or val == 0.0) and not any(x in status for x in STATUS_CANCELADOS_LISTA):
            total_valor_retirada += val 
            pedidos_retirada.append({
                'numero': str(row[2]),
                'data': str(row[0]),
                'hora': str(row[1]),
                'cliente': str(row[3]),
                'bairro': str(row[4]),
                'motoboy': str(row[6]),
                'valor': val,
                'itens': str(row[9]) if len(row) > 9 and row[9] else ""
            })

    print(f"🏁 Fim da varredura. Total de retiradas encontradas: {len(pedidos_retirada)}")
    
    if pedidos_retirada:
        print("🖨️ Imprimindo tickets de retirada...")
        imprimir_lote_continuo(pedidos_retirada)
        time.sleep(2)
        print("🖨️ Imprimindo resumo final de retiradas...")
        # Usamos a função de extrato, mas com dados zerados de entrega e total.
        imprimir_resumo_extrato(
            "TODAS RETIRADAS", pedidos_retirada, 0, 0, total_valor_retirada,
            data_personalizada=data_str, vale_total=0.0
        )
        print("✅ Impressão de retiradas completa!")
        return True

    print("⚠️ Nenhuma retirada encontrada para a data.")
    return False

def processar_comando_painel():
    if not os.path.exists(ARQUIVO_COMANDO): return

    # Renomeia atomicamente para evitar condição de corrida
    arquivo_temp = ARQUIVO_COMANDO + '.processing'
    try:
        os.rename(ARQUIVO_COMANDO, arquivo_temp)
    except FileNotFoundError:
        return  # Outro processo já pegou o comando
    except PermissionError:
        return  # Arquivo em uso

    try:
        with open(arquivo_temp, 'r', encoding='utf-8') as f:
            cmd = f.read().strip()
        os.remove(arquivo_temp)

        if cmd == "RECARREGAR_CONFIG":
            global CONFIG, ENDERECO_LOJA, NOME_GRUPO_FIXO, MOTOBOYS_API
            nova_config = carregar_configuracoes()
            if nova_config:
                CONFIG = nova_config
                ENDERECO_LOJA = CONFIG.get('endereco_loja', ENDERECO_LOJA)
                NOME_GRUPO_FIXO = CONFIG.get('nome_grupo', NOME_GRUPO_FIXO)
                MOTOBOYS_API = CONFIG.get('motoboys', MOTOBOYS_API)
                print("🔄 Configurações recarregadas do painel.")
            return

        if cmd == "VERIFICAR_HISTORICO":
            print("📜 COMANDO DO PAINEL: Verificar histórico do dia")
            buscar_historico_do_dia(limite_paginas=None)
            return


        if cmd == "FECHAMENTO_MANUAL":
            print("💰 COMANDO DO PAINEL: Fechamento manual")
            executar_fechamento_manual()
            return

        if cmd.startswith("IMPRIMIR_FECHAMENTO:"):
            payload = cmd.split(":", 1)[1]
            if "|" in payload:
                data_str, dados_str = payload.split("|", 1)
            else:
                data_str = datetime.now().strftime('%d-%m-%Y')
                dados_str = payload

            partes = dados_str.split("|")
            nome_alvo = partes[0] if partes else ""
            if nome_alvo:
                imprimir_extrato_por_nome(nome_alvo, data_str)
            print("\n🖨️ COMANDO RECEBIDO: Recibo de Garantia")
            imprimir_recibo_garantia(dados_str)

        elif cmd.startswith("IMPRIMIR:") or cmd.startswith("IMPRIMIR_COMANDA:"):
            if cmd.startswith("IMPRIMIR_COMANDA:"): dados_brutos = cmd.split(":", 1)[1]
            else: dados_brutos = cmd.split(":", 1)[1]
            
            partes_cmd = dados_brutos.split("|")
            nome_alvo = partes_cmd[0].strip()
            data_str = partes_cmd[1].strip() if len(partes_cmd) > 1 and partes_cmd[1].strip() else datetime.now().strftime('%d-%m-%Y')
            hora_ini = partes_cmd[2].strip() if len(partes_cmd) > 2 and partes_cmd[2].strip() else None
            hora_fim = partes_cmd[3].strip() if len(partes_cmd) > 3 and partes_cmd[3].strip() else None
            valor_gar = partes_cmd[4].strip() if len(partes_cmd) > 4 and partes_cmd[4].strip() else None
            
            termos_cancel = ["cancelada", "canceladas", "cancelamento", "abandoned"]
            if normalizar_texto(nome_alvo) in termos_cancel:
                print(f"\n🖨️ COMANDO ESPECIAL: Imprimir Canceladas ({data_str})")
                res = processar_relatorio_canceladas(data_str)
                print(res)
                return

            imprimir_extrato_por_nome(nome_alvo, data_str, hora_ini=hora_ini, hora_fim=hora_fim, valor_gar=valor_gar)

        elif cmd.startswith("IMPRIMIR_GARANTIA:"):
            dados_brutos = cmd.split(":", 1)[1]
            print(f"\n🖨️ COMANDO RECEBIDO: Recibo de Garantia")
            imprimir_recibo_garantia(dados_brutos)

        elif cmd.startswith("IMPRIMIR_CANCELADAS"):
            parts = cmd.split(":")
            data_cancel = parts[1] if len(parts) > 1 else None
            print(f"🖨️ COMANDO DO PAINEL: Canceladas (Data: {data_cancel if data_cancel else 'Hoje'})")
            res = processar_relatorio_canceladas(data_cancel)
            print(res)

        elif cmd.startswith("IMPRIMIR_RETIRADAS"):
            parts = cmd.split(":")
            data_retirada = parts[1] if len(parts) > 1 else None
            print(f"🖨️ COMANDO DO PAINEL: Retiradas (Data: {data_retirada if data_retirada else 'Hoje'})")
            processar_relatorio_retiradas(data_retirada)

        elif cmd.startswith("ENVIAR_WHATSAPP:"):
            mensagem = cmd.split(":", 1)[1]
            print(f"📤 COMANDO DO PAINEL: Enviar alerta no WhatsApp")
            enviar_mensagem_grupo(mensagem)
            fazer_barulho()
            print(f"✅ Alerta enviado: {mensagem[:50]}...")

    except Exception as e:
        print(f"❌ Erro ao processar comando do painel: {e}")
        # Limpa arquivo temporário se ainda existir
        if os.path.exists(arquivo_temp):
            try:
                os.remove(arquivo_temp)
            except:
                pass

ARQUIVO_COMANDO_TELEGRAM = 'telegram_command.txt'

def processar_comando_telegram():
    if not os.path.exists(ARQUIVO_COMANDO_TELEGRAM):
        return

    try:
        with open(ARQUIVO_COMANDO_TELEGRAM, 'r', encoding='utf-8') as f:
            cmd = f.read().strip()
        os.remove(ARQUIVO_COMANDO_TELEGRAM)

        print(f"🤖 Recebido comando do Telegram: {cmd}")


        if cmd == 'PARAR':
            enviar_telegram("🛑 Robô será parado por comando do Telegram.")
            print("[Telegram] Comando PARAR recebido. Encerrando...")
            os._exit(0)

        elif cmd == 'REINICIAR':
            enviar_telegram("🔄 Reiniciando o robô por comando do Telegram...")
            print("[Telegram] Comando REINICIAR recebido. Reiniciando...")
            os.execv(sys.executable, ['python'] + sys.argv)


        # INTEGRAÇÃO: Aqui são processados todos os comandos recebidos do Telegram (telegram_command.txt)
        # e do painel (comando_imprimir.txt). Mantenha este bloco sincronizado com os comandos disponíveis
        # em telegram_bot.py e painel.py para garantir operação remota consistente.
        if cmd == 'PARAR':
            enviar_telegram("🛑 Robô será parado por comando do Telegram.")
            print("[Telegram] Comando PARAR recebido. Encerrando...")
            os._exit(0)

        elif cmd == 'REINICIAR':
            enviar_telegram("🔄 Reiniciando o robô por comando do Telegram...")
            print("[Telegram] Comando REINICIAR recebido. Reiniciando...")
            os.execv(sys.executable, ['python'] + sys.argv)

        elif cmd == 'GERAR_EXCEL':
            enviar_telegram("📜 Ok, buscando o histórico completo do dia para gerar o Excel. Isso pode levar um momento...")
            buscar_historico_do_dia(limite_paginas=None)
            enviar_telegram("✅ Histórico verificado e Excel atualizado.")

        elif cmd == 'GERAR_RESUMO':
            enviar_telegram("📊 *Gerando resumo detalhado, aguarde...*")
            msg_resumo = gerar_relatorio_executivo()
            enviar_telegram(msg_resumo)

        elif cmd == 'GERAR_CANCELADAS':
            enviar_telegram("🖨️ *Gerando Relatório de Canceladas...*")
            res = processar_relatorio_canceladas()
            enviar_telegram(f"✅ {res}")


        elif cmd == 'VER_MOTOS':
            enviar_telegram("🔎 *Verificando entregadores na rua...*")
            try:
                pedidos = consultar_api_direta()
                status_rota = ['DISPATCHED', 'IN_TRANSIT', 'EN_ROUTE', 'EM_ROUTE', 'DM_EN_ROUTE', 'DM_PICKED_UP', 'PICKED_UP', 'OUT_FOR_DELIVERY', 'ON_THE_WAY', 'IN_DELIVERY', 'DM_ROUTING']
                
                motos_na_rua = {}
                for p in pedidos:
                    st = str(p.get('status', '')).upper()
                    if st in status_rota:
                        pid = str(p['number'])
                        try: email = p['delivery']['courier']['email']
                        except: email = None
                        
                        nome_moto = identificar_motoboy(email) if email else "Desconhecido"
                        
                        # Agrupa os pedidos por motoboy
                        if nome_moto not in motos_na_rua:
                            motos_na_rua[nome_moto] = []
                        motos_na_rua[nome_moto].append(pid)
                
                if not motos_na_rua:
                    enviar_telegram("🛵 *Nenhum motoboy na rua neste momento.*")
                else:
                    msg = "🛵 *MOTOBOYS EM ROTA:*\n\n"
                    for moto, pids in motos_na_rua.items():
                        msg += f"👤 *{moto.upper()}* - {len(pids)} entrega(s)\n"
                        msg += f"📦 Pedidos: {', '.join(pids)}\n\n"
                    enviar_telegram(msg)
            except Exception as e:
                enviar_telegram(f"❌ Erro ao buscar motoboys na rua: {e}")

       
        elif cmd == 'RECARREGAR_CONFIG':
            carregar_credenciais()
            carregar_motoboys_do_painel()
            enviar_telegram("🔄 Configurações recarregadas.")

        elif cmd.startswith('ENVIAR_MENSAGEM:'):
            mensagem = cmd.split(':', 1)[1]
            enviar_mensagem_grupo(mensagem)
            enviar_telegram("Mensagem enviada ao grupo WhatsApp.")

        elif cmd.startswith('IMPRIMIR_PEDIDO:'):
            nome = cmd.split(':', 1)[1]
            res = processar_impressao_individual(nome)
            enviar_telegram(res)

        # --- NOVO BLOCO: ENSINANDO O TELEGRAM A IMPRIMIR O FECHAMENTO ---
        elif cmd.startswith('IMPRIMIR:'):
            dados_brutos = cmd.split(':', 1)[1]
            
            # Separa o nome e a data cortando no "|"
            if "|" in dados_brutos:
                nome_alvo, data_str = dados_brutos.split("|")
            else:
                nome_alvo = dados_brutos
                data_str = datetime.now().strftime('%d-%m-%Y')
            
            enviar_telegram(f"🖨️ *Buscando fechamento...*\nMotoboy: {nome_alvo.upper()}\nData: {data_str}")
            
            # Chama a SUA função original que já faz todo o cálculo e impressão do Excel
            sucesso = imprimir_extrato_por_nome(nome_alvo, data_str)
            
            if sucesso:
                enviar_telegram(f"✅ Extrato impresso na impressora térmica com sucesso!")
            else:
                enviar_telegram(f"⚠️ *Nenhum pedido encontrado*\nNão achei corridas para '{nome_alvo}' na planilha do dia {data_str}.")
        
        # --- BLOCO CORRIGIDO: IMPRESSÃO DE GARANTIA ---
        elif cmd.startswith('GERAR_GARANTIA:') or cmd.startswith('IMPRIMIR_GARANTIA:'):
            dados = cmd.split(':', 1)[1]
            partes = dados.split('|')
            
            # Se vier do Telegram (/garantia), virão só 3 partes (Nome, Início, Fim)
            if len(partes) == 3:
                nome_alvo, hora_ini, hora_fim = partes
                data_str = datetime.now().strftime('%d-%m-%Y')
                
                enviar_telegram(f"🖨️ *Calculando garantia...*\nMotoboy: {nome_alvo.upper()}\nTurno: {hora_ini} às {hora_fim}")
                
                # Usamos a mesma função do painel que calcula tudo automaticamente!
                sucesso = imprimir_extrato_por_nome(nome_alvo, data_str, hora_ini=hora_ini, hora_fim=hora_fim)
                
                if sucesso:
                    enviar_telegram("✅ Extrato com garantia impresso com sucesso!")
                else:
                    enviar_telegram(f"⚠️ Não achei corridas para '{nome_alvo}' hoje.")
            
            # Mantém a estrutura original caso o painel envie o formato longo
            elif len(partes) >= 8:
                print(f"\n🖨️ COMANDO RECEBIDO: Recibo de Garantia (Painel)")
                imprimir_recibo_garantia(dados)
                enviar_telegram("✅ Recibo de garantia impresso com sucesso!")
            else:
                enviar_telegram("⚠️ Erro: Dados de garantia incompletos.")

        elif cmd == 'TOGGLE_ALERTA_AUTO':
            novo = not CONFIG.get('alerta_retirada_auto', False)
            atualizar_config_flag('alerta_retirada_auto', novo)
            enviar_telegram(f"Alerta automático {'ativada' if novo else 'desativada'}.")

        elif cmd == 'TOGGLE_MENCAO':
            novo = not CONFIG.get('whatsapp_mencao_ativa', False)
            atualizar_config_flag('whatsapp_mencao_ativa', novo)
            enviar_telegram(f"Menção automática {'ativada' if novo else 'desativada'}.")

    except Exception as e:
        print(f"❌ Erro ao processar comando do Telegram: {e}")
        enviar_telegram(f"❌ Erro ao executar comando: {e}")

# ================= LÓGICA DE COMANDOS DO TELEGRAM =================
# ==================================================================================
#  SEÇÃO 4: TELEGRAM BOT - PROCESSAMENTO DE COMANDOS
# ==================================================================================
# Responsável por: Interpretar comandos do Telegram (/imprimir, /resumo, /motos)
# e executar ações correspondentes, retornando resultados ao usuário.
# ==================================================================================

# ==================================================================================
#  SEÇÃO 14: ROTINA DE FECHAMENTO AUTOMÁTICO
# ==================================================================================
# Responsável por: Verificar e disparar relatório automático no horário
# agendado (geralmente 22:00 ou fim de expediente).
# ==================================================================================

def verificar_rotina_fechamento():
    global RELATORIO_ENVIADO_HOJE

    agora = datetime.now()
    dia_sem = agora.weekday() # 0=Seg, 1=Ter, ..., 5=Sab, 6=Dom
    hora = agora.hour
    minuto = agora.minute

    # 1. Reset da flag (as 06:00 da manhã reinicia o sistema para o próximo dia)
    if hora == 6 and RELATORIO_ENVIADO_HOJE:
        RELATORIO_ENVIADO_HOJE = False
        print("🌅 Sistema resetado para novo dia de trabalho.")
        return

    # Se já enviou hoje, não faz nada
    if RELATORIO_ENVIADO_HOJE: return

    # 2. Definição dos Horários de Disparo (5 min após o fechamento)
    # Segunda (0): Fecha 23:00 -> Relatório 23:05
    # Terça(1) a Sábado(5): Fecha 02:00 -> Relatório 02:05 (do dia seguinte)
    # Domingo (6): Fecha 00:00 -> Relatório 00:05 (do dia seguinte/Segunda)

    hora_disparo = False

    # Regra da SEGUNDA (Fecha 23h)
    if dia_sem == 0 and hora == 23 and 5 <= minuto <= 15:
        hora_disparo = True

    # Regra de TERÇA a SÁBADO (Fecha 02h da manhã seguinte)
    # Se agora é Quarta(2) até Domingo(6) e são 02:05h, refere-se ao turno anterior
    elif dia_sem in [2, 3, 4, 5, 6] and hora == 2 and 5 <= minuto <= 15:
        hora_disparo = True

    # Regra de DOMINGO (Fecha 00h)
    # Se agora é Segunda(0) e são 00:05h, refere-se ao domingo
    elif dia_sem == 0 and hora == 0 and 5 <= minuto <= 15:
        hora_disparo = True

    # 3. Execução
    if hora_disparo:
        print(f"⏰ Hora do Fechamento ({hora}:{minuto})! Gerando relatório...")
        enviar_telegram("🌙 *LOJA FECHADA! INICIANDO FECHAMENTO AUTOMÁTICO...*")
        
        # Gera o relatório
        msg_relatorio = gerar_relatorio_executivo()
        enviar_telegram(msg_relatorio)
        
        # Envia também as canceladas por garantia
        enviar_telegram(processar_relatorio_canceladas())
        
        RELATORIO_ENVIADO_HOJE = True
        print("✅ Relatório Automático Enviado.")

def executar_fechamento_manual():
    try:
        enviar_telegram("🧾 *FECHAMENTO MANUAL INICIADO...*")
        msg_relatorio = gerar_relatorio_executivo()
        enviar_telegram(msg_relatorio)
        enviar_telegram(processar_relatorio_canceladas())
        try:
            with open(ARQUIVO_FECHAMENTO_STATUS, 'w', encoding='utf-8') as f:
                json.dump({"status": "ok", "ts": datetime.now().isoformat()}, f)
        except Exception:
            pass
        print("✅ Fechamento manual enviado.")
    except Exception as e:
        print(f"❌ Erro no fechamento manual: {e}")
        try:
            with open(ARQUIVO_FECHAMENTO_STATUS, 'w', encoding='utf-8') as f:
                json.dump({"status": "erro", "ts": datetime.now().isoformat(), "msg": str(e)}, f)
        except Exception:
            pass

# ==================================================================================
# ==================================================================================

# ==================================================================================
#  SEÇÃO 15: INICIALIZAÇÃO DO ROBÔ (MAIN LOOP)
# ==================================================================================
# Responsável por: Ponto de entrada principal e loop infinito que coordena
# todas as variações do robô (API, WhatsApp, Telegram, Excel, etc).
# ==================================================================================

def start():
    # Inicialização do robô
    try:
        print("\n🚀 INICIANDO ZÉ-BOT TURBO (MODO STEALTH v2 - CHROME)")
        carregar_credenciais()
        carregar_motoboys_do_painel()
        inicializar_excel_agora()
        preparar_gps_loja()
        iniciar_chrome_persistente()
        print("📜 Sincronizando histórico inicial completo (aguarde)...")
        buscar_historico_do_dia(limite_paginas=None)
        print("✅ Sincronização concluída! Iniciando monitoramento...")
        enviar_telegram("🚀 *ROBÔ INICIADO COM SUCESSO!*\nDigite /ajuda para ver os comandos.")
    except Exception as e:
        print(f"Erro na inicialização: {e}")
        print("Tentando novamente em 5 segundos...")
        time.sleep(5)
        reiniciar_robo()
        return

    # Controle de re-sync do histórico às 2h
    _ULTIMO_RESYNC_2H = None

    # Loop de monitoramento
    while True:
        try:
            # Re-verifica histórico completo às 2:00h
            agora_loop = datetime.now()
            if agora_loop.hour == 2 and _ULTIMO_RESYNC_2H != agora_loop.date():
                print("\n🔄 2:00h - Re-sincronizando histórico completo do dia...")
                try:
                    buscar_historico_do_dia(limite_paginas=None)
                    _ULTIMO_RESYNC_2H = agora_loop.date()
                    print("✅ Re-sincronização das 2h concluída!")
                except Exception as e:
                    print(f"⚠️ Erro na re-sincronização das 2h: {e}")

            # Rotina de fechamento automático
            verificar_rotina_fechamento()

            print(".", end="", flush=True)
            if not _driver_ativo():
                _reiniciar_chrome_se_preciso("healthcheck")
                if not _driver_ativo():
                    # Chrome ainda morto (cooldown ou falha) - pula funções que precisam dele
                    processar_comando_painel()
                    processar_comando_telegram()
                    esperar_humano(2, 4)
                    continue
            refresh_whatsapp_periodically()
            refresh_main_periodically()
            garantir_foco_no_grupo()
            processar_comando_painel()      # Comandos da Interface
            processar_comando_telegram()    # Comandos do Bot Telegram
            monitorar()                     # API
            verificar_solicitacoes_whatsapp() # Ler Grupo Zap
            # Rechecagem rápida para reduzir latência de comandos do painel
            if os.path.exists(ARQUIVO_COMANDO):
                processar_comando_painel()
            esperar_humano(0.5, 1.5)  # Reduzido para responder mais rápido no WhatsApp
        except KeyboardInterrupt:
            print("\n🛑 Parando...")
            break
        except Exception as e:
            print(f"💥 ERRO FATAL NO LOOP: {e}")
            enviar_telegram(f"⚠️ Erro fatal detectado. Reiniciando robô completamente: {e}")
            time.sleep(5)
            reiniciar_robo()


if __name__ == "__main__":
    import sys
    manter_ativo = False
    try:
        # Inicializações extras podem ser adicionadas aqui
        pass
    except Exception as e:
        print(f"Erro na inicialização: {e}")

    if manter_ativo:
        while True:
            try:
                start()
            except Exception as e:
                print(f"Erro crítico no robô: {e}")
                print("Reiniciando robô em 5 segundos...")
                time.sleep(5)
    else:
        start()